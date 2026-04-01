# Professor_Student_Manage/admin.py
from django.contrib import admin
from django import forms
from django.utils.html import format_html
import requests
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import render, redirect
from django.urls import path
import csv
from io import TextIOWrapper
from django.contrib.auth.models import User
from django.db.models import Sum, F
from collections import defaultdict

# Register your models here.
from Select_Information.models import StudentProfessorChoice
from .models import Student, Professor, WeChatAccount, ProfessorDoctorQuota, ProfessorMasterQuota
from Enrollment_Manage.models import Subject
from django.http import JsonResponse, HttpResponse
import zipfile
import tempfile
import os
from django.core.cache import cache


WECHAT_CLOUD_ENV = os.environ.get('WECHAT_CLOUD_ENV', 'prod-2g1jrmkk21c1d283')
WECHAT_APPID = os.environ.get('WECHAT_APPID', 'wxa67ae78c4f1f6275')
WECHAT_SECRET = os.environ.get('WECHAT_SECRET', '7241b1950145a193f15b3584d50f3989')
WECHAT_API_BASE = os.environ.get('WECHAT_API_BASE', 'https://api.weixin.qq.com')


def get_wechat_access_token(force_refresh=False):
    cache_key = 'professor_student_manage_admin_wechat_access_token'
    if not force_refresh:
        cached_token = cache.get(cache_key)
        if cached_token:
            return cached_token

    response = requests.post(
        f'{WECHAT_API_BASE}/cgi-bin/stable_token',
        json={
            'grant_type': 'client_credential',
            'appid': WECHAT_APPID,
            'secret': WECHAT_SECRET,
            'force_refresh': bool(force_refresh),
        },
        timeout=15,
    )
    response.raise_for_status()
    payload = response.json()
    access_token = payload.get('access_token')
    if not access_token:
        raise ValueError(payload.get('errmsg') or '获取微信 access_token 失败。')

    expires_in = int(payload.get('expires_in') or 7200)
    cache.set(cache_key, access_token, timeout=max(expires_in - 300, 300))
    return access_token

class ProfessorDoctorQuotaInline(admin.TabularInline):
    model = ProfessorDoctorQuota
    extra = 0  # 不显示额外的空行
    fields = ['subject', 'total_quota', 'used_quota', 'remaining_quota']
    readonly_fields = ['used_quota', 'remaining_quota']  # 已用和剩余名额只读
    can_delete = False  # 禁止删除，确保每个博士专业都有记录


def swap_doctor_quotas(modeladmin, request, queryset):
    """
    互换两个导师的博士专业名额（专业、总名额和剩余名额）
    """
    # 验证只选择了两条记录
    if queryset.count() != 2:
        modeladmin.message_user(
            request, 
            "请恰好选择两个导师的博士专业名额进行互换", 
            level='error'
        )
        return
    
    # 获取两条记录
    quota1, quota2 = list(queryset)
    
    # 互换专业和名额
    quota1_subject = quota1.subject
    quota1_total = quota1.total_quota
    quota1_remaining = quota1.remaining_quota
    
    quota1.subject = quota2.subject
    quota1.total_quota = quota2.total_quota
    quota1.remaining_quota = quota2.remaining_quota
    
    quota2.subject = quota1_subject
    quota2.total_quota = quota1_total
    quota2.remaining_quota = quota1_remaining
    
    # 保存修改
    quota1.save()
    quota2.save()
    
    # 显示互换结果
    message = f"成功互换 {quota1.professor.name}({quota1.subject.subject_name}) 和 {quota2.professor.name}({quota2.subject.subject_name}) 的博士专业名额"
    
    modeladmin.message_user(request, message)

swap_doctor_quotas.short_description = "互换两个导师的博士专业名额"


@admin.register(ProfessorDoctorQuota)
class ProfessorDoctorQuotaAdmin(admin.ModelAdmin):
    list_display = ['professor', 'subject', 'total_quota', 'used_quota', 'remaining_quota']
    list_filter = ['subject']
    search_fields = ['professor__name', 'subject__subject_name']
    actions = [swap_doctor_quotas]

# ========= 新增硕士专业内联 =========
class ProfessorMasterQuotaInline(admin.TabularInline):
    model = ProfessorMasterQuota
    extra = 0
    fields = ['subject', 'beijing_quota', 'beijing_remaining_quota',
              'yantai_quota', 'yantai_remaining_quota', 'total_quota']
    # readonly_fields = ['beijing_remaining_quota', 'yantai_remaining_quota', 'total_quota']
    readonly_fields = ['total_quota']
    can_delete = False

@admin.register(ProfessorMasterQuota)
class ProfessorMasterQuotaAdmin(admin.ModelAdmin):
    list_display = ['professor', 'subject', 'total_quota', 'beijing_quota', 'beijing_remaining_quota',
                    'yantai_quota', 'yantai_remaining_quota']
    list_filter = ['subject']
    search_fields = ['professor__name', 'subject__subject_name']
    readonly_fields = ['beijing_remaining_quota', 'yantai_remaining_quota', 'total_quota']


@admin.action(description="重置导师指定类型的名额")
def reset_quota(modeladmin, request, queryset):
    # 获取用户选择的类型
    quota_type = request.POST.get('quota_type')

    # 根据类型重置名额
    if quota_type == 'academic':
        queryset.update(academic_quota=0)
    elif quota_type == 'professional':
        queryset.update(professional_quota=0)
    elif quota_type == 'professionalyt':
        queryset.update(professional_yt_quota=0)
    elif quota_type == 'doctor':
        queryset.update(doctor_quota=0)
    else:
        modeladmin.message_user(request, "请选择有效的名额类型", level='error')
        return

    modeladmin.message_user(request, f"已成功重置 {queryset.count()} 位导师的 {quota_type} 名额为 0")


@admin.action(description="重置导师状态为未开放选择: ")
def reset_proposed_quota_approved(modeladmin, request, queryset):
    # 将选中的导师的 proposed_quota_approved 字段重置为 False
    queryset.update(proposed_quota_approved=False)
    modeladmin.message_user(request, f"已成功重置 {queryset.count()} 位导师的“设置指标”为 False")


# 文件上传表单
class ImportQuotaForm(forms.Form):
    csv_file = forms.FileField(label="选择 CSV 文件")

class ImportDoctorQuotaForm(forms.Form):
    xlsx_file = forms.FileField(label="选择 XLSX 文件")


class ProfessorAdmin(admin.ModelAdmin):
    fieldsets = [
        ("导师信息更改", {"fields": ["name", "teacher_identity_id", "professor_title", "email", "department",
                                     "academic_quota", "professional_quota", "professional_yt_quota", "doctor_quota", "proposed_quota_approved",
                                     "have_qualification", "remaining_quota", "personal_page", "research_areas",
                                     "avatar", "contact_details", "department_position"]}),
    ]
    list_display = ["teacher_identity_id", "name", "department", "have_qualification", "proposed_quota_approved"]
    readonly_fields = ["remaining_quota"]
    actions = [reset_quota, reset_proposed_quota_approved, 'reset_password_to_teacher_id']
    change_list_template = 'admin/professor_change_list.html'  # 自定义列表页面模板
    inlines = [ProfessorMasterQuotaInline, ProfessorDoctorQuotaInline]  # 内联显示硕士、博士专业名额

    def reset_password_to_teacher_id(self, request, queryset):
        """
        将选中导师的密码重置为工号（teacher_identity_id）
        """
        for professor in queryset:
            if professor.user_name:  # 确保关联的 User 对象存在
                teacher_id = professor.teacher_identity_id
                professor.user_name.set_password(teacher_id)  # 重置密码
                professor.user_name.save()
                self.message_user(
                    request,
                    f"已重置导师 {professor.name} 的密码为工号: {teacher_id}",
                    level='success'
                )
    reset_password_to_teacher_id.short_description = "重置密码为工号"  # 动作显示名称

    def get_actions(self, request):
        actions = super().get_actions(request)
        # 添加自定义动作选项
        actions['reset_academic_quota'] = (
            reset_quota,
            'reset_academic_quota',
            '重置学硕名额为 0'
        )
        actions['reset_professional_quota'] = (
            reset_quota,
            'reset_professional_quota',
            '重置北京专硕名额为 0'
        )
        actions['reset_professionalyt_quota'] = (
            reset_quota,
            'reset_professionalyt_quota',
            '重置烟台专硕名额为 0'
        )
        actions['reset_doctor_quota'] = (
            reset_quota,
            'reset_doctor_quota',
            '重置博士名额为 0'
        )
        actions['reset_proposed_quota_approved'] = (
            reset_proposed_quota_approved,
            'reset_proposed_quota_approved',
            '重置所选导师状态为未开放选择'
        )
        return actions

    def response_action(self, request, queryset):
        # 获取用户选择的动作
        action = request.POST.get('action')
        if action in ['reset_academic_quota', 'reset_professional_quota', 'reset_professionalyt_quota', 'reset_doctor_quota']:
            # 设置 quota_type
            print("action: ", action)
            quota_type = action.split('_')[1]  # 从动作名称中提取类型
            print("quota_type: ", quota_type)
            request.POST = request.POST.copy()  # 使 POST 数据可变
            request.POST['quota_type'] = quota_type
        return super().response_action(request, queryset)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-quota/', self.admin_site.admin_view(self.import_quota_view), name='import_quota'),
            path('import-doctor-quota/', self.admin_site.admin_view(self.import_doctor_quota_view), name='import_doctor_quota'),
        ]
        return custom_urls + urls

    def import_quota_view(self, request):
        # 处理确认导入请求
        if request.method == 'POST' and 'confirm_import' in request.POST:
            import_data_json = request.POST.get('import_data')
            import_mode = request.POST.get('import_mode', 'full')
            sync_quotas = request.POST.get('sync_quotas') == 'yes'
            
            if import_data_json:
                import json
                import_data = json.loads(import_data_json)
                self._process_import_data(request, import_data, import_mode, sync_quotas)
                return redirect('admin:Professor_Student_Manage_professor_changelist')
        
        if request.method == 'POST':
            form = ImportQuotaForm(request.POST, request.FILES)
            if form.is_valid():
                csv_file = request.FILES['csv_file']
                import_mode = request.POST.get('import_mode', 'full')  # 获取导入模式
                
                try:
                    # 读取 CSV 文件
                    csv_file_wrapper = TextIOWrapper(csv_file, encoding='utf-8-sig')
                    reader = csv.DictReader(csv_file_wrapper)

                    # 先收集所有导入数据
                    import_data = []
                    for row in reader:
                        teacher_identity_id = str(row["工号"]).zfill(5)
                        try:
                            professor = Professor.objects.get(teacher_identity_id=teacher_identity_id)
                        except Professor.DoesNotExist:
                            continue

                        # 循环读取 5 个招生学科
                        for i in range(1, 6):
                            subject_name = row.get(f"招生学科{i}", "").strip()
                            subject_code = str(row.get(f"学科{i}代码", "")).strip()
                            subject_type = row.get(f"专业类型{i}", "").strip()
                            bj_quota = row.get(f"北京招生名额{i}", "").strip()
                            yt_quota = row.get(f"烟台招生名额{i}", "").strip()

                            # 跳过空或无效学科
                            if not subject_name or subject_name == "无":
                                continue

                            # 学科代码补齐为 6 位
                            if subject_code and len(subject_code) == 5:
                                subject_code = subject_code.zfill(6)

                            subject = Subject.objects.filter(subject_code=subject_code).first()
                            if not subject:
                                continue

                            # 转换为整数
                            try:
                                bj_quota = int(bj_quota) if bj_quota else 0
                                yt_quota = int(yt_quota) if yt_quota else 0
                            except ValueError:
                                continue

                            # 收集导入数据
                            import_data.append({
                                'professor_id': professor.id,
                                'professor_name': professor.name,
                                'subject_id': subject.id,
                                'subject_name': subject.subject_name,
                                'subject_code': subject.subject_code,
                                'bj_quota': bj_quota,
                                'yt_quota': yt_quota,
                            })

                    # 验证导入数据并生成对比信息
                    validation_result = self._validate_quota_import(import_data, import_mode)
                    
                    if validation_result['need_confirm']:
                        # 需要用户确认
                        import json
                        context = {
                            'form': form,
                            'opts': self.model._meta,
                            'title': validation_result['title'],
                            'comparison_data': validation_result['comparison_data'],
                            'warnings': validation_result.get('warnings', []),
                            'import_data_json': json.dumps(import_data),
                            'import_mode': import_mode,
                            'message': validation_result.get('message', ''),
                        }
                        return render(request, 'admin/import_quota_confirm.html', context)
                    else:
                        # 不需要确认，直接导入
                        self._process_import_data(request, import_data, import_mode, sync_quotas=False)
                        return redirect('admin:Professor_Student_Manage_professor_changelist')

                except Exception as e:
                    self.message_user(request, f"解析 CSV 文件时出错: {str(e)}", level='error')
                    return redirect('admin:Professor_Student_Manage_professor_changelist')
        else:
            form = ImportQuotaForm()

        context = {
            'form': form,
            'opts': self.model._meta,
            'title': '一键导入导师硕士名额',
        }
        return render(request, 'admin/import_quota.html', context)
    
    def import_doctor_quota_view(self, request):
        """
        博士名额部分导入功能
        支持XLSX格式，根据工号和学科代码更新博士招生名额
        """
        # 处理确认导入请求
        if request.method == 'POST' and 'confirm_import' in request.POST:
            import json
            import_data_json = request.POST.get('import_data')
            conflict_action = request.POST.get('conflict_action', 'replace')  # replace 或 add
            
            if import_data_json:
                import_data = json.loads(import_data_json)
                self._process_doctor_quota_import(request, import_data, conflict_action)
                return redirect('admin:Professor_Student_Manage_professor_changelist')
        
        if request.method == 'POST':
            form = ImportDoctorQuotaForm(request.POST, request.FILES)
            if form.is_valid():
                xlsx_file = request.FILES['xlsx_file']
                
                try:
                    import openpyxl
                    from collections import defaultdict
                    
                    # 读取XLSX文件
                    wb = openpyxl.load_workbook(xlsx_file)
                    ws = wb.active
                    
                    # 读取表头
                    headers = [cell.value for cell in ws[1]]
                    
                    # 查找核心列的索引
                    try:
                        col_name = headers.index('姓名') + 1
                        col_teacher_id = headers.index('工号') + 1
                        col_direction = headers.index('方向') + 1
                    except ValueError as e:
                        self.message_user(request, f"文件列名不正确，缺少必要列（姓名、工号、方向）: {str(e)}", level='error')
                        return redirect('admin:Professor_Student_Manage_professor_changelist')
                    
                    # 解析导入数据
                    import_data = []
                    conflicts = []  # 存储有冲突的记录
                    
                    for row in ws.iter_rows(min_row=2, values_only=False):
                        if not row[col_teacher_id - 1].value:
                            continue
                        
                        teacher_id = str(row[col_teacher_id - 1].value).strip().zfill(5)
                        teacher_name = str(row[col_name - 1].value).strip() if row[col_name - 1].value else ""
                        
                        # 查找导师
                        try:
                            professor = Professor.objects.get(teacher_identity_id=teacher_id)
                        except Professor.DoesNotExist:
                            self.message_user(request, f"工号 {teacher_id} 的导师不存在，跳过", level='warning')
                            continue
                        
                        # 遍历所有可能的招生学科列（假设最多有5组）
                        for i in range(1, 10):  # 支持最多9组学科
                            try:
                                # 查找学科相关列
                                subject_col_name = f'招生学科{i}'
                                code_col_name = f'学科{i}代码'
                                type_col_name = f'专业类型{i}'
                                quota_col_name = f'本次名额{i}'
                                
                                if subject_col_name not in headers:
                                    break
                                
                                col_subject = headers.index(subject_col_name) + 1
                                col_code = headers.index(code_col_name) + 1
                                col_type = headers.index(type_col_name) + 1
                                col_quota = headers.index(quota_col_name) + 1
                                
                                quota_value = row[col_quota - 1].value
                                
                                # 只处理本次名额>0的记录
                                if not quota_value or int(quota_value) <= 0:
                                    continue
                                
                                subject_code = str(row[col_code - 1].value).strip() if row[col_code - 1].value else ""
                                subject_type = str(row[col_type - 1].value).strip() if row[col_type - 1].value else ""
                                quota = int(quota_value)
                                
                                if not subject_code or subject_type != "博士":
                                    continue
                                
                                # 查找专业（博士类型）
                                subject = Subject.objects.filter(subject_code=subject_code, subject_type=2).first()
                                if not subject:
                                    self.message_user(request, f"博士专业代码 {subject_code} 不存在，跳过", level='warning')
                                    continue
                                
                                # 检查是否已存在名额记录
                                existing_quota = ProfessorDoctorQuota.objects.filter(
                                    professor=professor,
                                    subject=subject
                                ).first()
                                
                                import_item = {
                                    'professor_id': professor.id,
                                    'professor_name': professor.name,
                                    'teacher_identity_id': teacher_id,
                                    'subject_id': subject.id,
                                    'subject_name': subject.subject_name,
                                    'subject_code': subject_code,
                                    'quota': quota,
                                    'existing_quota': existing_quota.total_quota if existing_quota else 0,
                                    'has_conflict': existing_quota and existing_quota.total_quota > 0,
                                }
                                
                                import_data.append(import_item)
                                
                                # 如果有冲突，记录下来
                                if import_item['has_conflict']:
                                    conflicts.append(import_item)
                                    
                            except (ValueError, IndexError):
                                continue
                    
                    if not import_data:
                        self.message_user(request, "没有找到有效的导入数据（本次名额必须>0）", level='warning')
                        return redirect('admin:Professor_Student_Manage_professor_changelist')
                    
                    # 验证名额分配情况
                    validation_result = self._validate_doctor_quota_import(import_data)
                    
                    # 如果有冲突或超额，需要用户确认
                    if conflicts or validation_result['warnings']:
                        import json
                        context = {
                            'form': form,
                            'opts': self.model._meta,
                            'title': '博士名额部分导入确认',
                            'import_data': import_data,
                            'conflicts': conflicts,
                            'validation_result': validation_result,
                            'import_data_json': json.dumps(import_data),
                        }
                        return render(request, 'admin/import_doctor_quota_confirm.html', context)
                    else:
                        # 没有冲突和超额，直接导入
                        self._process_doctor_quota_import(request, import_data, 'replace')
                        return redirect('admin:Professor_Student_Manage_professor_changelist')
                    
                except Exception as e:
                    import traceback
                    self.message_user(request, f"解析 XLSX 文件时出错: {str(e)}\\n{traceback.format_exc()}", level='error')
                    return redirect('admin:Professor_Student_Manage_professor_changelist')
        else:
            form = ImportDoctorQuotaForm()
        
        context = {
            'form': form,
            'opts': self.model._meta,
            'title': '博士名额部分导入',
        }
        return render(request, 'admin/import_doctor_quota.html', context)
        return render(request, 'admin/import_quota.html', context)
    
    def _validate_quota_import(self, import_data, import_mode='full'):
        """
        验证导入数据，根据导入模式生成不同的验证结果
        全量模式：显示导入名额与总招生名额的对比
        增量模式：显示将要追加的名额并提示更新总名额
        """
        # 按学科代码汇总导入的名额
        subject_quotas = defaultdict(lambda: {
            'bj_total': 0, 
            'yt_total': 0,
            'total': 0, 
            'subject_obj': None,
            'subject_code': None,
            'subject_name': None
        })
        
        for item in import_data:
            subject_code = item['subject_code']
            bj_quota = item['bj_quota']
            yt_quota = item['yt_quota']
            
            subject_quotas[subject_code]['bj_total'] += bj_quota
            subject_quotas[subject_code]['yt_total'] += yt_quota
            subject_quotas[subject_code]['total'] += (bj_quota + yt_quota)
            if subject_quotas[subject_code]['subject_obj'] is None:
                subject_quotas[subject_code]['subject_obj'] = Subject.objects.get(id=item['subject_id'])
                subject_quotas[subject_code]['subject_code'] = subject_code
                subject_quotas[subject_code]['subject_name'] = item['subject_name']
        
        comparison_data = []
        warnings = []
        need_confirm = False
        
        if import_mode == 'full':
            # 全量导入模式：显示导入名额与总招生名额的对比
            for subject_code, data in subject_quotas.items():
                subject = data['subject_obj']
                import_total = data['total']
                current_total = subject.total_admission_quota or 0
                
                comparison_data.append({
                    'subject_name': data['subject_name'],
                    'subject_code': subject_code,
                    'current_total': current_total,
                    'import_total': import_total,
                    'bj_quota': data['bj_total'],
                    'yt_quota': data['yt_total'],
                    'difference': import_total - current_total,
                    'is_exceed': import_total > current_total,
                })
                
                # 如果导入名额超过总招生名额，需要确认
                if import_total > current_total:
                    need_confirm = True
                    warnings.append({
                        'subject_name': data['subject_name'],
                        'subject_code': subject_code,
                        'current_total': current_total,
                        'import_total': import_total,
                        'exceed': import_total - current_total,
                    })
            
            return {
                'need_confirm': need_confirm,
                'title': '全量导入确认 - 名额对比' if need_confirm else '全量导入',
                'comparison_data': comparison_data,
                'warnings': warnings,
                'message': '以下专业的导入名额超过当前总招生名额，是否更新总招生名额并调整学生候补状态？' if need_confirm else '',
            }
        
        else:  # incremental mode
            # 增量导入模式：显示追加信息并提示更新总名额
            for subject_code, data in subject_quotas.items():
                subject = data['subject_obj']
                import_total = data['total']
                current_total = subject.total_admission_quota or 0
                new_total = current_total + import_total
                
                comparison_data.append({
                    'subject_name': data['subject_name'],
                    'subject_code': subject_code,
                    'current_total': current_total,
                    'import_total': import_total,
                    'bj_quota': data['bj_total'],
                    'yt_quota': data['yt_total'],
                    'new_total': new_total,
                    'increase': import_total,
                })
            
            return {
                'need_confirm': True,  # 增量模式总是需要确认
                'title': '增量导入确认 - 追加名额',
                'comparison_data': comparison_data,
                'warnings': [],
                'message': '增量导入将在导师现有名额基础上追加新名额（包括剩余名额），并更新专业总招生名额，是否重新计算学生候补状态？',
            }

    def _process_import_data(self, request, import_data, import_mode='full', sync_quotas=False):
        """
        处理导入数据，保存到数据库
        import_mode: 'full' 全量导入（覆盖），'incremental' 增量导入（追加）
        sync_quotas: 是否同步更新专业总招生名额并调整候补状态（仅在增量模式下有效）
        """
        from Enrollment_Manage.models import sync_student_alternate_status
        
        # 按专业代码（subject_code）汇总导入的名额（分北京和烟台）
        subject_quotas = defaultdict(lambda: {
            'bj_total': 0, 
            'yt_total': 0, 
            'total': 0, 
            'subject_obj': None,
            'subject_code': None,
            'subject_name': None
        })
        for item in import_data:
            subject_code = item['subject_code']  # 使用学科代码作为key
            bj_quota = item['bj_quota']
            yt_quota = item['yt_quota']
            
            subject_quotas[subject_code]['bj_total'] += bj_quota
            subject_quotas[subject_code]['yt_total'] += yt_quota
            subject_quotas[subject_code]['total'] += (bj_quota + yt_quota)
            if subject_quotas[subject_code]['subject_obj'] is None:
                subject_quotas[subject_code]['subject_obj'] = Subject.objects.get(id=item['subject_id'])
                subject_quotas[subject_code]['subject_code'] = subject_code
                subject_quotas[subject_code]['subject_name'] = item['subject_name']
        
        # 处理专业总招生名额的更新
        if import_mode == 'incremental':
            # 增量模式：总名额 = 当前名额 + 导入名额
            for subject_code, data in subject_quotas.items():
                subject = data['subject_obj']
                import_total = data['total']
                old_quota = subject.total_admission_quota or 0
                new_quota = old_quota + import_total
                
                subject.total_admission_quota = new_quota
                subject.save()
                
                if sync_quotas:
                    # 同步候补状态
                    updated = sync_student_alternate_status(subject)
                    self.message_user(
                        request,
                        f"专业 {subject.subject_name} 总招生名额已从 {old_quota} 更新为 {new_quota}，"
                        f"同步调整了 {updated} 名学生的候补状态",
                        level='success'
                    )
                else:
                    self.message_user(
                        request,
                        f"专业 {subject.subject_name} 总招生名额已从 {old_quota} 更新为 {new_quota}（未调整候补状态）",
                        level='info'
                    )
        else:
            # 全量模式：更新专业总名额为导入的名额总和
            for subject_code, data in subject_quotas.items():
                subject = data['subject_obj']
                import_total = data['total']
                current_total = subject.total_admission_quota or 0
                
                if import_total != current_total:
                    subject.total_admission_quota = import_total
                    subject.save()
                    
                    if sync_quotas and import_total > current_total:
                        # 仅当名额增加且用户选择时才调整候补
                        updated = sync_student_alternate_status(subject)
                        self.message_user(
                            request,
                            f"专业 {subject.subject_name} 总招生名额已从 {current_total} 更新为 {import_total}，"
                            f"同步调整了 {updated} 名学生的候补状态",
                            level='success'
                        )
                    else:
                        self.message_user(
                            request,
                            f"专业 {subject.subject_name} 总招生名额已更新为 {import_total}",
                            level='info'
                        )
        
        # 保存导师名额数据
        success_count = 0
        for item in import_data:
            professor = Professor.objects.get(id=item['professor_id'])
            subject = Subject.objects.get(id=item['subject_id'])
            bj_quota = item['bj_quota']
            yt_quota = item['yt_quota']
            
            quota_obj, created = ProfessorMasterQuota.objects.get_or_create(
                professor=professor,
                subject=subject,
                defaults={
                    'beijing_quota': bj_quota,
                    'yantai_quota': yt_quota,
                    'beijing_remaining_quota': bj_quota,
                    'yantai_remaining_quota': yt_quota,
                }
            )
            
            if not created:
                if import_mode == 'incremental':
                    # 增量模式：在原有基础上增加（包括剩余名额）
                    quota_obj.beijing_quota += bj_quota
                    quota_obj.yantai_quota += yt_quota
                    quota_obj.beijing_remaining_quota += bj_quota  # 剩余名额也要追加
                    quota_obj.yantai_remaining_quota += yt_quota  # 剩余名额也要追加
                else:
                    # 全量模式：直接覆盖
                    quota_obj.beijing_quota = bj_quota
                    quota_obj.yantai_quota = yt_quota
                    quota_obj.beijing_remaining_quota = bj_quota
                    quota_obj.yantai_remaining_quota = yt_quota
                quota_obj.save()
            
            success_count += 1
        
        # 显示导入模式信息
        mode_text = "增量导入（追加）" if import_mode == 'incremental' else "全量导入（覆盖）"
        self.message_user(request, f"[{mode_text}] 成功导入 {success_count} 条导师招生名额数据", level='success')
        
        # 按学科代码排序后显示，为每个专业生成详细的名额统计信息
        for subject_code in sorted(subject_quotas.keys()):
            data = subject_quotas[subject_code]
            subject_name = data['subject_name']
            bj_total = data['bj_total']
            yt_total = data['yt_total']
            total = data['total']
            
            # 构建详细信息
            if yt_total > 0:
                quota_detail = f"北京 {bj_total} 人，烟台 {yt_total} 人，合计 {total} 人"
            else:
                quota_detail = f"北京 {bj_total} 人"
            
            # 显示专业名额统计
            self.message_user(
                request,
                f"📊 {subject_name} ({subject_code}): {quota_detail}",
                level='success'
            )

    def _validate_doctor_quota_import(self, import_data):
        """
        验证博士名额导入数据
        检查每个专业的导师总名额是否超过专业总招生名额
        """
        from collections import defaultdict
        
        # 按专业汇总导入的名额
        subject_quotas = defaultdict(lambda: {
            'total': 0,
            'subject_obj': None,
            'subject_code': None,
            'subject_name': None
        })
        
        for item in import_data:
            subject_code = item['subject_code']
            quota = item['quota']
            teacher_identity_id = item.get('teacher_identity_id', '')
            
            # 排除工号为csds开头的测试账号
            if teacher_identity_id.startswith('csds'):
                continue
            
            # 如果是覆盖模式（has_conflict=True且用户选择replace），使用新名额
            # 如果是增加模式（has_conflict=True且用户选择add），需要加上原有名额
            # 如果没有冲突，直接使用新名额
            subject_quotas[subject_code]['total'] += quota
            
            if subject_quotas[subject_code]['subject_obj'] is None:
                subject = Subject.objects.get(id=item['subject_id'])
                subject_quotas[subject_code]['subject_obj'] = subject
                subject_quotas[subject_code]['subject_code'] = subject_code
                subject_quotas[subject_code]['subject_name'] = item['subject_name']
        
        # 检查每个专业的总名额
        warnings = []
        comparison_data = []
        
        for subject_code, data in subject_quotas.items():
            subject = data['subject_obj']
            import_total = data['total']
            subject_total = subject.total_admission_quota or 0
            
            # 注意：这里只是导入的名额总和，实际还需要加上已存在的名额（如果不是覆盖）
            comparison_data.append({
                'subject_name': data['subject_name'],
                'subject_code': subject_code,
                'import_total': import_total,
                'subject_total': subject_total,
                'is_exceed': import_total > subject_total,
            })
            
            if import_total > subject_total:
                warnings.append({
                    'subject_name': data['subject_name'],
                    'subject_code': subject_code,
                    'import_total': import_total,
                    'subject_total': subject_total,
                    'exceed': import_total - subject_total,
                })
        
        return {
            'comparison_data': comparison_data,
            'warnings': warnings,
        }
    
    def _process_doctor_quota_import(self, request, import_data, conflict_action='replace'):
        """
        处理博士名额导入
        conflict_action: 'replace' 覆盖，'add' 增加
        """
        from collections import defaultdict
        from Enrollment_Manage.models import sync_student_alternate_status
        
        success_count = 0
        
        # 按专业汇总导入的名额
        subject_totals = defaultdict(int)
        
        for item in import_data:
            professor = Professor.objects.get(id=item['professor_id'])
            subject = Subject.objects.get(id=item['subject_id'])
            quota = item['quota']
            
            # 跳过测试账号
            if item.get('teacher_identity_id', '').startswith('csds'):
                continue
            
            quota_obj, created = ProfessorDoctorQuota.objects.get_or_create(
                professor=professor,
                subject=subject,
                defaults={
                    'total_quota': quota,
                    'used_quota': 0,
                    'remaining_quota': quota,
                }
            )
            
            if not created:
                # 已存在的记录
                if conflict_action == 'add':
                    # 增加模式：在原有基础上增加
                    quota_obj.total_quota += quota
                    quota_obj.remaining_quota += quota
                else:
                    # 覆盖模式：直接覆盖
                    used_quota = quota_obj.used_quota or 0
                    quota_obj.total_quota = quota
                    quota_obj.remaining_quota = quota - used_quota
                
                quota_obj.save()
            
            success_count += 1
            subject_totals[subject.id] += quota
        
        # 更新专业总招生名额并同步候补状态（仅当导师名额总和超过专业总招生名额时）
        for subject_id, allocated_total in subject_totals.items():
            subject = Subject.objects.get(id=subject_id)
            old_quota = subject.total_admission_quota or 0
            
            # 计算该专业所有导师的总名额（排除测试账号）
            all_quotas = ProfessorDoctorQuota.objects.filter(subject=subject).exclude(
                professor__teacher_identity_id__startswith='csds'
            )
            new_total_quota = all_quotas.aggregate(total=Sum('total_quota'))['total'] or 0
            
            # 只有当导师名额总和超过专业总招生名额时，才更新总招生名额
            if new_total_quota > old_quota:
                # 更新专业总招生名额
                subject.total_admission_quota = new_total_quota
                subject.save()
                
                # 同步候补状态
                updated_count = sync_student_alternate_status(subject)
                
                self.message_user(
                    request,
                    f"✅ 专业 {subject.subject_name} 总招生名额已从 {old_quota} 自动增加为 {new_total_quota}，"
                    f"同步调整了 {updated_count} 名学生的候补状态",
                    level='success'
                )
            elif new_total_quota < old_quota:
                # 导师名额总和小于专业总招生名额，不修改专业总招生名额
                self.message_user(
                    request,
                    f"📊 专业 {subject.subject_name} 导师名额总和为 {new_total_quota} 人，"
                    f"专业总招生名额保持为 {old_quota} 人（未调整）",
                    level='info'
                )
            else:
                # 名额相等，无需调整
                self.message_user(
                    request,
                    f"✓ 专业 {subject.subject_name} 导师名额总和为 {new_total_quota} 人，"
                    f"与专业总招生名额一致",
                    level='success'
                )
        
        # 显示导入结果
        action_text = "增加" if conflict_action == 'add' else "覆盖"
        self.message_user(request, f"[博士名额部分导入] 成功处理 {success_count} 条记录（{action_text}模式）", level='success')
        
        # 按专业统计并显示
        subject_stats = defaultdict(int)
        for item in import_data:
            if not item.get('teacher_identity_id', '').startswith('csds'):
                subject_stats[item['subject_name']] += item['quota']
        
        for subject_name in sorted(subject_stats.keys()):
            total = subject_stats[subject_name]
            self.message_user(request, f"📊 {subject_name}: 导入名额 {total} 人", level='success')

    def get_readonly_fields(self, request, obj=None):
        readonly_fields = self.readonly_fields
        if obj and obj.proposed_quota_approved == True:  # 根据字段值来判断是否设置为只读
            # readonly_fields = ["academic_quota", "professional_quota", "professional_yt_quota", "doctor_quota",
            #                    "have_qualification", "remaining_quota"]
            readonly_fields = ["have_qualification", "remaining_quota"]
        return readonly_fields

    list_filter = ["department_id", "proposed_quota_approved", "have_qualification"]
    search_fields = ["name"]


class ImportStudentForm(forms.Form):
    IMPORT_TYPE_CHOICES = [
        ('doctor', '博士生申请导入'),
        ('master_exam', '硕士统考生导入'),
        ('master_recommend', '硕士推免生导入'),
    ]
    import_type = forms.ChoiceField(
        label="导入类型",
        choices=IMPORT_TYPE_CHOICES,
        widget=forms.RadioSelect,
        initial='doctor'
    )
    file = forms.FileField(label="选择文件（支持CSV或XLSX）")
    update_quota = forms.BooleanField(
        label="是否根据表格信息更新专业总招生人数",
        required=False,
        initial=True
    )


class CustomModelChoiceField(forms.ModelChoiceField):
    def label_from_instance(self, obj):
        # 自定义下拉菜单选项显示为：专业名称 (专业类别)
        return f"{obj.subject_name} ({obj.get_subject_type_display()})"


class StudentAdmin(admin.ModelAdmin):
    # fieldsets 元组中的第一个元素是字段集的标题
    fieldsets = [
        ("学生信息更改", {"fields": ["name", "candidate_number", "student_type", "subject", 
                               "postgraduate_type", "study_mode", "resume", "avatar", "phone_number",
                               "initial_exam_score", "initial_rank", "secondary_exam_score",
                               "secondary_rank", "final_rank", "signature_table_student_signatured", 
                               "signature_table_professor_signatured", "signature_table_review_status", 
                               "is_selected", "is_giveup", "is_alternate"]}),
    ]
    list_display = ["candidate_number", "name", "subject", "study_mode", "student_type", "postgraduate_type", "is_selected", 
                    "selected_professor", "is_giveup", "is_alternate", "download_hx_file", "download_fq_file"]
    list_filter = ["subject"]
    search_fields = ["name"]
    actions = ['reset_password_to_exam_id', 'download_all_signature_tables', 'download_all_giveup_tables']  # 添加自定义动作
    change_list_template = 'admin/student_change_list.html'
    
    def selected_professor(self, obj):
        """显示学生选中的导师姓名"""
        if obj.is_selected:
            # 查找状态为"已同意"且被导师选中的互选记录
            choice = StudentProfessorChoice.objects.filter(
                student=obj, 
                status=1,
                chosen_by_professor=True
            ).first()
            return choice.professor.name if choice else "未找到"
        return "-"
    selected_professor.short_description = "选中导师"

    @admin.action(description="批量下载所有已签名互选表")
    def download_all_signature_tables(self, request, queryset):
        """
        批量下载所有学生和导师均已签名的互选表，打包成zip
        """
        # 过滤出所有已签名的学生
        signed_students = queryset.filter(
            signature_table_student_signatured=True,
            signature_table_professor_signatured=True,
            signature_table__isnull=False
        )
        # print(signed_students)

        if not signed_students.exists():
            self.message_user(request, "没有符合条件的已签名互选表", level='warning')
            return

        # 临时文件夹
        temp_dir = tempfile.mkdtemp()
        zip_filename = os.path.join(temp_dir, "互选表打包下载.zip")

        with zipfile.ZipFile(zip_filename, 'w') as zipf:
            for student in signed_students:
                try:
                    # 找到该学生已同意的导师
                    choice = StudentProfessorChoice.objects.filter(
                        student=student, status=1
                    ).first()
                    professor_name = choice.professor.name if choice else "无导师"

                    # 获取文件下载地址
                    response_data = self.get_fileid_download_url(student.signature_table)
                    if response_data.get("errcode") == 0:
                        download_url = response_data['file_list'][0]['download_url']
                        file_content = requests.get(download_url).content

                        # 文件命名: 准考证号-学生姓名-导师姓名.pdf
                        filename = f"{student.candidate_number}-{student.name}-{professor_name}.pdf"
                        print(filename)

                        # 写入zip
                        file_path = os.path.join(temp_dir, filename)
                        with open(file_path, 'wb') as f:
                            f.write(file_content)
                        zipf.write(file_path, arcname=filename)
                except Exception as e:
                    print(f"下载学生 {student.name} 的互选表失败: {e}")
                    continue

        with open(zip_filename, 'rb') as f:
            response = HttpResponse(f.read(), content_type="application/zip")
            response['Content-Disposition'] = 'attachment; filename="互选表打包下载.zip"'
            return response

    @admin.action(description="批量下载所有已签名弃选表")
    def download_all_giveup_tables(self, request, queryset):
        """
        批量下载所有已签名弃选说明表的学生文件，打包成zip
        """
        # 过滤出所有已放弃且已签名弃选表的学生
        giveup_students = queryset.filter(
            is_giveup=True,
            is_signate_giveup_table=True,
            giveup_signature_table__isnull=False
        )

        if not giveup_students.exists():
            self.message_user(request, "没有符合条件的已签名弃选表", level='warning')
            return

        # 临时文件夹
        temp_dir = tempfile.mkdtemp()
        zip_filename = os.path.join(temp_dir, "弃选表打包下载.zip")

        with zipfile.ZipFile(zip_filename, 'w') as zipf:
            for student in giveup_students:
                try:
                    # 获取文件下载地址
                    response_data = self.get_fileid_download_url(student.giveup_signature_table)
                    if response_data.get("errcode") == 0:
                        download_url = response_data['file_list'][0]['download_url']
                        file_content = requests.get(download_url).content

                        # 文件命名: 准考证号-学生姓名-弃选表.pdf
                        filename = f"{student.candidate_number}-{student.name}-弃选表.pdf"
                        print(filename)

                        # 写入zip
                        file_path = os.path.join(temp_dir, filename)
                        with open(file_path, 'wb') as f:
                            f.write(file_content)
                        zipf.write(file_path, arcname=filename)
                except Exception as e:
                    print(f"下载学生 {student.name} 的弃选表失败: {e}")
                    continue

        with open(zip_filename, 'rb') as f:
            response = HttpResponse(f.read(), content_type="application/zip")
            response['Content-Disposition'] = 'attachment; filename="弃选表打包下载.zip"'
            return response

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "subject":
            # 使用自定义 ModelChoiceField 显示 subject_name (subject_type_display)
            kwargs["form_class"] = CustomModelChoiceField
            kwargs["queryset"] = Subject.objects.all()
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    # 添加自定义URL
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-students/', self.admin_site.admin_view(self.import_students_view), name='import_students'),
            path('get-download-url/', self.admin_site.admin_view(self.get_download_url_view), name='get_download_url'),
        ]

        # 调试：打印所有 URL 模式
        # print("Custom URLs:", custom_urls)
        # print("All URLs:", urls)
        return custom_urls + urls

    # 处理CSV文件上传和学生创建的视图
    # def import_students_view(self, request):
    #     if request.method == 'POST':
    #         form = ImportStudentForm(request.POST, request.FILES)
    #         if form.is_valid():
    #             csv_file = request.FILES['csv_file']
    #             try:
    #                 # 读取 CSV 文件
    #                 csv_file_wrapper = TextIOWrapper(csv_file, encoding='utf-8-sig')
    #                 reader = csv.DictReader(csv_file_wrapper)

                    
    #                 # 检查列名是否正确
    #                 required_columns = ["专业代码", "专业", "考生编号", "姓名", "初试成绩", "复试成绩", "综合成绩", "综合排名", "研究生类型", "学生类型", "手机号"]
    #                 if not all(column in reader.fieldnames for column in required_columns):
    #                     self.message_user(request, "CSV 文件列名不正确，请确保包含：准考证号、姓名、学生类型、研究生类型、学习方式", level='error')
    #                     return redirect('admin:Professor_Student_Manage_student_changelist')

    #                 # 创建学生账号
    #                 success_count = 0
    #                 for row in reader:
    #                     print(row)
    #                     subject_number = row["专业代码"]
    #                     subject_number = str(subject_number).zfill(6)
    #                     subject_name = row["专业"]

    #                     subject = Subject.objects.filter(subject_code=subject_number).first()
    #                     # print(subject)

    #                     candidate_number = str(row["考生编号"]).strip()
    #                     name = row["姓名"]
    #                     initial_exam_score = float(row["初试成绩"])
    #                     secondary_exam_score = float(row["复试成绩"])
    #                     # name = row["综合成绩"]
    #                     final_rank = row["综合排名"]

    #                     postgraduate_type = int(row["研究生类型"])  # 需转换为整数

    #                     student_type = int(row["学生类型"])  # 需转换为整数

    #                     phone_number = str(row["手机号"]).strip()
    #                     # study_mode  == "全日制"  # 转换为布尔值

    #                     try:
    #                         # 检查准考证号是否已存在
    #                         if Student.objects.filter(candidate_number=candidate_number).exists():
    #                             self.message_user(request, f"考生编号 {candidate_number} 已存在，跳过此记录", level='warning')
    #                             continue

    #                         # 创建关联的 User 对象
    #                         username = candidate_number  # 使用准考证号作为用户名
    #                         if User.objects.filter(username=username).exists():
    #                             self.message_user(request, f"用户名 {username} 已存在，跳过此记录", level='warning')
    #                             continue

    #                         user = User.objects.create_user(
    #                             username=username,
    #                             password=phone_number  # 初始密码设置为手机号
    #                         )
                            

    #                         # 创建 Student 对象
    #                         student = Student(
    #                             user_name=user,
    #                             name=name,
    #                             candidate_number=candidate_number,
    #                             subject = subject,
    #                             student_type=student_type,
    #                             postgraduate_type=postgraduate_type,
    #                             phone_number = phone_number,
    #                             initial_exam_score = initial_exam_score,
    #                             secondary_exam_score = secondary_exam_score,
    #                             final_rank = final_rank
    #                             # 其他字段使用默认值
    #                         )
    #                         student.save()
    #                         success_count += 1

    #                     except ValueError as e:
    #                         self.message_user(request, f"准考证号 {candidate_number} 的数据格式不正确: {str(e)}", level='warning')
    #                         continue
    #                     except Exception as e:
    #                         self.message_user(request, f"创建学生 {candidate_number} 时出错: {str(e)}", level='error')
    #                         continue
                        
    #                     # break

    #                 self.message_user(request, f"成功创建 {success_count} 个学生账号")
    #                 return redirect('admin:Professor_Student_Manage_student_changelist')

    #             except Exception as e:
    #                 self.message_user(request, f"解析 CSV 文件时出错: {str(e)}", level='error')
    #                 return redirect('admin:Professor_Student_Manage_student_changelist')
    #     else:
    #         form = ImportStudentForm()

    #     context = {
    #         'form': form,
    #         'opts': self.model._meta,
    #         'title': '一键导入学生账号',
    #     }
    #     return render(request, 'admin/import_students.html', context)
    
    def import_students_view(self, request):
        if request.method == 'POST':
            form = ImportStudentForm(request.POST, request.FILES)
            if form.is_valid():
                import_type = form.cleaned_data['import_type']
                file = request.FILES['file']
                update_quota = form.cleaned_data.get('update_quota', True)
                
                # 根据导入类型调用不同的处理方法
                if import_type == 'doctor':
                    return self._import_doctor_students(request, file, update_quota)
                elif import_type == 'master_exam':
                    return self._import_master_exam_students(request, file, update_quota)
                elif import_type == 'master_recommend':
                    return self._import_master_recommend_students(request, file, update_quota)
        else:
            form = ImportStudentForm()

        context = {
            'form': form,
            'opts': self.model._meta,
            'title': '一键导入学生账号',
        }
        return render(request, 'admin/import_students.html', context)
    
    def _import_doctor_students(self, request, file, update_quota):
        """
        博士生申请导入
        列：考生编号、考生姓名、报考专业代码、复试成绩、排名、备注、递补顺序、手机号
        """
        try:
            import openpyxl
            from collections import defaultdict
            
            # 读取xlsx文件
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            # 读取表头
            headers = [cell.value for cell in ws[1]]
            
            # 查找需要的列索引
            try:
                col_candidate_number = headers.index('考生编号') + 1
                col_name = headers.index('考生姓名') + 1
                col_subject_code = headers.index('报考专业代码') + 1
                col_retest_score = headers.index('复核成绩（满分100分）') + 1
                col_rank = headers.index('排名') + 1
                col_remark = headers.index('备注') + 1
                col_phone = headers.index('手机号') + 1
            except ValueError as e:
                self.message_user(request, f"文件列名不正确，缺少必要列: {str(e)}", level='error')
                return redirect('admin:Professor_Student_Manage_student_changelist')
            
            # 第一遍扫描：统计各专业的总招生人数（备注=选导师的人数）
            subject_quota_map = defaultdict(int)
            rows_data = []
            
            for row in ws.iter_rows(min_row=2, values_only=False):
                if not row[col_candidate_number - 1].value:
                    continue
                    
                candidate_number = str(row[col_candidate_number - 1].value).strip()
                name = str(row[col_name - 1].value).strip() if row[col_name - 1].value else ""
                subject_code = str(row[col_subject_code - 1].value).strip()
                retest_score = row[col_retest_score - 1].value
                rank = int(row[col_rank - 1].value) if row[col_rank - 1].value else 0
                remark = str(row[col_remark - 1].value).strip() if row[col_remark - 1].value else ""
                phone = str(row[col_phone - 1].value).strip() if row[col_phone - 1].value else ""
                
                rows_data.append({
                    'candidate_number': candidate_number,
                    'name': name,
                    'subject_code': subject_code,
                    'retest_score': retest_score,
                    'rank': rank,
                    'remark': remark,
                    'phone': phone
                })
                
                # 统计备注为"选导师"的人数
                if remark == '选导师':
                    subject_quota_map[subject_code] += 1
            
            # 如果勾选了更新专业总招生人数，则更新
            if update_quota:
                for subject_code, quota in subject_quota_map.items():
                    subject = Subject.objects.filter(subject_code=subject_code, subject_type=2).first()  # subject_type=2 表示博士
                    if subject:
                        subject.total_admission_quota = quota
                        subject.save()
                        self.message_user(request, f"已更新专业 {subject.subject_name}({subject_code}) 的总招生人数为 {quota}", level='info')
                    else:
                        self.message_user(request, f"未找到博士专业代码 {subject_code}，跳过更新", level='warning')
            
            # 第二遍处理：创建学生账号
            success_count = 0
            skip_count = 0
            
            for row_data in rows_data:
                try:
                    candidate_number = row_data['candidate_number']
                    name = row_data['name']
                    subject_code = row_data['subject_code']
                    retest_score = row_data['retest_score']
                    rank = row_data['rank']
                    phone = row_data['phone']
                    
                    # 查找专业（必须是博士专业，subject_type=2）
                    subject = Subject.objects.filter(subject_code=subject_code, subject_type=2).first()
                    if not subject:
                        self.message_user(request, f"博士专业代码 {subject_code} 不存在，跳过考生 {candidate_number}", level='warning')
                        skip_count += 1
                        continue
                    
                    # 检查是否已存在
                    if Student.objects.filter(candidate_number=candidate_number).exists():
                        self.message_user(request, f"考生编号 {candidate_number} 已存在，跳过", level='warning')
                        skip_count += 1
                        continue
                    
                    # 创建用户账号
                    username = candidate_number
                    if User.objects.filter(username=username).exists():
                        self.message_user(request, f"用户名 {username} 已存在，跳过", level='warning')
                        skip_count += 1
                        continue
                    
                    user = User.objects.create_user(
                        username=username,
                        password=phone  # 密码为手机号
                    )
                    
                    # 计算是否候补及候补顺序
                    total_quota = subject.total_admission_quota or 0
                    if rank <= total_quota:
                        is_alternate = False
                        alternate_rank = None
                    else:
                        is_alternate = True
                        alternate_rank = rank - total_quota
                    
                    # 创建学生
                    student = Student(
                        user_name=user,
                        name=name,
                        candidate_number=candidate_number,
                        subject=subject,
                        student_type=3,  # 博士统考生
                        postgraduate_type=3,  # 博士
                        study_mode=True,  # 全日制
                        phone_number=phone,
                        secondary_exam_score=float(retest_score) if retest_score else None,
                        final_rank=rank,
                        is_alternate=is_alternate,
                        alternate_rank=alternate_rank
                    )
                    student.save()
                    success_count += 1
                    
                except Exception as e:
                    self.message_user(request, f"处理考生 {row_data.get('candidate_number', 'unknown')} 时出错: {str(e)}", level='error')
                    skip_count += 1
                    continue
            
            self.message_user(request, f"博士生导入完成：成功创建 {success_count} 个账号，跳过 {skip_count} 个", level='success')
            return redirect('admin:Professor_Student_Manage_student_changelist')
            
        except Exception as e:
            self.message_user(request, f"解析文件时出错: {str(e)}", level='error')
            return redirect('admin:Professor_Student_Manage_student_changelist')
    
    def _import_master_exam_students(self, request, file, update_quota):
        """
        硕士统考生导入（暂用原CSV逻辑）
        """
        try:
            # 读取 CSV 文件
            csv_file_wrapper = TextIOWrapper(file, encoding='utf-8-sig')
            reader = csv.DictReader(csv_file_wrapper)

            # 检查列名是否正确
            required_columns = ["专业代码", "专业", "考生编号", "姓名", "初试成绩", "复试成绩",
                                "综合成绩", "综合排名", "研究生类型", "学生类型", "手机号", "身份证号"]
            if not all(column in reader.fieldnames for column in required_columns):
                self.message_user(
                    request,
                    "CSV 文件列名不正确，请确保包含：专业代码、专业、考生编号、姓名、初试成绩、复试成绩、综合成绩、综合排名、研究生类型、学生类型、手机号、身份证号",
                    level='error'
                )
                return redirect('admin:Professor_Student_Manage_student_changelist')

            success_count = 0
            for row in reader:
                subject_number = str(row["专业代码"]).zfill(6)
                subject = Subject.objects.filter(subject_code=subject_number).first()
                if not subject:
                    self.message_user(request, f"学科代码 {subject_number} 不存在，跳过", level='warning')
                    continue

                candidate_number = str(row["考生编号"]).strip()
                name = row["姓名"].strip()
                secondary_exam_score = float(row["综合成绩"])
                final_rank = int(row["综合排名"])
                postgraduate_type = int(row["研究生类型"])
                student_type = int(row["学生类型"])
                phone_number = str(row["手机号"]).strip()
                identity_number = str(row["身份证号"]).strip()

                try:
                    if Student.objects.filter(candidate_number=candidate_number).exists():
                        self.message_user(request, f"考生编号 {candidate_number} 已存在，跳过", level='warning')
                        continue

                    username = candidate_number
                    if User.objects.filter(username=username).exists():
                        self.message_user(request, f"用户名 {username} 已存在，跳过", level='warning')
                        continue

                    user = User.objects.create_user(
                        username=username,
                        password=phone_number
                    )

                    total_quota = subject.total_admission_quota or 0
                    if final_rank <= total_quota:
                        is_alternate = False
                        alternate_rank = None
                    else:
                        is_alternate = True
                        alternate_rank = final_rank - total_quota

                    student = Student(
                        user_name=user,
                        name=name,
                        candidate_number=candidate_number,
                        subject=subject,
                        identify_number=identity_number,
                        student_type=student_type,
                        postgraduate_type=postgraduate_type,
                        phone_number=phone_number,
                        secondary_exam_score=secondary_exam_score,
                        final_rank=final_rank,
                        is_alternate=is_alternate,
                        alternate_rank=alternate_rank,
                    )
                    student.save()
                    success_count += 1

                except ValueError as e:
                    self.message_user(request, f"考生编号 {candidate_number} 的数据格式不正确: {str(e)}", level='warning')
                    continue
                except Exception as e:
                    self.message_user(request, f"创建学生 {candidate_number} 时出错: {str(e)}", level='error')
                    continue

            self.message_user(request, f"硕士统考生导入完成：成功创建 {success_count} 个学生账号")
            return redirect('admin:Professor_Student_Manage_student_changelist')

        except Exception as e:
            self.message_user(request, f"解析 CSV 文件时出错: {str(e)}", level='error')
            return redirect('admin:Professor_Student_Manage_student_changelist')
    
    def _import_master_recommend_students(self, request, file, update_quota):
        """
        硕士推免生导入（暂用原CSV逻辑）
        """
        # 与硕士统考生导入逻辑相同，后续可以自定义
        return self._import_master_exam_students(request, file, update_quota)


    def reset_password_to_exam_id(self, request, queryset):
        """
        将选中学生的密码重置为准考证号（exam_id）
        """
        for student in queryset:
            if student.user_name:  # 确保关联的 User 对象存在
                candidate_number = student.candidate_number
                student.user_name.set_password(candidate_number)  # 重置密码
                student.user_name.save()
                self.message_user(
                    request,
                    f"已重置学生 {student.name} 的密码为准考证号: {candidate_number}",
                    level='success'
                )

    reset_password_to_exam_id.short_description = "重置密码为准考证号"  # 动作显示名称

    # def download_fq_file(self, obj):
    #     """
    #     若学生已放弃拟录取并且 hx_file 有文件，则显示下载链接；否则显示 '-'
    #     """
    #     if obj.is_giveup == True:

    #         # 获取下载地址
    #         response_data_signature = self.get_fileid_download_url(obj.giveup_signature_table)
    #         if response_data_signature.get("errcode") == 0:
    #             signature_download_url = response_data_signature['file_list'][0]['download_url']
    #             print(f"放弃说明表下载地址: {signature_download_url}")
    #         else:
    #             return Response({'message': '获取放弃说明表下载地址失败'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    #         return format_html(
    #             "<a href='{}' download>下载</a>", signature_download_url
    #         )
    #     return '未完成'

    # def download_hx_file(self, obj):
    #     """
    #     若学生已放弃拟录取并且 hx_file 有文件，则显示下载链接；否则显示 '-'
    #     """
    #     if obj.is_selected == True and obj.signature_table_review_status == 1:

    #         # 获取下载地址
    #         response_data_signature = self.get_fileid_download_url(obj.signature_table)
    #         if response_data_signature.get("errcode") == 0:
    #             signature_download_url = response_data_signature['file_list'][0]['download_url']
    #             print(f"签名图片下载地址: {signature_download_url}")
    #         else:
    #             return Response({'message': '获取签名图片下载地址失败'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    #         return format_html(
    #             "<a href='{}' download>下载</a>", signature_download_url
    #         )
    #     return '未完成'

    def get_download_url_view(self, request):
        """
        AJAX 视图，用于根据文件 ID 获取下载 URL
        """
        print("url访问成功")
        if request.method == 'POST':
            file_id = request.POST.get('file_id')
            if not file_id:
                return JsonResponse({'error': '未提供文件 ID'}, status=400)

            response_data = self.get_fileid_download_url(file_id)
            if response_data.get("errcode") == 0:
                download_url = response_data['file_list'][0]['download_url']
                return JsonResponse({'download_url': download_url})
            else:
                return JsonResponse({'error': '获取下载 URL 失败'}, status=500)

        return JsonResponse({'error': '无效的请求方法'}, status=405)

    def download_hx_file(self, obj):
        """
        如果满足条件，显示下载互选表的按钮
        """
        if obj.is_selected and obj.signature_table_review_status == 1 and obj.signature_table:
            return format_html(
                '<button class="download-btn" data-file-id="{}" data-type="hx">下载</button>',
                obj.signature_table
            )
        return '未完成'

    def download_fq_file(self, obj):
        """
        如果满足条件，显示下载弃选表的按钮
        """
        if obj.is_giveup and obj.giveup_signature_table:
            return format_html(
                '<button class="download-btn" data-file-id="{}" data-type="fq">下载</button>',
                obj.giveup_signature_table
            )
        return '未完成'

    def get_fileid_download_url(self, file_id):
        """
        根据 file_id 获取下载地址
        """
        access_token = get_wechat_access_token()
        url = f'{WECHAT_API_BASE}/tcb/batchdownloadfile?access_token={access_token}'
        data = {
            "env": WECHAT_CLOUD_ENV,
            "file_list": [
                {
                    "fileid": file_id,
                    "max_age":7200
                }
            ]
        }

        # 发送POST请求
        response = requests.post(url, json=data, timeout=15)
        response.raise_for_status()
        payload = response.json()
        if payload.get('errcode') == 41001:
            access_token = get_wechat_access_token(force_refresh=True)
            refresh_url = f'{WECHAT_API_BASE}/tcb/batchdownloadfile?access_token={access_token}'
            response = requests.post(refresh_url, json=data, timeout=15)
            response.raise_for_status()
            payload = response.json()
        return payload

    # change_list_template = 'admin/student_change_list.html'  # 自定义列表页面模板

    download_hx_file.short_description = "互选表下载"
    download_fq_file.short_description = "弃选表下载"


class WeChatAccountAdmin(admin.ModelAdmin):
    list_display = ["user", "openid", "session_key"]


admin.site.register(Student, StudentAdmin)
admin.site.register(Professor, ProfessorAdmin)
admin.site.register(WeChatAccount, WeChatAccountAdmin)
