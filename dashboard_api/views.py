import os
import io
import zipfile
import csv
from decimal import Decimal
from datetime import datetime, timedelta
from io import TextIOWrapper

import requests
from django.contrib.auth.models import User
from django.contrib.auth import authenticate
from django.core.exceptions import ValidationError
from django.core.cache import cache
from django.http import HttpResponse
from django.db import transaction
from django.db.models import Model
from django.db.models import Count, OuterRef, Prefetch, Q, Subquery, Sum, Value
from django.db.models.functions import Coalesce, TruncDate
from django.utils import timezone
from rest_framework import status
from rest_framework.authentication import TokenAuthentication
from rest_framework.authtoken.models import Token
from rest_framework.permissions import AllowAny
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView
from openpyxl import Workbook, load_workbook

from Enrollment_Manage.models import Department, Subject, sync_student_alternate_status
from Professor_Student_Manage.models import (
    AvailableStudentDisplaySetting,
    AdmissionBatch,
    get_professor_heat_display_metrics,
    get_available_student_display_setting,
    normalize_available_student_display_values,
    get_professor_heat_display_setting,
    Professor,
    ProfessorHeatDisplaySetting,
    ProfessorDoctorQuota,
    ProfessorMasterQuota,
    ProfessorSharedQuotaPool,
    Student,
    get_default_admission_year,
    get_matching_shared_quota_pool,
    get_quota_source_for_student,
    recalculate_professor_quota_summary,
)
from Professor_Student_Manage.models import WeChatAccount
from Select_Information.models import ReviewRecord, StudentProfessorChoice
from Select_Information.models import SelectionTime

from .models import DashboardAuditLog
from .permissions import IsDashboardAdmin
from .serializers import (
    AdmissionBatchSerializer,
    AvailableStudentDisplaySettingSerializer,
    AlternateStudentSerializer,
    ChoiceListSerializer,
    DashboardAdminSerializer,
    DashboardAuditLogDetailSerializer,
    DashboardAuditLogListSerializer,
    DashboardUserSerializer,
    DepartmentListSerializer,
    GiveupStudentSerializer,
    DashboardTokenSerializer,
    ProfessorDetailSerializer,
    ProfessorHeatDisplaySettingSerializer,
    ProfessorHeatListSerializer,
    ProfessorDoctorQuotaSerializer,
    ProfessorListSerializer,
    ProfessorMasterQuotaSerializer,
    ReviewRecordDetailSerializer,
    ReviewRecordListSerializer,
    SelectionTimeSerializer,
    SharedQuotaPoolSerializer,
    SubjectListSerializer,
    StudentDetailSerializer,
    StudentListSerializer,
    WeChatAccountSerializer,
)


WECHAT_CLOUD_ENV = os.environ.get('WECHAT_CLOUD_ENV', 'prod-2g1jrmkk21c1d283')
WECHAT_APPID = os.environ.get('WECHAT_APPID', 'wxa67ae78c4f1f6275')
WECHAT_SECRET = os.environ.get('WECHAT_SECRET', '7241b1950145a193f15b3584d50f3989')


def get_display_name_for_user(user):
    if not user:
        return ''
    full_name = f'{user.first_name} {user.last_name}'.strip()
    return full_name or user.username


def get_request_ip(request):
    forwarded = request.META.get('HTTP_X_FORWARDED_FOR', '')
    if forwarded:
        return forwarded.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR', '')


def serialize_audit_value(value):
    if isinstance(value, dict):
        return {str(key): serialize_audit_value(item) for key, item in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [serialize_audit_value(item) for item in value]
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return timezone.localtime(value).isoformat() if timezone.is_aware(value) else value.isoformat()
    if hasattr(value, 'isoformat') and not isinstance(value, str):
        try:
            return value.isoformat()
        except Exception:
            pass
    if isinstance(value, Model):
        return {'id': value.pk, 'display': str(value)}
    return value


def parse_request_bool(value, default=False):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in ('1', 'true', 'yes', 'on')


def snapshot_instance(instance, extra=None):
    if instance is None:
        return None

    data = {}
    for field in instance._meta.concrete_fields:
        if getattr(field, 'many_to_one', False) or getattr(field, 'one_to_one', False):
            data[field.name] = getattr(instance, f'{field.name}_id', None)
        else:
            data[field.name] = serialize_audit_value(getattr(instance, field.name))
    data['_display'] = str(instance)
    if extra:
        data.update(serialize_audit_value(extra))
    return data


def create_audit_log(
    request,
    *,
    action,
    module,
    level=DashboardAuditLog.LEVEL_INFO,
    status_value=DashboardAuditLog.STATUS_SUCCESS,
    operator=None,
    operator_username='',
    target_type='',
    target_id='',
    target_display='',
    detail='',
    before_data=None,
    after_data=None,
):
    if operator is None:
        operator = getattr(request, 'user', None)
        if operator is not None and getattr(operator, 'is_authenticated', False) is False:
            operator = None
    if not operator_username:
        operator_username = operator.username if operator else ''

    DashboardAuditLog.objects.create(
        operator=operator,
        operator_username=operator_username,
        action=action,
        module=module,
        level=level,
        status=status_value,
        target_type=target_type,
        target_id=str(target_id or ''),
        target_display=target_display or '',
        detail=detail or '',
        before_data=serialize_audit_value(before_data),
        after_data=serialize_audit_value(after_data),
        request_method=request.method,
        request_path=request.path,
        ip_address=get_request_ip(request),
        user_agent=request.META.get('HTTP_USER_AGENT', ''),
    )


def build_audit_log_queryset(request):
    queryset = DashboardAuditLog.objects.select_related('operator')

    search = request.query_params.get('search', '').strip()
    module = request.query_params.get('module', '').strip()
    action = request.query_params.get('action', '').strip()
    status_value = request.query_params.get('status', '').strip()
    level = request.query_params.get('level', '').strip()
    operator_id = request.query_params.get('operator_id', '').strip()
    date_from = request.query_params.get('date_from', '').strip()
    date_to = request.query_params.get('date_to', '').strip()

    if search:
        queryset = queryset.filter(
            Q(operator_username__icontains=search) |
            Q(target_display__icontains=search) |
            Q(detail__icontains=search) |
            Q(request_path__icontains=search)
        )
    if module:
        queryset = queryset.filter(module=module)
    if action:
        queryset = queryset.filter(action=action)
    if status_value:
        queryset = queryset.filter(status=status_value)
    if level:
        queryset = queryset.filter(level=level)
    if operator_id:
        queryset = queryset.filter(operator_id=operator_id)
    if date_from:
        queryset = queryset.filter(created_at__date__gte=date_from)
    if date_to:
        queryset = queryset.filter(created_at__date__lte=date_to)
    return queryset


def get_wechat_access_token(force_refresh=False):
    cache_key = 'dashboard_wechat_access_token'
    if not force_refresh:
        cached_token = cache.get(cache_key)
        if cached_token:
            return cached_token

    response = requests.get(
        'https://api.weixin.qq.com/cgi-bin/token',
        params={
            'grant_type': 'client_credential',
            'appid': WECHAT_APPID,
            'secret': WECHAT_SECRET,
        },
        timeout=15,
    )
    response.raise_for_status()
    payload = response.json()
    access_token = payload.get('access_token')
    if not access_token:
        raise ValueError(payload.get('errmsg') or 'Failed to fetch access token.')

    expires_in = int(payload.get('expires_in') or 7200)
    cache.set(cache_key, access_token, timeout=max(expires_in - 300, 300))
    return access_token


def get_file_download_url(file_id):
    access_token = get_wechat_access_token()
    response = requests.post(
        f'https://api.weixin.qq.com/tcb/batchdownloadfile?access_token={access_token}',
        json={
            'env': WECHAT_CLOUD_ENV,
            'file_list': [{'fileid': file_id, 'max_age': 7200}],
        },
        timeout=15,
    )
    response.raise_for_status()
    payload = response.json()
    if payload.get('errcode') == 41001:
        access_token = get_wechat_access_token(force_refresh=True)
        response = requests.post(
            f'https://api.weixin.qq.com/tcb/batchdownloadfile?access_token={access_token}',
            json={
                'env': WECHAT_CLOUD_ENV,
                'file_list': [{'fileid': file_id, 'max_age': 7200}],
            },
            timeout=15,
        )
        response.raise_for_status()
        payload = response.json()
    if payload.get('errcode') not in (None, 0):
        raise ValueError(payload.get('errmsg') or 'Failed to fetch download url.')
    file_list = payload.get('file_list') or []
    if not file_list or not file_list[0].get('download_url'):
        raise ValueError('Download url is missing in WeChat cloud response.')
    return file_list[0]['download_url']


def professor_has_quota_for_student(professor, student):
    _, quota_source = get_quota_source_for_student(professor, student, remaining_only=True)
    return quota_source is not None


def apply_department_used_quota_delta(department, student, delta):
    if not department:
        return
    if student.postgraduate_type == 3:
        department.used_doctor_quota = max(0, (department.used_doctor_quota or 0) + delta)
        department.save(update_fields=['used_doctor_quota'])
    elif student.postgraduate_type == 2:
        department.used_academic_quota = max(0, (department.used_academic_quota or 0) + delta)
        department.save(update_fields=['used_academic_quota'])
    elif student.postgraduate_type == 1:
        department.used_professional_quota = max(0, (department.used_professional_quota or 0) + delta)
        department.save(update_fields=['used_professional_quota'])
    elif student.postgraduate_type == 4:
        department.used_professional_yt_quota = max(0, (department.used_professional_yt_quota or 0) + delta)
        department.save(update_fields=['used_professional_yt_quota'])


def consume_quota_for_choice(choice):
    student = choice.student
    professor = choice.professor
    department = professor.department

    if choice.shared_quota_pool_id:
        shared_pool = ProfessorSharedQuotaPool.objects.filter(
            pk=choice.shared_quota_pool_id,
            professor=professor,
            is_active=True,
        ).first()
        if not shared_pool or (shared_pool.remaining_quota or 0) <= 0:
            raise ValidationError('原共享名额池已无可用名额，无法恢复该录取。')
        shared_pool.remaining_quota -= 1
        shared_pool.used_quota += 1
        shared_pool.save(update_fields=['remaining_quota', 'used_quota', 'updated_at'])
        apply_department_used_quota_delta(department, student, 1)
        recalculate_professor_quota_summary(professor)
        return

    if student.postgraduate_type == 3:
        quota = ProfessorDoctorQuota.objects.filter(professor=professor, subject=student.subject).first()
        if not quota or (quota.remaining_quota or 0) <= 0:
            raise ValidationError('导师在该博士专业下已无可用名额，无法恢复该录取。')
        quota.remaining_quota -= 1
        quota.used_quota += 1
        quota.save(update_fields=['remaining_quota', 'used_quota'])
        apply_department_used_quota_delta(department, student, 1)
        recalculate_professor_quota_summary(professor)
        return

    quota = ProfessorMasterQuota.objects.filter(professor=professor, subject=student.subject).first()
    if not quota:
        raise ValidationError('导师在该硕士专业下不存在名额配置，无法恢复该录取。')
    update_fields = []
    if student.postgraduate_type in [1, 2]:
        if (quota.beijing_remaining_quota or 0) <= 0:
            raise ValidationError('导师在该硕士专业北京校区已无可用名额，无法恢复该录取。')
        quota.beijing_remaining_quota -= 1
        update_fields.append('beijing_remaining_quota')
    elif student.postgraduate_type == 4:
        if (quota.yantai_remaining_quota or 0) <= 0:
            raise ValidationError('导师在该硕士专业烟台校区已无可用名额，无法恢复该录取。')
        quota.yantai_remaining_quota -= 1
        update_fields.append('yantai_remaining_quota')
    if update_fields:
        quota.save(update_fields=update_fields)
    apply_department_used_quota_delta(department, student, 1)
    recalculate_professor_quota_summary(professor)


def reinstate_revoked_choice(choice):
    consume_quota_for_choice(choice)

    student = choice.student
    choice.status = 1
    choice.finish_time = timezone.now()
    choice.save(update_fields=['status', 'finish_time'])

    student.is_selected = True
    student.is_giveup = False
    student.signature_table_student_signatured = False
    student.signature_table_professor_signatured = False
    student.signature_table_review_status = 4
    student.save(
        update_fields=[
            'is_selected',
            'is_giveup',
            'signature_table_student_signatured',
            'signature_table_professor_signatured',
            'signature_table_review_status',
        ]
    )


def normalize_alternate_ranks(subject, admission_year=None):
    changed = 0
    base_queryset = Student.objects.filter(
        subject=subject,
        is_alternate=True,
        is_giveup=False,
    )
    if admission_year not in (None, ''):
        admission_years = [admission_year]
    else:
        admission_years = list(
            base_queryset.exclude(admission_year__isnull=True)
            .order_by('admission_year')
            .values_list('admission_year', flat=True)
            .distinct()
        )
        if not admission_years and base_queryset.exists():
            admission_years = [None]

    for year in admission_years:
        alternates = base_queryset
        if year is None:
            alternates = alternates.filter(admission_year__isnull=True)
        else:
            alternates = alternates.filter(admission_year=year)
        alternates = alternates.order_by('alternate_rank', 'final_rank', 'id')

        for index, student in enumerate(alternates, start=1):
            if student.alternate_rank != index:
                student.alternate_rank = index
                student.save(update_fields=['alternate_rank'])
                changed += 1
    return changed


def promote_next_alternate(subject, require_available_quota=False, admission_year=None):
    student_queryset = Student.objects.filter(subject=subject, is_alternate=True, is_giveup=False)
    if admission_year not in (None, ''):
        student_queryset = student_queryset.filter(admission_year=admission_year)
    student = student_queryset.order_by('alternate_rank', 'final_rank', 'id').first()
    if not student:
        return None, 'missing'
    if require_available_quota:
        has_available_quota = any(
            professor_has_quota_for_student(professor, student)
            for professor in Professor.objects.filter(have_qualification=True).iterator()
        )
        if not has_available_quota:
            return None, 'no_quota'
    student.is_alternate = False
    student.alternate_rank = None
    student.save(update_fields=['is_alternate', 'alternate_rank'])
    normalize_alternate_ranks(subject, student.admission_year if admission_year in (None, '') else admission_year)
    return student, None


def cancel_waiting_choices_for_giveup_students():
    waiting_choices = StudentProfessorChoice.objects.select_related('student').filter(status=3)
    cancelled_count = 0
    for choice in waiting_choices:
        if choice.student and choice.student.is_giveup:
            choice.status = 4
            choice.finish_time = timezone.now()
            choice.save(update_fields=['status', 'finish_time'])
            cancelled_count += 1
    return cancelled_count


def reject_waiting_choices_without_quota():
    waiting_choices = StudentProfessorChoice.objects.select_related('student', 'student__subject', 'professor').filter(status=3)
    rejected_count = 0
    for choice in waiting_choices:
        if not professor_has_quota_for_student(choice.professor, choice.student):
            choice.status = 2
            choice.finish_time = timezone.now()
            choice.save(update_fields=['status', 'finish_time'])
            rejected_count += 1
    return rejected_count


def cancel_approved_choice(choice):
    student = choice.student
    professor = choice.professor
    department = professor.department

    if choice.shared_quota_pool_id:
        shared_pool = choice.shared_quota_pool
        shared_pool.used_quota = max(0, (shared_pool.used_quota or 0) - 1)
        shared_pool.remaining_quota = (shared_pool.remaining_quota or 0) + 1
        shared_pool.save(update_fields=['used_quota', 'remaining_quota', 'updated_at'])
        if student.postgraduate_type == 3:
            department.used_doctor_quota = max(0, (department.used_doctor_quota or 0) - 1)
            department.save(update_fields=['used_doctor_quota'])
        elif student.postgraduate_type == 2:
            department.used_academic_quota = max(0, (department.used_academic_quota or 0) - 1)
            department.save(update_fields=['used_academic_quota'])
        elif student.postgraduate_type == 1:
            department.used_professional_quota = max(0, (department.used_professional_quota or 0) - 1)
            department.save(update_fields=['used_professional_quota'])
        elif student.postgraduate_type == 4:
            department.used_professional_yt_quota = max(0, (department.used_professional_yt_quota or 0) - 1)
            department.save(update_fields=['used_professional_yt_quota'])
        recalculate_professor_quota_summary(professor)
    elif student.postgraduate_type == 3:
        quota = ProfessorDoctorQuota.objects.filter(professor=professor, subject=student.subject).first()
        if quota:
            quota.used_quota = max(0, (quota.used_quota or 0) - 1)
            quota.remaining_quota = (quota.remaining_quota or 0) + 1
            quota.save(update_fields=['used_quota', 'remaining_quota'])
        if department:
            department.used_doctor_quota = max(0, (department.used_doctor_quota or 0) - 1)
            department.save(update_fields=['used_doctor_quota'])
    else:
        quota = ProfessorMasterQuota.objects.filter(professor=professor, subject=student.subject).first()
        if quota:
            update_fields = []
            if student.postgraduate_type == 1:
                quota.beijing_remaining_quota = (quota.beijing_remaining_quota or 0) + 1
                update_fields.append('beijing_remaining_quota')
                if department:
                    department.used_professional_quota = max(0, (department.used_professional_quota or 0) - 1)
                    department.save(update_fields=['used_professional_quota'])
            elif student.postgraduate_type == 2:
                quota.beijing_remaining_quota = (quota.beijing_remaining_quota or 0) + 1
                update_fields.append('beijing_remaining_quota')
                if department:
                    department.used_academic_quota = max(0, (department.used_academic_quota or 0) - 1)
                    department.save(update_fields=['used_academic_quota'])
            elif student.postgraduate_type == 4:
                quota.yantai_remaining_quota = (quota.yantai_remaining_quota or 0) + 1
                update_fields.append('yantai_remaining_quota')
                if department:
                    department.used_professional_yt_quota = max(0, (department.used_professional_yt_quota or 0) - 1)
                    department.save(update_fields=['used_professional_yt_quota'])
            if update_fields:
                quota.save(update_fields=update_fields)

    recalculate_professor_quota_summary(professor)

    choice.status = 5
    choice.finish_time = timezone.now()
    choice.save(update_fields=['status', 'finish_time'])

    student.is_selected = False
    student.signature_table_student_signatured = False
    student.signature_table_professor_signatured = False
    student.signature_table_review_status = 4
    student.save(
        update_fields=[
            'is_selected',
            'signature_table_student_signatured',
            'signature_table_professor_signatured',
            'signature_table_review_status',
        ]
    )
    ReviewRecord.objects.filter(student=student, professor=professor).delete()


def to_int(value, default=0):
    try:
        if value in (None, ''):
            return default
        return int(str(value).strip())
    except (TypeError, ValueError):
        return default


def normalize_header_name(value):
    return str(value or '').replace('\ufeff', '').replace(' ', '').replace('\n', '').replace('\r', '').strip()


def resolve_header_name(headers, aliases):
    normalized_headers = {normalize_header_name(header): header for header in headers if header is not None}
    for alias in aliases:
        normalized_alias = normalize_header_name(alias)
        if normalized_alias in normalized_headers:
            return normalized_headers[normalized_alias]
    for alias in aliases:
        normalized_alias = normalize_header_name(alias)
        for normalized_header, raw_header in normalized_headers.items():
            if normalized_alias in normalized_header or normalized_header in normalized_alias:
                return raw_header
    return None


def get_csv_value(row, aliases, default=''):
    header = resolve_header_name(row.keys(), aliases)
    if header is None:
        return default
    value = row.get(header, default)
    if value is None:
        return default
    return value


def normalize_subject_code(value):
    subject_code = str(value or '').strip()
    if subject_code.isdigit() and len(subject_code) < 6:
        subject_code = subject_code.zfill(6)
    return subject_code


def refresh_professor_summary_quotas(professor):
    recalculate_professor_quota_summary(professor)


def get_uploaded_file(request):
    return request.FILES.get('file') or request.FILES.get('csv_file') or request.FILES.get('xlsx_file')


def parse_bool_param(value):
    if isinstance(value, bool):
        return value
    if value in (None, ''):
        return None
    normalized = str(value).strip().lower()
    if normalized in {'true', '1', 'yes', 'on'}:
        return True
    if normalized in {'false', '0', 'no', 'off'}:
        return False
    return None


def get_excel_value(row, header_map, header, default=''):
    if header is None:
        return default
    index = header_map.get(header)
    if index is None or index >= len(row):
        return default
    value = row[index]
    return default if value is None else value


def normalize_marker_value(value):
    return str(value or '').strip()


def is_blank_marker(value):
    return normalize_marker_value(value) in {'', '\\'}


def is_empty_quota_cell(value):
    return normalize_marker_value(value) == ''


def find_master_subject_by_code(subject_code):
    return Subject.objects.filter(subject_code=subject_code, subject_type__in=[0, 1]).first()


def add_incremental_master_quota(professor, subject, beijing_quota=0, yantai_quota=0):
    quota, _ = ProfessorMasterQuota.objects.get_or_create(professor=professor, subject=subject)
    quota.beijing_quota = (quota.beijing_quota or 0) + max(0, beijing_quota)
    quota.yantai_quota = (quota.yantai_quota or 0) + max(0, yantai_quota)
    quota.save()
    return quota


def get_or_create_incremental_shared_pool(professor, subjects, quota_value, campus):
    subject_ids = sorted(subject.id for subject in subjects)
    pools = (
        ProfessorSharedQuotaPool.objects.filter(
            professor=professor,
            quota_scope=ProfessorSharedQuotaPool.SCOPE_MASTER,
            campus=campus,
            is_active=True,
        )
        .prefetch_related('subjects')
    )

    for pool in pools:
        if sorted(pool.subjects.values_list('id', flat=True)) == subject_ids:
            pool.total_quota = (pool.total_quota or 0) + quota_value
            pool.remaining_quota = min(pool.total_quota, (pool.remaining_quota or 0) + quota_value)
            pool.save()
            return pool

    pool = ProfessorSharedQuotaPool.objects.create(
        professor=professor,
        pool_name='、'.join(subject.subject_name for subject in subjects) + '共享名额',
        quota_scope=ProfessorSharedQuotaPool.SCOPE_MASTER,
        campus=campus,
        total_quota=quota_value,
        used_quota=0,
        remaining_quota=quota_value,
        is_active=True,
    )
    pool.subjects.set(subject_ids)
    return pool


def sync_master_subject_total_quota(subject):
    fixed_totals = ProfessorMasterQuota.objects.filter(subject=subject).aggregate(
        bj_total=Coalesce(Sum('beijing_quota'), 0),
        yt_total=Coalesce(Sum('yantai_quota'), 0),
    )
    shared_total = ProfessorSharedQuotaPool.objects.filter(
        quota_scope=ProfessorSharedQuotaPool.SCOPE_MASTER,
        is_active=True,
        subjects=subject,
    ).aggregate(total=Coalesce(Sum('total_quota'), 0))['total'] or 0
    subject.total_admission_quota = (
        (fixed_totals.get('bj_total') or 0)
        + (fixed_totals.get('yt_total') or 0)
        + shared_total
    )
    subject.save(update_fields=['total_admission_quota'])
    sync_student_alternate_status(subject)


def record_subject_import_summary(summary_map, subject, beijing_delta=0, yantai_delta=0, shared_beijing_delta=0, shared_yantai_delta=0):
    if not subject:
        return
    item = summary_map.setdefault(
        subject.id,
        {
            'subject_id': subject.id,
            'subject_name': subject.subject_name,
            'subject_code': subject.subject_code,
            'beijing_quota': 0,
            'yantai_quota': 0,
            'shared_beijing_quota': 0,
            'shared_yantai_quota': 0,
            'total_quota': 0,
        },
    )
    item['beijing_quota'] += max(0, beijing_delta or 0)
    item['yantai_quota'] += max(0, yantai_delta or 0)
    item['shared_beijing_quota'] += max(0, shared_beijing_delta or 0)
    item['shared_yantai_quota'] += max(0, shared_yantai_delta or 0)
    item['total_quota'] = (
        item['beijing_quota']
        + item['yantai_quota']
        + item['shared_beijing_quota']
        + item['shared_yantai_quota']
    )


def ensure_professor_for_master_quota_import(teacher_id, professor_name, subject_candidates):
    professor = Professor.objects.filter(teacher_identity_id=teacher_id).first()
    if professor:
        if professor_name and professor.name != professor_name:
            professor.name = professor_name
            professor.save(update_fields=['name'])
        return professor, False

    department = None
    for subject in subject_candidates:
        department = subject.subject_department.order_by('id').first()
        if department:
            break
    if department is None:
        raise ValidationError(f'工号 {teacher_id} 对应导师不存在，且无法从专业关联中推断所属招生方向。')

    user = User.objects.filter(username=teacher_id).first()
    if user is None:
        user = User.objects.create_user(username=teacher_id, password=teacher_id)

    professor = Professor.objects.create(
        user_name=user,
        name=professor_name or f'导师{teacher_id}',
        teacher_identity_id=teacher_id,
        department=department,
        proposed_quota_approved=True,
        have_qualification=True,
    )
    return professor, True


def import_master_quota_workbook(upload, sync_quotas=False):
    workbook = load_workbook(upload, data_only=True)
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    header_map = {header: index for index, header in enumerate(headers)}

    teacher_id_header = resolve_header_name(headers, ['工号'])
    if teacher_id_header is None:
        raise ValidationError('未找到工号列。')
    name_header = resolve_header_name(headers, ['姓名', '导师姓名'])

    subject1_code_header = resolve_header_name(headers, ['学科1代码'])
    subject1_quota_header = resolve_header_name(headers, ['名额1', '本次名额1'])
    subject2_code1_header = resolve_header_name(headers, ['学科2代码1'])
    subject2_code2_header = resolve_header_name(headers, ['学科2代码2'])
    subject2_beijing_header = resolve_header_name(headers, ['北京名额2', '北京招生名额2'])
    subject2_yantai_header = resolve_header_name(headers, ['烟台名额2', '烟台招生名额2'])
    subject3_code_header = resolve_header_name(headers, ['学科3代码'])
    subject3_beijing_header = resolve_header_name(headers, ['北京名额3', '北京招生名额3'])
    subject3_yantai_header = resolve_header_name(headers, ['烟台名额3', '烟台招生名额3'])
    subject4_code_header = resolve_header_name(headers, ['学科4代码'])
    subject4_quota_header = resolve_header_name(headers, ['名额4', '本次名额4'])

    updated_count = 0
    skipped_rows = 0
    touched_professors = set()
    touched_subjects = set()
    created_professor_ids = []
    subject_summary_map = {}

    with transaction.atomic():
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if all(normalize_marker_value(value) == '' for value in row):
                continue

            teacher_id = normalize_marker_value(get_excel_value(row, header_map, teacher_id_header)).zfill(5)
            professor_name = normalize_marker_value(get_excel_value(row, header_map, name_header))
            if not teacher_id:
                skipped_rows += 1
                continue

            row_imported = False

            subject1_code = normalize_subject_code(get_excel_value(row, header_map, subject1_code_header))
            subject1_quota_raw = get_excel_value(row, header_map, subject1_quota_header)
            subject1 = find_master_subject_by_code(subject1_code) if subject1_code else None

            subject2_code1 = normalize_subject_code(get_excel_value(row, header_map, subject2_code1_header))
            subject2_code2_raw = get_excel_value(row, header_map, subject2_code2_header)
            subject2_code2 = normalize_subject_code(subject2_code2_raw)
            subject2_main = find_master_subject_by_code(subject2_code1) if subject2_code1 else None
            subject2_shared = (
                find_master_subject_by_code(subject2_code2)
                if subject2_code2 and not is_blank_marker(subject2_code2_raw)
                else None
            )
            subject2_beijing_raw = get_excel_value(row, header_map, subject2_beijing_header)
            subject2_yantai_raw = get_excel_value(row, header_map, subject2_yantai_header)

            subject3_code = normalize_subject_code(get_excel_value(row, header_map, subject3_code_header))
            subject3_beijing_raw = get_excel_value(row, header_map, subject3_beijing_header)
            subject3_yantai_raw = get_excel_value(row, header_map, subject3_yantai_header)
            subject3 = find_master_subject_by_code(subject3_code) if subject3_code else None

            subject4_code = normalize_subject_code(get_excel_value(row, header_map, subject4_code_header))
            subject4_quota_raw = get_excel_value(row, header_map, subject4_quota_header)
            subject4 = find_master_subject_by_code(subject4_code) if subject4_code else None

            subject_candidates = [subject for subject in [subject1, subject2_main, subject2_shared, subject3, subject4] if subject]
            if not subject_candidates:
                skipped_rows += 1
                continue

            professor, professor_created = ensure_professor_for_master_quota_import(teacher_id, professor_name, subject_candidates)
            if professor_created:
                created_professor_ids.append(teacher_id)

            if subject1_code and not is_empty_quota_cell(subject1_quota_raw):
                if subject1:
                    quota_value = to_int(subject1_quota_raw)
                    if quota_value > 0:
                        add_incremental_master_quota(professor, subject1, beijing_quota=quota_value)
                        record_subject_import_summary(subject_summary_map, subject1, beijing_delta=quota_value)
                        touched_subjects.add(subject1.id)
                        updated_count += 1
                        row_imported = True

            if subject2_code1 and not (is_empty_quota_cell(subject2_beijing_raw) and is_empty_quota_cell(subject2_yantai_raw)):
                if subject2_main:
                    beijing_quota = to_int(subject2_beijing_raw)
                    yantai_quota = to_int(subject2_yantai_raw)
                    if not is_blank_marker(subject2_code2_raw):
                        if subject2_shared:
                            if beijing_quota > 0:
                                get_or_create_incremental_shared_pool(
                                    professor,
                                    [subject2_main, subject2_shared],
                                    beijing_quota,
                                    ProfessorSharedQuotaPool.CAMPUS_BEIJING,
                                )
                                record_subject_import_summary(subject_summary_map, subject2_main, shared_beijing_delta=beijing_quota)
                                record_subject_import_summary(subject_summary_map, subject2_shared, shared_beijing_delta=beijing_quota)
                                touched_subjects.add(subject2_main.id)
                                touched_subjects.add(subject2_shared.id)
                                updated_count += 1
                                row_imported = True
                            if yantai_quota > 0:
                                get_or_create_incremental_shared_pool(
                                    professor,
                                    [subject2_main, subject2_shared],
                                    yantai_quota,
                                    ProfessorSharedQuotaPool.CAMPUS_YANTAI,
                                )
                                record_subject_import_summary(subject_summary_map, subject2_main, shared_yantai_delta=yantai_quota)
                                record_subject_import_summary(subject_summary_map, subject2_shared, shared_yantai_delta=yantai_quota)
                                touched_subjects.add(subject2_main.id)
                                touched_subjects.add(subject2_shared.id)
                                updated_count += 1
                                row_imported = True
                    else:
                        if beijing_quota > 0:
                            add_incremental_master_quota(professor, subject2_main, beijing_quota=beijing_quota)
                            record_subject_import_summary(subject_summary_map, subject2_main, beijing_delta=beijing_quota)
                            touched_subjects.add(subject2_main.id)
                            updated_count += 1
                            row_imported = True
                        if yantai_quota > 0:
                            add_incremental_master_quota(professor, subject2_main, yantai_quota=yantai_quota)
                            record_subject_import_summary(subject_summary_map, subject2_main, yantai_delta=yantai_quota)
                            touched_subjects.add(subject2_main.id)
                            updated_count += 1
                            row_imported = True

            if subject3_code and not (is_empty_quota_cell(subject3_beijing_raw) and is_empty_quota_cell(subject3_yantai_raw)):
                if subject3:
                    beijing_quota = to_int(subject3_beijing_raw)
                    yantai_quota = to_int(subject3_yantai_raw)
                    if beijing_quota > 0 or yantai_quota > 0:
                        add_incremental_master_quota(
                            professor,
                            subject3,
                            beijing_quota=beijing_quota,
                            yantai_quota=yantai_quota,
                        )
                        record_subject_import_summary(
                            subject_summary_map,
                            subject3,
                            beijing_delta=beijing_quota,
                            yantai_delta=yantai_quota,
                        )
                        touched_subjects.add(subject3.id)
                        updated_count += 1
                        row_imported = True

            if subject4_code and not is_empty_quota_cell(subject4_quota_raw):
                if subject4:
                    quota_value = to_int(subject4_quota_raw)
                    if quota_value > 0:
                        add_incremental_master_quota(professor, subject4, beijing_quota=quota_value)
                        record_subject_import_summary(subject_summary_map, subject4, beijing_delta=quota_value)
                        touched_subjects.add(subject4.id)
                        updated_count += 1
                        row_imported = True

            if row_imported:
                touched_professors.add(professor.id)
            else:
                skipped_rows += 1

        for professor in Professor.objects.filter(id__in=touched_professors):
            refresh_professor_summary_quotas(professor)

        if sync_quotas:
            for subject in Subject.objects.filter(id__in=touched_subjects):
                sync_master_subject_total_quota(subject)

    return {
        'detail': f'硕士名额导入完成，更新 {updated_count} 条记录，跳过 {skipped_rows} 条。',
        'updated_count': updated_count,
        'skipped_rows': skipped_rows,
        'created_professor_count': len(created_professor_ids),
        'created_professor_teacher_ids': created_professor_ids,
        'subject_quota_summary': sorted(subject_summary_map.values(), key=lambda item: (item['subject_code'], item['subject_name'])),
    }


def parse_int_list_param(value):
    if value in (None, ''):
        return []
    if isinstance(value, (list, tuple, set)):
        items = value
    else:
        items = str(value).split(',')
    parsed_values = []
    for item in items:
        item_value = to_int(item, default=None)
        if item_value is not None:
            parsed_values.append(item_value)
    return parsed_values


def serialize_available_student_display_setting_payload(data):
    return {
        'enabled': parse_bool_param(data.get('enabled')),
        'require_resume': parse_bool_param(data.get('require_resume')),
        'allowed_admission_years': parse_int_list_param(data.get('allowed_admission_years')),
        'allowed_batch_ids': parse_int_list_param(data.get('allowed_batch_ids')),
        'allowed_postgraduate_types': parse_int_list_param(data.get('allowed_postgraduate_types')),
    }


def save_available_student_display_setting_from_payload(data):
    setting = get_available_student_display_setting()
    payload = serialize_available_student_display_setting_payload(data)

    if payload['enabled'] is not None:
        setting.enabled = payload['enabled']
    if payload['require_resume'] is not None:
        setting.require_resume = payload['require_resume']
    setting.allowed_admission_years = normalize_available_student_display_values(payload['allowed_admission_years'])
    setting.allowed_batch_ids = normalize_available_student_display_values(payload['allowed_batch_ids'])
    setting.allowed_postgraduate_types = normalize_available_student_display_values(payload['allowed_postgraduate_types'])
    setting.save()
    return setting


def update_professor_summary_fields(professor):
    refresh_professor_summary_quotas(professor)
    professor.save(update_fields=['name_fk_search'])


def serialize_professor_form_data(data):
    return {
        'name': str(data.get('name') or '').strip(),
        'teacher_identity_id': str(data.get('teacher_identity_id') or '').strip(),
        'professor_title': str(data.get('professor_title') or '').strip(),
        'department_id': data.get('department_id'),
        'email': str(data.get('email') or '').strip(),
        'phone_number': str(data.get('phone_number') or '').strip(),
        'avatar': str(data.get('avatar') or '').strip(),
        'contact_details': str(data.get('contact_details') or '').strip(),
        'research_areas': str(data.get('research_areas') or '').strip(),
        'personal_page': str(data.get('personal_page') or '').strip(),
        'have_qualification': parse_bool_param(data.get('have_qualification')),
        'proposed_quota_approved': parse_bool_param(data.get('proposed_quota_approved')),
        'department_position': data.get('department_position'),
        'website_order': data.get('website_order'),
        'password': str(data.get('password') or '').strip(),
    }


def save_professor_from_payload(payload, professor=None):
    name = payload['name']
    teacher_identity_id = payload['teacher_identity_id']
    if not name:
        raise ValidationError('导师姓名不能为空。')
    if not teacher_identity_id:
        raise ValidationError('导师工号不能为空。')

    existing_professor = Professor.objects.filter(teacher_identity_id=teacher_identity_id)
    if professor:
        existing_professor = existing_professor.exclude(pk=professor.pk)
    if existing_professor.exists():
        raise ValidationError('导师工号已存在。')

    department = Department.objects.filter(pk=payload['department_id']).first() if payload['department_id'] else None
    if payload['department_id'] and not department:
        raise ValidationError('招生方向不存在。')

    if professor is None:
        username = teacher_identity_id
        if User.objects.filter(username=username).exists():
            raise ValidationError('同名用户已存在，请先调整工号。')
        user = User.objects.create_user(username=username, password=payload['password'] or teacher_identity_id)
        professor = Professor(user_name=user)
    else:
        user = professor.user_name
        if user and teacher_identity_id != professor.teacher_identity_id:
            duplicate_user = User.objects.filter(username=teacher_identity_id).exclude(pk=user.pk)
            if duplicate_user.exists():
                raise ValidationError('新工号对应的登录账号已存在。')
            user.username = teacher_identity_id
            user.save(update_fields=['username'])
        if user and payload['password']:
            user.set_password(payload['password'])
            user.save(update_fields=['password'])

    professor.name = name
    professor.teacher_identity_id = teacher_identity_id
    professor.professor_title = payload['professor_title'] or professor.professor_title or '副教授'
    professor.department = department
    professor.email = payload['email'] or None
    professor.phone_number = payload['phone_number'] or None
    professor.avatar = payload['avatar'] or None
    professor.contact_details = payload['contact_details'] or None
    professor.research_areas = payload['research_areas'] or None
    professor.personal_page = payload['personal_page'] or ''
    if payload['have_qualification'] is not None:
        professor.have_qualification = payload['have_qualification']
    if payload['proposed_quota_approved'] is not None:
        professor.proposed_quota_approved = payload['proposed_quota_approved']
    if payload['department_position'] not in (None, ''):
        professor.department_position = to_int(payload['department_position'], default=0)
    if payload['website_order'] not in (None, ''):
        professor.website_order = to_int(payload['website_order'], default=0)
    professor.save()
    update_professor_summary_fields(professor)
    return professor


def serialize_student_form_data(data):
    return {
        'name': str(data.get('name') or '').strip(),
        'candidate_number': str(data.get('candidate_number') or '').strip(),
        'identify_number': str(data.get('identify_number') or '').strip(),
        'subject_id': data.get('subject_id'),
        'admission_year': data.get('admission_year'),
        'admission_batch_id': data.get('admission_batch_id'),
        'can_login': parse_bool_param(data.get('can_login')),
        'selection_display_enabled': parse_bool_param(data.get('selection_display_enabled')),
        'student_type': data.get('student_type'),
        'postgraduate_type': data.get('postgraduate_type'),
        'study_mode': parse_bool_param(data.get('study_mode')),
        'phone_number': str(data.get('phone_number') or '').strip(),
        'initial_exam_score': data.get('initial_exam_score'),
        'secondary_exam_score': data.get('secondary_exam_score'),
        'initial_rank': data.get('initial_rank'),
        'secondary_rank': data.get('secondary_rank'),
        'final_rank': data.get('final_rank'),
        'is_selected': parse_bool_param(data.get('is_selected')),
        'is_alternate': parse_bool_param(data.get('is_alternate')),
        'alternate_rank': data.get('alternate_rank'),
        'is_giveup': parse_bool_param(data.get('is_giveup')),
        'password': str(data.get('password') or '').strip(),
    }


def save_student_from_payload(payload, student=None):
    name = payload['name']
    candidate_number = payload['candidate_number']
    if not name:
        raise ValidationError('学生姓名不能为空。')
    if not candidate_number:
        raise ValidationError('考生编号不能为空。')

    existing_student = Student.objects.filter(candidate_number=candidate_number)
    if student:
        existing_student = existing_student.exclude(pk=student.pk)
    if existing_student.exists():
        raise ValidationError('考生编号已存在。')

    subject = Subject.objects.filter(pk=payload['subject_id']).first() if payload['subject_id'] else None
    if payload['subject_id'] and not subject:
        raise ValidationError('报考专业不存在。')
    admission_batch = AdmissionBatch.objects.filter(pk=payload['admission_batch_id']).first() if payload['admission_batch_id'] else None
    if payload['admission_batch_id'] and not admission_batch:
        raise ValidationError('招生批次不存在。')

    previous_subject = student.subject if student else None
    previous_admission_year = student.admission_year if student else None

    if student is None:
        if User.objects.filter(username=candidate_number).exists():
            raise ValidationError('同名用户已存在，请先调整考生编号。')
        user = User.objects.create_user(username=candidate_number, password=payload['password'] or candidate_number)
        student = Student(user_name=user)
    else:
        user = student.user_name
        if user and candidate_number != student.candidate_number:
            duplicate_user = User.objects.filter(username=candidate_number).exclude(pk=user.pk)
            if duplicate_user.exists():
                raise ValidationError('新考生编号对应的登录账号已存在。')
            user.username = candidate_number
            user.save(update_fields=['username'])
        if user and payload['password']:
            user.set_password(payload['password'])
            user.save(update_fields=['password'])

    student.name = name
    student.candidate_number = candidate_number
    student.identify_number = payload['identify_number'] or None
    student.subject = subject
    if payload['admission_year'] not in (None, ''):
        student.admission_year = to_int(payload['admission_year'], default=get_default_admission_year())
    elif student.pk is None or not student.admission_year:
        student.admission_year = get_default_admission_year()
    student.admission_batch = admission_batch
    if payload['student_type'] not in (None, ''):
        student.student_type = to_int(payload['student_type'], default=student.student_type or 1)
    if payload['postgraduate_type'] not in (None, ''):
        student.postgraduate_type = to_int(payload['postgraduate_type'], default=student.postgraduate_type or 1)
    if payload['study_mode'] is not None:
        student.study_mode = payload['study_mode']
    student.phone_number = payload['phone_number'] or None
    student.initial_exam_score = float(payload['initial_exam_score']) if payload['initial_exam_score'] not in (None, '') else None
    student.secondary_exam_score = float(payload['secondary_exam_score']) if payload['secondary_exam_score'] not in (None, '') else None
    student.initial_rank = to_int(payload['initial_rank'], default=None) if payload['initial_rank'] not in (None, '') else None
    student.secondary_rank = to_int(payload['secondary_rank'], default=None) if payload['secondary_rank'] not in (None, '') else None
    student.final_rank = to_int(payload['final_rank'], default=None) if payload['final_rank'] not in (None, '') else None
    if payload['is_selected'] is not None:
        student.is_selected = payload['is_selected']
    if payload['is_alternate'] is not None:
        student.is_alternate = payload['is_alternate']
    student.alternate_rank = to_int(payload['alternate_rank'], default=None) if payload['alternate_rank'] not in (None, '') else None
    if payload['is_giveup'] is not None:
        student.is_giveup = payload['is_giveup']
    if payload['can_login'] is not None:
        student.can_login = payload['can_login']
    if payload['selection_display_enabled'] is not None:
        student.selection_display_enabled = payload['selection_display_enabled']
    student.save()

    if previous_subject and previous_subject != student.subject:
        sync_student_alternate_status(previous_subject, previous_admission_year)
    if student.subject:
        sync_student_alternate_status(student.subject, student.admission_year)
    return student


def save_department_from_payload(data, department=None):
    department_name = str(data.get('department_name') or '').strip()
    if not department_name:
        raise ValidationError('方向名称不能为空。')
    existing = Department.objects.filter(department_name=department_name)
    if department:
        existing = existing.exclude(pk=department.pk)
    if existing.exists():
        raise ValidationError('方向名称已存在。')

    if department is None:
        department = Department()

    department.department_name = department_name
    for field in ['total_academic_quota', 'total_professional_quota', 'total_professional_yt_quota', 'total_doctor_quota']:
        if field in data and data.get(field) not in (None, ''):
            value = to_int(data.get(field), default=0)
            if value < 0:
                raise ValidationError(f'{field} 必须是非负整数。')
            setattr(department, field, value)
    department.save()
    return department


def save_subject_from_payload(data, subject=None):
    subject_name = str(data.get('subject_name') or '').strip()
    subject_code = str(data.get('subject_code') or '').strip()
    if not subject_name:
        raise ValidationError('专业名称不能为空。')
    if not subject_code:
        raise ValidationError('专业代码不能为空。')

    existing = Subject.objects.filter(subject_code=subject_code)
    if subject:
        existing = existing.exclude(pk=subject.pk)
    if existing.exists():
        raise ValidationError('专业代码已存在。')

    if subject is None:
        subject = Subject()

    subject.subject_name = subject_name
    subject.subject_code = subject_code
    if data.get('subject_type') in (None, ''):
        raise ValidationError('专业类型不能为空。')
    subject.subject_type = to_int(data.get('subject_type'), default=0)
    total_admission_quota = to_int(data.get('total_admission_quota'), default=0)
    if total_admission_quota < 0:
        raise ValidationError('total_admission_quota 必须是非负整数。')
    subject.total_admission_quota = total_admission_quota
    subject.save()

    department_ids = data.get('department_ids')
    if department_ids is not None:
      subject.subject_department.set(Department.objects.filter(id__in=department_ids))
    return subject


def validate_fixed_quota_conflict(professor, subject, campus_values, exclude_pool_id=None):
    conflicting_pools = ProfessorSharedQuotaPool.objects.filter(
        professor=professor,
        is_active=True,
        subjects=subject,
    )
    if subject.subject_type == 2:
        conflicting_pools = conflicting_pools.filter(quota_scope=ProfessorSharedQuotaPool.SCOPE_DOCTOR)
    else:
        conflicting_pools = conflicting_pools.filter(
            quota_scope=ProfessorSharedQuotaPool.SCOPE_MASTER,
            campus__in=campus_values,
        )
    if exclude_pool_id:
        conflicting_pools = conflicting_pools.exclude(pk=exclude_pool_id)
    if conflicting_pools.exists():
        raise ValidationError('该专业已被同导师的共享名额池覆盖，请先移除共享池配置。')


def validate_shared_pool_conflict(professor, quota_scope, campus, subject_ids, exclude_pool=None):
    subject_queryset = Subject.objects.filter(id__in=subject_ids)
    if quota_scope == ProfessorSharedQuotaPool.SCOPE_DOCTOR:
        invalid_subject = subject_queryset.exclude(subject_type=2).first()
        if invalid_subject:
            raise ValidationError('博士共享名额池只能关联博士专业。')
        fixed_conflicts = ProfessorDoctorQuota.objects.filter(
            professor=professor,
            subject_id__in=subject_ids,
        ).filter(Q(total_quota__gt=0) | Q(used_quota__gt=0) | Q(remaining_quota__gt=0))
        if fixed_conflicts.exists():
            raise ValidationError('所选专业中存在已配置的博士单专业名额，请先清理后再使用共享名额池。')
        overlap_queryset = ProfessorSharedQuotaPool.objects.filter(
            professor=professor,
            quota_scope=quota_scope,
            subjects__id__in=subject_ids,
            is_active=True,
        )
    else:
        invalid_subject = subject_queryset.exclude(subject_type__in=[0, 1]).first()
        if invalid_subject:
            raise ValidationError('硕士共享名额池只能关联硕士专业。')
        fixed_conflicts = ProfessorMasterQuota.objects.filter(
            professor=professor,
            subject_id__in=subject_ids,
        )
        if campus == ProfessorSharedQuotaPool.CAMPUS_BEIJING:
            fixed_conflicts = fixed_conflicts.filter(Q(beijing_quota__gt=0) | Q(beijing_remaining_quota__gt=0))
        else:
            fixed_conflicts = fixed_conflicts.filter(Q(yantai_quota__gt=0) | Q(yantai_remaining_quota__gt=0))
        if fixed_conflicts.exists():
            raise ValidationError('所选专业中存在已配置的硕士单专业名额，请先清理后再使用共享名额池。')
        overlap_queryset = ProfessorSharedQuotaPool.objects.filter(
            professor=professor,
            quota_scope=quota_scope,
            campus=campus,
            subjects__id__in=subject_ids,
            is_active=True,
        )
    if exclude_pool:
        overlap_queryset = overlap_queryset.exclude(pk=exclude_pool.pk)
    if overlap_queryset.exists():
        raise ValidationError('同一导师下，所选专业已经被另一个共享名额池覆盖。')


def save_shared_quota_pool_from_payload(data, pool=None):
    professor_id = data.get('professor_id')
    pool_name = str(data.get('pool_name') or '').strip()
    quota_scope = str(data.get('quota_scope') or '').strip()
    campus = str(data.get('campus') or '').strip() or ProfessorSharedQuotaPool.CAMPUS_GENERAL
    subject_ids = data.get('subject_ids') or []

    if not professor_id:
        raise ValidationError('导师不能为空。')
    if not pool_name:
        raise ValidationError('共享名额池名称不能为空。')
    if quota_scope not in {ProfessorSharedQuotaPool.SCOPE_MASTER, ProfessorSharedQuotaPool.SCOPE_DOCTOR}:
        raise ValidationError('名额类型不合法。')
    if quota_scope == ProfessorSharedQuotaPool.SCOPE_MASTER and campus not in {
        ProfessorSharedQuotaPool.CAMPUS_BEIJING,
        ProfessorSharedQuotaPool.CAMPUS_YANTAI,
    }:
        raise ValidationError('硕士共享名额池必须指定北京或烟台。')
    if quota_scope == ProfessorSharedQuotaPool.SCOPE_DOCTOR:
        campus = ProfessorSharedQuotaPool.CAMPUS_GENERAL
    if not subject_ids:
        raise ValidationError('至少选择一个专业。')

    professor = Professor.objects.filter(pk=professor_id).first()
    if not professor:
        raise ValidationError('导师不存在。')

    validate_shared_pool_conflict(professor, quota_scope, campus, subject_ids, exclude_pool=pool)

    if pool is None:
        pool = ProfessorSharedQuotaPool(professor=professor)
    else:
        pool.professor = professor

    total_quota = to_int(data.get('total_quota'), default=pool.total_quota or 0)
    used_quota = to_int(data.get('used_quota'), default=pool.used_quota or 0)
    remaining_value = data.get('remaining_quota')
    remaining_quota = (
        to_int(remaining_value, default=total_quota - used_quota)
        if remaining_value not in (None, '')
        else total_quota - used_quota
    )
    if total_quota < 0 or used_quota < 0 or remaining_quota < 0:
        raise ValidationError('名额字段必须是非负整数。')
    if used_quota > total_quota:
        raise ValidationError('已用名额不能大于总名额。')

    pool.pool_name = pool_name
    pool.quota_scope = quota_scope
    pool.campus = campus
    pool.total_quota = total_quota
    pool.used_quota = used_quota
    pool.remaining_quota = remaining_quota
    if data.get('is_active') is not None:
        is_active = parse_bool_param(data.get('is_active'))
        pool.is_active = True if is_active is None else is_active
    pool.notes = str(data.get('notes') or '').strip()
    pool.save()
    pool.subjects.set(Subject.objects.filter(id__in=subject_ids))
    recalculate_professor_quota_summary(professor)
    return pool


def save_doctor_quota_from_payload(data, quota=None):
    professor_id = data.get('professor_id')
    subject_id = data.get('subject_id')
    if not professor_id:
        raise ValidationError('导师不能为空。')
    if not subject_id:
        raise ValidationError('博士专业不能为空。')

    professor = Professor.objects.filter(pk=professor_id).first()
    if not professor:
        raise ValidationError('导师不存在。')
    subject = Subject.objects.filter(pk=subject_id, subject_type=2).first()
    if not subject:
        raise ValidationError('博士专业不存在。')

    existing = ProfessorDoctorQuota.objects.filter(professor=professor, subject=subject)
    if quota:
        existing = existing.exclude(pk=quota.pk)
    if existing.exists():
        raise ValidationError('该导师的博士专业名额已存在。')
    validate_fixed_quota_conflict(professor, subject, [ProfessorSharedQuotaPool.CAMPUS_GENERAL])

    if quota is None:
        quota = ProfessorDoctorQuota(professor=professor, subject=subject)
    else:
        quota.professor = professor
        quota.subject = subject

    total_quota = to_int(data.get('total_quota'), default=quota.total_quota or 0)
    used_quota = to_int(data.get('used_quota'), default=quota.used_quota or 0)
    remaining_quota_value = data.get('remaining_quota')
    remaining_quota = to_int(remaining_quota_value, default=total_quota - used_quota) if remaining_quota_value not in (None, '') else total_quota - used_quota

    if total_quota < 0 or used_quota < 0 or remaining_quota < 0:
        raise ValidationError('名额字段必须是非负整数。')
    if used_quota > total_quota:
        raise ValidationError('已用名额不能大于总名额。')

    quota.total_quota = total_quota
    quota.used_quota = used_quota
    quota.remaining_quota = remaining_quota
    quota.save()
    refresh_professor_summary_quotas(professor)
    return quota


def save_master_quota_from_payload(data, quota=None):
    professor_id = data.get('professor_id')
    subject_id = data.get('subject_id')
    if not professor_id:
        raise ValidationError('导师不能为空。')
    if not subject_id:
        raise ValidationError('硕士专业不能为空。')

    professor = Professor.objects.filter(pk=professor_id).first()
    if not professor:
        raise ValidationError('导师不存在。')
    subject = Subject.objects.filter(pk=subject_id, subject_type__in=[0, 1]).first()
    if not subject:
        raise ValidationError('硕士专业不存在。')

    existing = ProfessorMasterQuota.objects.filter(professor=professor, subject=subject)
    if quota:
        existing = existing.exclude(pk=quota.pk)
    if existing.exists():
        raise ValidationError('该导师的硕士专业名额已存在。')
    campus_values = []
    if to_int(data.get('beijing_quota'), default=quota.beijing_quota if quota else 0) > 0:
        campus_values.append(ProfessorSharedQuotaPool.CAMPUS_BEIJING)
    if to_int(data.get('yantai_quota'), default=quota.yantai_quota if quota else 0) > 0:
        campus_values.append(ProfessorSharedQuotaPool.CAMPUS_YANTAI)
    if campus_values:
        validate_fixed_quota_conflict(professor, subject, campus_values)

    if quota is None:
        quota = ProfessorMasterQuota(professor=professor, subject=subject)
    else:
        quota.professor = professor
        quota.subject = subject

    beijing_quota = to_int(data.get('beijing_quota'), default=quota.beijing_quota or 0)
    yantai_quota = to_int(data.get('yantai_quota'), default=quota.yantai_quota or 0)
    beijing_remaining = data.get('beijing_remaining_quota')
    yantai_remaining = data.get('yantai_remaining_quota')
    quota.beijing_quota = beijing_quota
    quota.yantai_quota = yantai_quota
    if beijing_remaining not in (None, ''):
        quota.beijing_remaining_quota = to_int(beijing_remaining, default=beijing_quota)
    if yantai_remaining not in (None, ''):
        quota.yantai_remaining_quota = to_int(yantai_remaining, default=yantai_quota)
    if quota.beijing_quota < 0 or quota.yantai_quota < 0 or quota.beijing_remaining_quota < 0 or quota.yantai_remaining_quota < 0:
        raise ValidationError('名额字段必须是非负整数。')
    quota.save()
    refresh_professor_summary_quotas(professor)
    return quota


def save_wechat_account_from_payload(data, account=None):
    user_id = data.get('user_id')
    username = str(data.get('username') or '').strip()
    openid = str(data.get('openid') or '').strip()
    session_key = str(data.get('session_key') or '').strip()

    if not user_id and not username:
        raise ValidationError('用户不能为空。')
    if not openid:
        raise ValidationError('openid 不能为空。')
    if not session_key:
        raise ValidationError('session_key 不能为空。')

    user = User.objects.filter(pk=user_id).first() if user_id else User.objects.filter(username=username).first()
    if not user:
        raise ValidationError('用户不存在。')

    duplicate_openid = WeChatAccount.objects.filter(openid=openid)
    duplicate_user = WeChatAccount.objects.filter(user=user)
    if account:
        duplicate_openid = duplicate_openid.exclude(pk=account.pk)
        duplicate_user = duplicate_user.exclude(pk=account.pk)
    if duplicate_openid.exists():
        raise ValidationError('该 openid 已绑定其他账号。')
    if duplicate_user.exists():
        raise ValidationError('该用户已绑定微信账号。')

    if account is None:
        account = WeChatAccount(user=user)
    else:
        account.user = user
    account.openid = openid
    account.session_key = session_key
    account.save()
    return account


def save_user_from_payload(data, user=None):
    username = str(data.get('username') or '').strip()
    email = str(data.get('email') or '').strip()
    first_name = str(data.get('first_name') or '').strip()
    last_name = str(data.get('last_name') or '').strip()
    password = str(data.get('password') or '').strip()

    if not username:
        raise ValidationError('用户名不能为空。')

    existing = User.objects.filter(username=username)
    if user:
        existing = existing.exclude(pk=user.pk)
    if existing.exists():
        raise ValidationError('用户名已存在。')

    if user is None:
        user = User(username=username)

    user.username = username
    user.email = email
    user.first_name = first_name
    user.last_name = last_name

    is_active = parse_bool_param(data.get('is_active'))
    is_staff = parse_bool_param(data.get('is_staff'))
    is_superuser = parse_bool_param(data.get('is_superuser'))
    if is_active is not None:
        user.is_active = is_active
    if is_staff is not None:
        user.is_staff = is_staff
    if is_superuser is not None:
        user.is_superuser = is_superuser
        if is_superuser:
            user.is_staff = True

    if password:
        user.set_password(password)
    elif user.pk is None:
        user.set_password(username)

    user.save()
    return user


def save_selection_time_from_payload(data, selection_time=None):
    open_time = data.get('open_time')
    close_time = data.get('close_time')
    if not open_time:
        raise ValidationError('开始时间不能为空。')
    if not close_time:
        raise ValidationError('结束时间不能为空。')

    parsed_open_time = datetime.fromisoformat(str(open_time).replace('Z', '+00:00'))
    parsed_close_time = datetime.fromisoformat(str(close_time).replace('Z', '+00:00'))
    if timezone.is_naive(parsed_open_time):
        parsed_open_time = timezone.make_aware(parsed_open_time, timezone.get_current_timezone())
    if timezone.is_naive(parsed_close_time):
        parsed_close_time = timezone.make_aware(parsed_close_time, timezone.get_current_timezone())

    if parsed_open_time >= parsed_close_time:
        raise ValidationError('结束时间必须晚于开始时间。')

    if selection_time is None:
        selection_time = SelectionTime()
    selection_time.open_time = parsed_open_time
    selection_time.close_time = parsed_close_time
    selection_time.save()
    return selection_time


def ensure_selection_time(target):
    defaults = {
        'open_time': timezone.now(),
        'close_time': timezone.now() + timedelta(days=1),
    }
    selection_time, _ = SelectionTime.objects.get_or_create(target=target, defaults=defaults)
    return selection_time


def get_selection_time_payload():
    student_config = ensure_selection_time(SelectionTime.TARGET_STUDENT)
    professor_config = ensure_selection_time(SelectionTime.TARGET_PROFESSOR)
    return {
        'student': SelectionTimeSerializer(student_config).data,
        'professor': SelectionTimeSerializer(professor_config).data,
    }


def parse_batch_delete_request(request):
    ids = request.data.get('ids') or []
    delete_all_filtered = str(request.data.get('delete_all_filtered') or '').lower() == 'true'
    return ids, delete_all_filtered


def build_professor_queryset(request):
    queryset = Professor.objects.select_related('department', 'user_name').annotate(
        pending_choice_count=Count('studentprofessorchoice', filter=Q(studentprofessorchoice__status=3), distinct=True),
        accepted_choice_count=Count('studentprofessorchoice', filter=Q(studentprofessorchoice__status=1), distinct=True),
        rejected_choice_count=Count('studentprofessorchoice', filter=Q(studentprofessorchoice__status=2), distinct=True),
    )
    search = request.query_params.get('search', '').strip()
    department_id = request.query_params.get('department_id')
    qualification = request.query_params.get('have_qualification')
    reviewer_only = request.query_params.get('reviewer_only')
    if search:
        queryset = queryset.filter(
            Q(name__icontains=search) |
            Q(teacher_identity_id__icontains=search) |
            Q(research_areas__icontains=search)
        )
    if department_id:
        queryset = queryset.filter(department_id=department_id)
    if qualification in {'true', 'false'}:
        queryset = queryset.filter(have_qualification=(qualification == 'true'))
    if reviewer_only == 'true':
        queryset = queryset.filter(department_position__in=[1, 2])
    return queryset


def build_student_queryset(request):
    queryset = Student.objects.select_related('user_name', 'subject', 'admission_batch')
    search = request.query_params.get('search', '').strip()
    subject_id = request.query_params.get('subject_id')
    subject_type = request.query_params.get('subject_type')
    admission_year = request.query_params.get('admission_year')
    admission_batch_id = request.query_params.get('admission_batch_id')
    can_login = request.query_params.get('can_login')
    selection_display_enabled = request.query_params.get('selection_display_enabled')
    student_type = request.query_params.get('student_type')
    postgraduate_type = request.query_params.get('postgraduate_type')
    is_selected = request.query_params.get('is_selected')
    is_alternate = request.query_params.get('is_alternate')
    is_giveup = request.query_params.get('is_giveup')
    review_status = request.query_params.get('review_status')
    if search:
        queryset = queryset.filter(Q(name__icontains=search) | Q(candidate_number__icontains=search))
    if subject_id:
        queryset = queryset.filter(subject_id=subject_id)
    if subject_type not in (None, ''):
        queryset = queryset.filter(subject__subject_type=subject_type)
    if admission_year not in (None, ''):
        queryset = queryset.filter(admission_year=admission_year)
    if admission_batch_id not in (None, ''):
        queryset = queryset.filter(admission_batch_id=admission_batch_id)
    if can_login in {'true', 'false'}:
        queryset = queryset.filter(can_login=(can_login == 'true'))
    if selection_display_enabled in {'true', 'false'}:
        queryset = queryset.filter(selection_display_enabled=(selection_display_enabled == 'true'))
    if student_type not in (None, ''):
        queryset = queryset.filter(student_type=student_type)
    if postgraduate_type not in (None, ''):
        queryset = queryset.filter(postgraduate_type=postgraduate_type)
    if is_selected in {'true', 'false'}:
        queryset = queryset.filter(is_selected=(is_selected == 'true'))
    if is_alternate in {'true', 'false'}:
        queryset = queryset.filter(is_alternate=(is_alternate == 'true'))
    if is_giveup in {'true', 'false'}:
        queryset = queryset.filter(is_giveup=(is_giveup == 'true'))
    if review_status not in (None, ''):
        queryset = queryset.filter(signature_table_review_status=review_status)
    return queryset


def build_admission_batch_queryset(request):
    queryset = AdmissionBatch.objects.all()
    search = request.query_params.get('search', '').strip()
    admission_year = request.query_params.get('admission_year')
    is_active = request.query_params.get('is_active')
    if search:
        queryset = queryset.filter(
            Q(name__icontains=search) |
            Q(batch_code__icontains=search) |
            Q(description__icontains=search)
        )
    if admission_year not in (None, ''):
        queryset = queryset.filter(admission_year=admission_year)
    if is_active in {'true', 'false'}:
        queryset = queryset.filter(is_active=(is_active == 'true'))
    return queryset


def save_admission_batch_from_payload(data, admission_batch=None):
    name = str(data.get('name') or '').strip()
    if not name:
        raise ValidationError('批次名称不能为空。')

    admission_year = data.get('admission_year')
    if admission_year in (None, ''):
        admission_year = get_default_admission_year()
    admission_year = to_int(admission_year, default=get_default_admission_year())

    existing = AdmissionBatch.objects.filter(admission_year=admission_year, name=name)
    if admission_batch:
        existing = existing.exclude(pk=admission_batch.pk)
    if existing.exists():
        raise ValidationError('同届别下已存在同名招生批次。')

    if admission_batch is None:
        admission_batch = AdmissionBatch()

    admission_batch.name = name
    admission_batch.admission_year = admission_year
    admission_batch.batch_code = str(data.get('batch_code') or '').strip()
    admission_batch.sort_order = to_int(data.get('sort_order'), default=0)
    is_active = parse_bool_param(data.get('is_active'))
    if is_active is not None:
        admission_batch.is_active = is_active
    admission_batch.description = str(data.get('description') or '').strip()
    admission_batch.save()
    return admission_batch


def build_choice_queryset(request):
    queryset = StudentProfessorChoice.objects.select_related('student__subject', 'professor__department')
    status_value = request.query_params.get('status')
    student_id = request.query_params.get('student_id')
    professor_id = request.query_params.get('professor_id')
    subject_id = request.query_params.get('subject_id')
    search = request.query_params.get('search', '').strip()
    department_id = request.query_params.get('department_id')
    if status_value:
        queryset = queryset.filter(status=status_value)
    if student_id:
        queryset = queryset.filter(student_id=student_id)
    if professor_id:
        queryset = queryset.filter(professor_id=professor_id)
    if subject_id:
        queryset = queryset.filter(student__subject_id=subject_id)
    if department_id:
        queryset = queryset.filter(professor__department_id=department_id)
    if search:
        queryset = queryset.filter(
            Q(student__name__icontains=search) |
            Q(student__candidate_number__icontains=search) |
            Q(professor__name__icontains=search) |
            Q(professor__teacher_identity_id__icontains=search)
        )
    return queryset


def build_review_record_queryset(request):
    queryset = ReviewRecord.objects.select_related('student__subject', 'professor__department', 'reviewer')
    status_value = request.query_params.get('status')
    reviewer_id = request.query_params.get('reviewer_id')
    professor_id = request.query_params.get('professor_id')
    search = request.query_params.get('search', '').strip()
    subject_id = request.query_params.get('subject_id')
    if status_value:
        queryset = queryset.filter(status=status_value)
    if reviewer_id:
        queryset = queryset.filter(reviewer_id=reviewer_id)
    if professor_id:
        queryset = queryset.filter(professor_id=professor_id)
    if subject_id:
        queryset = queryset.filter(student__subject_id=subject_id)
    if search:
        queryset = queryset.filter(
            Q(student__name__icontains=search) |
            Q(student__candidate_number__icontains=search) |
            Q(professor__name__icontains=search)
        )
    return queryset


def build_department_queryset(request):
    queryset = Department.objects.prefetch_related(
        Prefetch(
            'professor_set',
            queryset=Professor.objects.filter(department_position__in=[1, 2]).only('id', 'name', 'department_id', 'department_position'),
            to_attr='prefetched_reviewers',
        )
    ).annotate(subject_count=Count('subjects', distinct=True))
    search = request.query_params.get('search', '').strip()
    if search:
        queryset = queryset.filter(department_name__icontains=search)
    return queryset


def build_subject_queryset(request):
    queryset = Subject.objects.prefetch_related(Prefetch('subject_department', to_attr='prefetched_departments')).annotate(
        student_count=Count('student', distinct=True),
        selected_student_count=Count('student', filter=Q(student__is_selected=True), distinct=True),
        alternate_student_count=Count('student', filter=Q(student__is_alternate=True, student__is_giveup=False), distinct=True),
        giveup_student_count=Count('student', filter=Q(student__is_giveup=True), distinct=True),
    )
    search = request.query_params.get('search', '').strip()
    subject_type = request.query_params.get('subject_type')
    department_id = request.query_params.get('department_id')
    if search:
        queryset = queryset.filter(Q(subject_name__icontains=search) | Q(subject_code__icontains=search))
    if subject_type:
        queryset = queryset.filter(subject_type=subject_type)
    if department_id:
        queryset = queryset.filter(subject_department__id=department_id)
    return queryset.distinct()


def build_alternate_queryset(request):
    queryset = Student.objects.select_related('user_name', 'subject').filter(is_alternate=True)
    search = request.query_params.get('search', '').strip()
    subject_id = request.query_params.get('subject_id')
    is_giveup = request.query_params.get('is_giveup')
    admission_year = request.query_params.get('admission_year')
    if search:
        queryset = queryset.filter(Q(name__icontains=search) | Q(candidate_number__icontains=search))
    if subject_id:
        queryset = queryset.filter(subject_id=subject_id)
    if is_giveup in {'true', 'false'}:
        queryset = queryset.filter(is_giveup=(is_giveup == 'true'))
    if admission_year not in (None, ''):
        queryset = queryset.filter(admission_year=admission_year)
    return queryset


def build_giveup_queryset(request):
    queryset = Student.objects.select_related('user_name', 'subject').filter(is_giveup=True)
    search = request.query_params.get('search', '').strip()
    subject_id = request.query_params.get('subject_id')
    is_selected = request.query_params.get('is_selected')
    if search:
        queryset = queryset.filter(Q(name__icontains=search) | Q(candidate_number__icontains=search))
    if subject_id:
        queryset = queryset.filter(subject_id=subject_id)
    if is_selected in {'true', 'false'}:
        queryset = queryset.filter(is_selected=(is_selected == 'true'))
    return queryset


class DashboardLoginView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        username = request.data.get('username', '').strip()
        password = request.data.get('password', '')
        if not username or not password:
            return Response({'detail': '请输入用户名和密码。'}, status=status.HTTP_400_BAD_REQUEST)

        user = authenticate(username=username, password=password)
        if not user or not user.is_staff:
            create_audit_log(
                request,
                action='auth.login',
                module='认证登录',
                level=DashboardAuditLog.LEVEL_WARNING,
                status_value=DashboardAuditLog.STATUS_FAILED,
                operator_username=username,
                target_type='user',
                target_display=username,
                detail='管理员登录失败。',
            )
            return Response({'detail': '用户名或密码错误，或该账号没有后台登录权限。'}, status=status.HTTP_401_UNAUTHORIZED)

        Token.objects.filter(user=user).delete()
        token = Token.objects.create(user=user)
        create_audit_log(
            request,
            action='auth.login',
            module='认证登录',
            operator=user,
            target_type='user',
            target_id=user.pk,
            target_display=get_display_name_for_user(user),
            detail='管理员登录成功。',
            after_data={'token_created_at': token.created},
        )
        return Response({'token': token.key, 'user': DashboardAdminSerializer(user).data}, status=status.HTTP_200_OK)


class DashboardLogoutView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def post(self, request):
        create_audit_log(
            request,
            action='auth.logout',
            module='认证登录',
            target_type='user',
            target_id=request.user.pk,
            target_display=get_display_name_for_user(request.user),
            detail='管理员退出登录。',
        )
        if request.auth:
            request.auth.delete()
        return Response({'detail': '已退出登录。'}, status=status.HTTP_200_OK)


class DashboardMeView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def get(self, request):
        return Response(DashboardAdminSerializer(request.user).data, status=status.HTTP_200_OK)


class DashboardAuditLogListView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def get(self, request):
        page = max(int(request.query_params.get('page', 1)), 1)
        page_size = min(max(int(request.query_params.get('page_size', 20)), 1), 200)
        queryset = build_audit_log_queryset(request)

        order_by = request.query_params.get('order_by', 'created_at')
        order_direction = request.query_params.get('order_direction', 'desc')
        order_map = {
            'created_at': 'created_at',
            'module': 'module',
            'action': 'action',
            'status': 'status',
            'level': 'level',
            'operator_username': 'operator_username',
            'target_display': 'target_display',
        }
        order_field = order_map.get(order_by, 'created_at')
        if order_direction == 'desc':
            order_field = f'-{order_field}'
        queryset = queryset.order_by(order_field, '-id')

        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        serializer = DashboardAuditLogListSerializer(queryset[start:end], many=True)
        return Response({'count': total, 'page': page, 'page_size': page_size, 'results': serializer.data}, status=status.HTTP_200_OK)


class DashboardAuditLogDetailView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def get(self, request, pk):
        audit_log = DashboardAuditLog.objects.select_related('operator').filter(pk=pk).first()
        if not audit_log:
            return Response({'detail': '操作日志不存在。'}, status=status.HTTP_404_NOT_FOUND)
        return Response(DashboardAuditLogDetailSerializer(audit_log).data, status=status.HTTP_200_OK)


class DashboardUserListView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def get(self, request):
        page = max(int(request.query_params.get('page', 1)), 1)
        page_size = min(max(int(request.query_params.get('page_size', 20)), 1), 200)
        queryset = User.objects.all()

        search = request.query_params.get('search', '').strip()
        user_type = request.query_params.get('user_type')
        is_active = request.query_params.get('is_active')
        order_by = request.query_params.get('order_by', 'date_joined')
        order_direction = request.query_params.get('order_direction', 'desc')

        if search:
            queryset = queryset.filter(
                Q(username__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search) |
                Q(professor__name__icontains=search) |
                Q(student__name__icontains=search)
            ).distinct()

        if user_type == 'professor':
            queryset = queryset.filter(professor__isnull=False)
        elif user_type == 'student':
            queryset = queryset.filter(student__isnull=False)
        elif user_type == 'staff':
            queryset = queryset.filter(is_staff=True, professor__isnull=True, student__isnull=True)
        elif user_type == 'normal':
            queryset = queryset.filter(is_staff=False, professor__isnull=True, student__isnull=True)

        if is_active in {'true', 'false'}:
            queryset = queryset.filter(is_active=(is_active == 'true'))

        order_map = {
            'username': 'username',
            'email': 'email',
            'date_joined': 'date_joined',
            'last_login': 'last_login',
            'first_name': 'first_name',
            'is_active': 'is_active',
            'is_staff': 'is_staff',
            'is_superuser': 'is_superuser',
        }
        order_field = order_map.get(order_by, 'date_joined')
        if order_direction == 'desc':
            order_field = f'-{order_field}'
        queryset = queryset.order_by(order_field, '-id')

        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        serializer = DashboardUserSerializer(queryset[start:end], many=True)
        return Response({'count': total, 'page': page, 'page_size': page_size, 'results': serializer.data}, status=status.HTTP_200_OK)

    def post(self, request):
        try:
            user = save_user_from_payload(request.data)
        except ValidationError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        create_audit_log(
            request,
            action='user.create',
            module='用户管理',
            target_type='user',
            target_id=user.pk,
            target_display=get_display_name_for_user(user),
            detail='创建后台用户。',
            after_data=snapshot_instance(user),
        )
        return Response(DashboardUserSerializer(user).data, status=status.HTTP_201_CREATED)


class DashboardUserDetailView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def get(self, request, pk):
        user = User.objects.filter(pk=pk).first()
        if not user:
            return Response({'detail': '用户不存在。'}, status=status.HTTP_404_NOT_FOUND)
        return Response(DashboardUserSerializer(user).data, status=status.HTTP_200_OK)

    def patch(self, request, pk):
        user = User.objects.filter(pk=pk).first()
        if not user:
            return Response({'detail': '用户不存在。'}, status=status.HTTP_404_NOT_FOUND)
        before_data = snapshot_instance(user)
        try:
            user = save_user_from_payload(request.data, user=user)
        except ValidationError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        create_audit_log(
            request,
            action='user.update',
            module='用户管理',
            target_type='user',
            target_id=user.pk,
            target_display=get_display_name_for_user(user),
            detail='更新后台用户。',
            before_data=before_data,
            after_data=snapshot_instance(user),
        )
        return Response(DashboardUserSerializer(user).data, status=status.HTTP_200_OK)

    def delete(self, request, pk):
        user = User.objects.filter(pk=pk).first()
        if not user:
            return Response({'detail': '用户不存在。'}, status=status.HTTP_404_NOT_FOUND)
        if user.pk == request.user.pk:
            return Response({'detail': '不能删除当前登录用户。'}, status=status.HTTP_400_BAD_REQUEST)
        before_data = snapshot_instance(user)
        target_display = get_display_name_for_user(user)
        user.delete()
        create_audit_log(
            request,
            action='user.delete',
            module='用户管理',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='user',
            target_id=pk,
            target_display=target_display,
            detail='删除后台用户。',
            before_data=before_data,
        )
        return Response({'detail': '用户已删除。'}, status=status.HTTP_200_OK)


class DashboardSelectionTimeListView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def get(self, request):
        return Response(get_selection_time_payload(), status=status.HTTP_200_OK)

    def patch(self, request):
        student_payload = request.data.get('student')
        professor_payload = request.data.get('professor')
        if not isinstance(student_payload, dict) or not isinstance(professor_payload, dict):
            return Response({'detail': '学生端和导师端的互选时间都需要填写。'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            student_config = ensure_selection_time(SelectionTime.TARGET_STUDENT)
            professor_config = ensure_selection_time(SelectionTime.TARGET_PROFESSOR)
            before_data = {
                'student': snapshot_instance(student_config, extra={'target_display': student_config.get_target_display()}),
                'professor': snapshot_instance(professor_config, extra={'target_display': professor_config.get_target_display()}),
            }
            save_selection_time_from_payload(student_payload, selection_time=student_config)
            save_selection_time_from_payload(professor_payload, selection_time=professor_config)
        except ValidationError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        after_payload = get_selection_time_payload()
        create_audit_log(
            request,
            action='selection_time.update',
            module='互选时间设置',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='selection_time',
            target_display='学生端 / 导师端互选时间',
            detail='更新互选时间配置。',
            before_data=before_data,
            after_data=after_payload,
        )
        return Response(after_payload, status=status.HTTP_200_OK)


class DashboardFileDownloadUrlView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def post(self, request):
        file_id = request.data.get('file_id', '').strip()
        if not file_id:
            return Response({'detail': 'file_id is required.'}, status=status.HTTP_400_BAD_REQUEST)
        if file_id.startswith('http://') or file_id.startswith('https://'):
            create_audit_log(
                request,
                action='file.preview',
                module='材料访问',
                target_type='file',
                target_display=file_id,
                detail='访问直链文件。',
            )
            return Response({'download_url': file_id}, status=status.HTTP_200_OK)

        try:
            download_url = get_file_download_url(file_id)
        except requests.RequestException as exc:
            return Response({'detail': f'Failed to request download url: {exc}'}, status=status.HTTP_502_BAD_GATEWAY)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        create_audit_log(
            request,
            action='file.preview',
            module='材料访问',
            target_type='file',
            target_display=file_id,
            detail='获取云存储文件访问地址。',
        )
        return Response({'download_url': download_url}, status=status.HTTP_200_OK)


class DashboardStatsView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def get(self, request):
        admission_year = to_int(request.query_params.get('admission_year'), default=2026)
        student_type = to_int(request.query_params.get('student_type'), default=2)
        if student_type not in {1, 2, 3}:
            student_type = 2

        cache_key = f'dashboard_stats_v2:{admission_year}:{student_type}'
        cached_payload = cache.get(cache_key)
        if cached_payload:
            return Response(cached_payload, status=status.HTTP_200_OK)

        today = timezone.localdate()
        trend_days = [today - timedelta(days=offset) for offset in range(6, -1, -1)]
        student_queryset = Student.objects.filter(admission_year=admission_year, student_type=student_type)
        student_ids = student_queryset.values('id')
        choice_queryset = StudentProfessorChoice.objects.filter(student_id__in=student_ids)
        review_queryset = ReviewRecord.objects.filter(student_id__in=student_ids)

        subject_type_filter = 2 if student_type == 3 else [0, 1]
        subject_quota = []
        if isinstance(subject_type_filter, list):
            subject_queryset = Subject.objects.filter(subject_type__in=subject_type_filter).order_by('subject_code', 'id')
        else:
            subject_queryset = Subject.objects.filter(subject_type=subject_type_filter).order_by('subject_code', 'id')

        for subject in subject_queryset:
            total = subject.total_admission_quota or 0
            selected = student_queryset.filter(subject=subject, is_selected=True, is_giveup=False).count()
            alternate = student_queryset.filter(subject=subject, is_alternate=True, is_giveup=False).count()
            giveup = student_queryset.filter(subject=subject, is_giveup=True).count()
            subject_quota.append(
                {
                    'subject_id': subject.id,
                    'subject_name': subject.subject_name,
                    'subject_code': subject.subject_code,
                    'total_quota': total,
                    'selected_count': selected,
                    'remaining_quota': max(total - selected, 0),
                    'alternate_count': alternate,
                    'giveup_count': giveup,
                }
            )

        department_usage = []
        for department in Department.objects.all().order_by('department_name'):
            total_master = (
                (department.total_academic_quota or 0)
                + (department.total_professional_quota or 0)
                + (department.total_professional_yt_quota or 0)
            )
            used_master = (
                (department.used_academic_quota or 0)
                + (department.used_professional_quota or 0)
                + (department.used_professional_yt_quota or 0)
            )
            total_doctor = department.total_doctor_quota or 0
            used_doctor = department.used_doctor_quota or 0
            department_usage.append(
                {
                    'department_id': department.id,
                    'department_name': department.department_name,
                    'master_total': total_master,
                    'master_used': used_master,
                    'master_remaining': max(total_master - used_master, 0),
                    'doctor_total': total_doctor,
                    'doctor_used': used_doctor,
                    'doctor_remaining': max(total_doctor - used_doctor, 0),
                }
            )

        review_distribution = [
            {'label': '待审核', 'value': review_queryset.filter(status=3).count()},
            {'label': '已通过', 'value': review_queryset.filter(status=1).count()},
            {'label': '已驳回', 'value': review_queryset.filter(status=2).count()},
            {'label': '已撤销', 'value': review_queryset.filter(status=4).count()},
        ]

        student_status_distribution = [
            {'label': '已录取', 'value': student_queryset.filter(is_selected=True, is_giveup=False).count()},
            {'label': '候补中', 'value': student_queryset.filter(is_alternate=True, is_giveup=False).count()},
            {'label': '已放弃', 'value': student_queryset.filter(is_giveup=True).count()},
            {
                'label': '未完成',
                'value': student_queryset.filter(is_selected=False, is_alternate=False, is_giveup=False).count(),
            },
        ]

        top_pending_professors = []
        professor_rows = (
            Professor.objects.select_related('department')
            .annotate(
                pending_count=Count(
                    'studentprofessorchoice',
                    filter=Q(
                        studentprofessorchoice__status=3,
                        studentprofessorchoice__student__admission_year=admission_year,
                        studentprofessorchoice__student__student_type=student_type,
                    ),
                    distinct=True,
                ),
                accepted_count=Count(
                    'studentprofessorchoice',
                    filter=Q(
                        studentprofessorchoice__status=1,
                        studentprofessorchoice__student__admission_year=admission_year,
                        studentprofessorchoice__student__student_type=student_type,
                    ),
                    distinct=True,
                ),
            )
            .order_by('-pending_count', 'remaining_quota', 'id')[:8]
        )
        for professor in professor_rows:
            top_pending_professors.append(
                {
                    'professor_id': professor.id,
                    'professor_name': professor.name,
                    'teacher_identity_id': professor.teacher_identity_id,
                    'department_name': professor.department.department_name if professor.department_id else '',
                    'pending_count': professor.pending_count or 0,
                    'accepted_count': professor.accepted_count or 0,
                    'remaining_quota': professor.remaining_quota or 0,
                }
            )

        low_quota_professors = []
        low_quota_rows = (
            Professor.objects.select_related('department')
            .filter(have_qualification=True)
            .order_by('remaining_quota', '-proposed_quota_approved', 'id')[:8]
        )
        for professor in low_quota_rows:
            low_quota_professors.append(
                {
                    'professor_id': professor.id,
                    'professor_name': professor.name,
                    'teacher_identity_id': professor.teacher_identity_id,
                    'department_name': professor.department.department_name if professor.department_id else '',
                    'remaining_quota': professor.remaining_quota or 0,
                    'have_qualification': professor.have_qualification,
                    'proposed_quota_approved': professor.proposed_quota_approved,
                }
            )

        choice_trend = []
        review_trend = []
        accepted_choice_trend = []
        for day in trend_days:
            next_day = day + timedelta(days=1)
            choice_trend.append(
                {
                    'date': day.isoformat(),
                    'label': day.strftime('%m-%d'),
                    'value': choice_queryset.filter(submit_date__gte=day, submit_date__lt=next_day).count(),
                }
            )
            review_trend.append(
                {
                    'date': day.isoformat(),
                    'label': day.strftime('%m-%d'),
                    'value': review_queryset.filter(review_time__gte=day, review_time__lt=next_day, status=1).count(),
                }
            )
            accepted_choice_trend.append(
                {
                    'date': day.isoformat(),
                    'label': day.strftime('%m-%d'),
                    'value': choice_queryset.filter(finish_time__gte=day, finish_time__lt=next_day, status=1).count(),
                }
            )

        payload = {
                'admission_year': admission_year,
                'student_type': student_type,
                'available_admission_years': sorted(Student.objects.values_list('admission_year', flat=True).distinct()),
                'student_type_options': [
                    {'value': 2, 'label': '硕士统考生'},
                    {'value': 1, 'label': '硕士推荐生'},
                    {'value': 3, 'label': '博士统考生'},
                ],
                'professor_count': Professor.objects.count(),
                'student_count': student_queryset.count(),
                'pending_choice_count': choice_queryset.filter(status=3).count(),
                'accepted_choice_count': choice_queryset.filter(status=1).count(),
                'pending_review_count': review_queryset.filter(status=3).count(),
                'approved_review_count': review_queryset.filter(status=1).count(),
                'rejected_review_count': review_queryset.filter(status=2).count(),
                'revoked_review_count': review_queryset.filter(status=4).count(),
                'alternate_student_count': student_queryset.filter(is_alternate=True, is_giveup=False).count(),
                'giveup_student_count': student_queryset.filter(is_giveup=True).count(),
                'department_usage': department_usage,
                'subject_quota': subject_quota,
                'subject_quota_title': '博士专业剩余名额' if student_type == 3 else '硕士专业剩余名额',
                'review_distribution': review_distribution,
                'student_status_distribution': student_status_distribution,
                'top_pending_professors': top_pending_professors,
                'low_quota_professors': low_quota_professors,
                'choice_trend': choice_trend,
                'review_trend': review_trend,
                'accepted_choice_trend': accepted_choice_trend,
            }
        cache.set('dashboard_stats_v1', payload, timeout=60)
        return Response(payload, status=status.HTTP_200_OK)


class DashboardProfessorListView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def get(self, request):
        page = max(int(request.query_params.get('page', 1)), 1)
        page_size = min(max(int(request.query_params.get('page_size', 10)), 1), 100)
        queryset = build_professor_queryset(request).select_related('department').prefetch_related(
            'master_quotas__subject',
            'doctor_quotas__subject',
            'shared_quota_pools__subjects',
        )

        order_by = request.query_params.get('order_by', 'website_order')
        order_direction = request.query_params.get('order_direction', 'asc')

        order_map = {
            'name': 'name',
            'teacher_identity_id': 'teacher_identity_id',
            'professor_title': 'professor_title',
            'department_name': 'department__department_name',
            'remaining_quota': 'remaining_quota',
            'pending_choice_count': 'pending_choice_count',
            'accepted_choice_count': 'accepted_choice_count',
            'website_order': 'website_order',
        }
        order_field = order_map.get(order_by, 'website_order')
        if order_direction == 'desc':
            order_field = f'-{order_field}'
        queryset = queryset.order_by(order_field, 'id')

        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        serializer = ProfessorListSerializer(queryset[start:end], many=True)
        return Response(
            {'count': total, 'page': page, 'page_size': page_size, 'results': serializer.data},
            status=status.HTTP_200_OK,
        )

    def post(self, request):
        try:
            professor = save_professor_from_payload(serialize_professor_form_data(request.data))
        except ValidationError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        create_audit_log(
            request,
            action='professor.create',
            module='导师管理',
            target_type='professor',
            target_id=professor.pk,
            target_display=professor.name,
            detail='创建导师。',
            after_data=snapshot_instance(professor),
        )
        return Response(ProfessorDetailSerializer(professor).data, status=status.HTTP_201_CREATED)


class DashboardProfessorHeatListView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def get(self, request):
        page = max(int(request.query_params.get('page', 1)), 1)
        page_size = min(max(int(request.query_params.get('page_size', 10)), 1), 100)
        setting = get_professor_heat_display_setting()
        subject_id = request.query_params.get('subject_id')
        postgraduate_type = request.query_params.get('postgraduate_type')
        student_type = request.query_params.get('student_type')
        heat_subject = Subject.objects.filter(pk=subject_id).first() if subject_id else None
        heat_postgraduate_type = None
        heat_student_type = None
        if postgraduate_type not in (None, ''):
            try:
                heat_postgraduate_type = int(postgraduate_type)
            except (TypeError, ValueError):
                return Response({'detail': '培养类型参数不正确。'}, status=status.HTTP_400_BAD_REQUEST)
        if student_type not in (None, ''):
            try:
                heat_student_type = int(student_type)
            except (TypeError, ValueError):
                return Response({'detail': '学生类型参数不正确。'}, status=status.HTTP_400_BAD_REQUEST)

        queryset = build_professor_queryset(request).select_related('department')
        professors = list(queryset)
        heat_level = request.query_params.get('heat_level')

        filtered_professors = []
        for professor in professors:
            metrics = get_professor_heat_display_metrics(
                professor,
                global_setting=setting,
                subject=heat_subject,
                postgraduate_type=heat_postgraduate_type,
                student_type=heat_student_type,
            )
            if heat_subject and (metrics.get('available_quota_total') or 0) <= 0:
                continue
            if heat_level and metrics['heat_level'] != heat_level:
                continue
            filtered_professors.append(professor)

        filtered_professors.sort(
            key=lambda professor: (
                -get_professor_heat_display_metrics(
                    professor,
                    global_setting=setting,
                    subject=heat_subject,
                    postgraduate_type=heat_postgraduate_type,
                    student_type=heat_student_type,
                )['pending_count'],
                professor.id,
            )
        )

        total = len(filtered_professors)
        start = (page - 1) * page_size
        end = start + page_size
        serializer = ProfessorHeatListSerializer(
            filtered_professors[start:end],
            many=True,
            context={
                'heat_setting': setting,
                'heat_subject': heat_subject,
                'heat_postgraduate_type': heat_postgraduate_type,
                'heat_student_type': heat_student_type,
            },
        )
        return Response(
            {
                'count': total,
                'page': page,
                'page_size': page_size,
                'results': serializer.data,
                'subject_id': heat_subject.id if heat_subject else None,
                'subject_name': heat_subject.subject_name if heat_subject else '',
                'postgraduate_type': heat_postgraduate_type,
                'student_type': heat_student_type,
                'calculation_scope': ProfessorHeatDisplaySetting.CALCULATION_SCOPE_SUBJECT,
                'target_admission_year': setting.target_admission_year,
            },
            status=status.HTTP_200_OK,
        )


class DashboardProfessorHeatSettingView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def get(self, request):
        setting = get_professor_heat_display_setting()
        return Response(ProfessorHeatDisplaySettingSerializer(setting).data, status=status.HTTP_200_OK)

    def patch(self, request):
        setting = get_professor_heat_display_setting()
        before_data = {
            'show_professor_heat': setting.show_professor_heat,
            'calculation_scope': ProfessorHeatDisplaySetting.CALCULATION_SCOPE_SUBJECT,
            'target_admission_year': setting.target_admission_year,
            'medium_threshold': str(setting.medium_threshold),
            'high_threshold': str(setting.high_threshold),
            'very_high_threshold': str(setting.very_high_threshold),
        }

        if 'show_professor_heat' in request.data:
            show_professor_heat = request.data.get('show_professor_heat')
            setting.show_professor_heat = parse_request_bool(show_professor_heat, default=setting.show_professor_heat)

        setting.calculation_scope = ProfessorHeatDisplaySetting.CALCULATION_SCOPE_SUBJECT

        if 'target_admission_year' in request.data:
            try:
                target_admission_year = int(request.data.get('target_admission_year'))
            except (TypeError, ValueError):
                return Response({'detail': '统计届别不是有效整数。'}, status=status.HTTP_400_BAD_REQUEST)
            if target_admission_year <= 0:
                return Response({'detail': '统计届别必须大于 0。'}, status=status.HTTP_400_BAD_REQUEST)
            setting.target_admission_year = target_admission_year

        decimal_fields = ['medium_threshold', 'high_threshold', 'very_high_threshold']
        for field_name in decimal_fields:
            if field_name not in request.data:
                continue
            raw_value = request.data.get(field_name)
            try:
                decimal_value = Decimal(str(raw_value))
            except Exception:
                return Response({'detail': f'{field_name} 不是有效数字。'}, status=status.HTTP_400_BAD_REQUEST)
            if decimal_value < 0:
                return Response({'detail': f'{field_name} 不能小于 0。'}, status=status.HTTP_400_BAD_REQUEST)
            setattr(setting, field_name, decimal_value)

        if not (setting.medium_threshold < setting.high_threshold < setting.very_high_threshold):
            return Response(
                {'detail': '热度阈值必须满足：中热度 < 高热度 < 很高热度。'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        setting.save()
        create_audit_log(
            request,
            action='professor_heat.setting_update',
            module='导师热度管理',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='professor_heat_setting',
            target_id=setting.pk,
            target_display='导师热度配置',
            detail='更新导师热度显示与计算规则。',
            before_data=before_data,
            after_data={
                'show_professor_heat': setting.show_professor_heat,
                'calculation_scope': ProfessorHeatDisplaySetting.CALCULATION_SCOPE_SUBJECT,
                'target_admission_year': setting.target_admission_year,
                'medium_threshold': str(setting.medium_threshold),
                'high_threshold': str(setting.high_threshold),
                'very_high_threshold': str(setting.very_high_threshold),
            },
        )
        return Response(ProfessorHeatDisplaySettingSerializer(setting).data, status=status.HTTP_200_OK)


class DashboardProfessorHeatDetailView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def patch(self, request, pk):
        professor = Professor.objects.filter(pk=pk).first()
        if not professor:
            return Response({'detail': '导师不存在。'}, status=status.HTTP_404_NOT_FOUND)

        before_data = {
            'heat_display_enabled': professor.heat_display_enabled,
            'manual_heat_score': str(professor.manual_heat_score) if professor.manual_heat_score is not None else None,
            'manual_heat_level': professor.manual_heat_level or '',
        }

        if 'heat_display_enabled' in request.data:
            professor.heat_display_enabled = parse_request_bool(
                request.data.get('heat_display_enabled'),
                default=professor.heat_display_enabled,
            )

        if 'manual_heat_score' in request.data:
            raw_score = request.data.get('manual_heat_score')
            professor.manual_heat_score = None if raw_score in (None, '', 'null') else raw_score

        if 'manual_heat_level' in request.data:
            raw_level = str(request.data.get('manual_heat_level') or '').strip()
            professor.manual_heat_level = raw_level or None

        professor.save(update_fields=['heat_display_enabled', 'manual_heat_score', 'manual_heat_level'])

        create_audit_log(
            request,
            action='professor_heat.update',
            module='导师热度管理',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='professor',
            target_id=professor.pk,
            target_display=professor.name,
            detail='更新导师热度显示开关或手动热度。',
            before_data=before_data,
            after_data={
                'heat_display_enabled': professor.heat_display_enabled,
                'manual_heat_score': str(professor.manual_heat_score) if professor.manual_heat_score is not None else None,
                'manual_heat_level': professor.manual_heat_level or '',
            },
        )
        return Response(ProfessorHeatListSerializer(professor).data, status=status.HTTP_200_OK)


class DashboardProfessorDetailView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def get(self, request, pk):
        professor = Professor.objects.select_related('department').prefetch_related(
            'master_quotas__subject',
            'doctor_quotas__subject',
            'shared_quota_pools__subjects',
        ).annotate(
            pending_choice_count=Count('studentprofessorchoice', filter=Q(studentprofessorchoice__status=3), distinct=True),
            accepted_choice_count=Count('studentprofessorchoice', filter=Q(studentprofessorchoice__status=1), distinct=True),
        ).filter(pk=pk).first()
        if not professor:
            return Response({'detail': 'Professor not found.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(ProfessorDetailSerializer(professor).data, status=status.HTTP_200_OK)

    def patch(self, request, pk):
        professor = Professor.objects.select_related('user_name').filter(pk=pk).first()
        if not professor:
            return Response({'detail': 'Professor not found.'}, status=status.HTTP_404_NOT_FOUND)
        before_data = snapshot_instance(professor)
        try:
            professor = save_professor_from_payload(serialize_professor_form_data(request.data), professor=professor)
        except ValidationError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        create_audit_log(
            request,
            action='professor.update',
            module='导师管理',
            target_type='professor',
            target_id=professor.pk,
            target_display=professor.name,
            detail='更新导师信息。',
            before_data=before_data,
            after_data=snapshot_instance(professor),
        )
        return Response(ProfessorDetailSerializer(professor).data, status=status.HTTP_200_OK)

    def delete(self, request, pk):
        professor = Professor.objects.select_related('user_name').filter(pk=pk).first()
        if not professor:
            return Response({'detail': '导师不存在。'}, status=status.HTTP_404_NOT_FOUND)
        user = professor.user_name
        before_data = snapshot_instance(professor)
        target_display = professor.name
        professor.delete()
        if user:
            user.delete()
        create_audit_log(
            request,
            action='professor.delete',
            module='导师管理',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='professor',
            target_id=pk,
            target_display=target_display,
            detail='删除导师。',
            before_data=before_data,
        )
        return Response({'detail': '导师已删除。'}, status=status.HTTP_200_OK)


class DashboardProfessorBatchDeleteView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def post(self, request):
        ids, delete_all_filtered = parse_batch_delete_request(request)
        queryset = build_professor_queryset(request).select_related('user_name')
        if not delete_all_filtered:
            queryset = queryset.filter(id__in=ids)
        if not queryset.exists():
            return Response({'detail': '没有可删除的导师记录。'}, status=status.HTTP_400_BAD_REQUEST)
        users = [professor.user_name for professor in queryset if professor.user_name_id]
        target_names = list(queryset.values_list('name', flat=True))
        deleted_count = queryset.count()
        queryset.delete()
        for user in users:
            User.objects.filter(pk=user.pk).delete()
        create_audit_log(
            request,
            action='professor.batch_delete',
            module='导师管理',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='professor',
            target_display='批量删除导师',
            detail=f'批量删除 {deleted_count} 位导师。',
            before_data={'count': deleted_count, 'names': target_names, 'delete_all_filtered': delete_all_filtered},
        )
        return Response({'detail': f'已删除 {deleted_count} 位导师。'}, status=status.HTTP_200_OK)


class DashboardProfessorResetPasswordView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def post(self, request):
        ids = request.data.get('ids') or []
        queryset = Professor.objects.select_related('user_name').filter(id__in=ids)
        updated_count = 0

        for professor in queryset:
            if professor.user_name:
                professor.user_name.set_password(professor.teacher_identity_id)
                professor.user_name.save(update_fields=['password'])
                updated_count += 1
        create_audit_log(
            request,
            action='professor.reset_password',
            module='导师管理',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='professor',
            target_display='批量重置导师密码',
            detail=f'重置 {updated_count} 位导师密码。',
            after_data={'ids': ids, 'updated_count': updated_count},
        )
        return Response({'detail': f'已重置 {updated_count} 位导师的密码。'}, status=status.HTTP_200_OK)


class DashboardProfessorResetSelectionStatusView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def post(self, request):
        ids = request.data.get('ids') or []
        updated_count = Professor.objects.filter(id__in=ids).update(proposed_quota_approved=False)
        create_audit_log(
            request,
            action='professor.reset_selection_status',
            module='导师管理',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='professor',
            target_display='批量重置开放选择资格',
            detail=f'重置 {updated_count} 位导师的开放选择资格。',
            after_data={'ids': ids, 'updated_count': updated_count},
        )
        return Response({'detail': f'已重置 {updated_count} 位导师的“开放选择资格”状态。'}, status=status.HTTP_200_OK)


class DashboardProfessorResetQuotaView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def post(self, request):
        ids = request.data.get('ids') or []
        quota_type = str(request.data.get('quota_type') or '').strip()
        if not ids:
            return Response({'detail': 'Please select at least one professor.'}, status=status.HTTP_400_BAD_REQUEST)
        if quota_type not in {'academic', 'professional', 'professionalyt', 'doctor'}:
            return Response({'detail': 'Invalid quota_type.'}, status=status.HTTP_400_BAD_REQUEST)

        professors = Professor.objects.filter(id__in=ids).prefetch_related('master_quotas__subject', 'doctor_quotas__subject')
        updated_count = 0
        skipped = []

        with transaction.atomic():
            for professor in professors:
                if quota_type == 'academic':
                    for quota in professor.master_quotas.filter(subject__subject_type=1):
                        quota.beijing_quota = 0
                        quota.beijing_remaining_quota = 0
                        quota.save(update_fields=['beijing_quota', 'beijing_remaining_quota'])
                elif quota_type == 'professional':
                    for quota in professor.master_quotas.filter(subject__subject_type=0):
                        quota.beijing_quota = 0
                        quota.beijing_remaining_quota = 0
                        quota.save(update_fields=['beijing_quota', 'beijing_remaining_quota'])
                elif quota_type == 'professionalyt':
                    for quota in professor.master_quotas.filter(subject__subject_type=0):
                        quota.yantai_quota = 0
                        quota.yantai_remaining_quota = 0
                        quota.save(update_fields=['yantai_quota', 'yantai_remaining_quota'])
                else:
                    doctor_quotas = list(professor.doctor_quotas.all())
                    if any((quota.used_quota or 0) > 0 for quota in doctor_quotas):
                        skipped.append(professor.name)
                        continue
                    for quota in doctor_quotas:
                        quota.total_quota = 0
                        quota.remaining_quota = 0
                        quota.save(update_fields=['total_quota', 'remaining_quota'])

                refresh_professor_summary_quotas(professor)
                updated_count += 1

        detail = f'Reset quotas for {updated_count} professor(s).'
        if skipped:
            detail = f"{detail} Skipped {len(skipped)} professor(s) with used doctor quotas."
        create_audit_log(
            request,
            action='professor.reset_quota',
            module='导师管理',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='professor',
            target_display='批量重置导师名额',
            detail=detail,
            after_data={'ids': ids, 'quota_type': quota_type, 'updated_count': updated_count, 'skipped_professors': skipped},
        )
        return Response({'detail': detail, 'skipped_professors': skipped}, status=status.HTTP_200_OK)


"""
class DashboardProfessorImportMasterQuotaView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        upload = get_uploaded_file(request)
        if not upload:
            return Response({'detail': 'Please upload a master quota XLSX file.'}, status=status.HTTP_400_BAD_REQUEST)

        sync_quotas = str(request.data.get('sync_quotas') or 'false').lower() in {'1', 'true', 'yes', 'on'}
        result = import_master_quota_workbook(upload, sync_quotas=sync_quotas)
        return Response(result, status=status.HTTP_200_OK)


class DashboardProfessorImportDoctorQuotaView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        upload = get_uploaded_file(request)
        if not upload:
            return Response({'detail': 'Please upload a doctor quota XLSX file.'}, status=status.HTTP_400_BAD_REQUEST)

        conflict_action = str(request.data.get('conflict_action') or 'replace').strip().lower()
        if conflict_action not in {'replace', 'add'}:
            return Response({'detail': 'conflict_action must be replace or add.'}, status=status.HTTP_400_BAD_REQUEST)

        workbook = load_workbook(upload, data_only=True)
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[1]]
        teacher_id_header = resolve_header_name(headers, ['工号'])
        if teacher_id_header is None:
            return Response({'detail': 'Teacher id column was not found.'}, status=status.HTTP_400_BAD_REQUEST)
        header_map = {header: index for index, header in enumerate(headers)}

        updated_count = 0
        skipped_rows = 0
        touched_professors = set()

        with transaction.atomic():
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                teacher_id = str(row[header_map[teacher_id_header]] or '').strip().zfill(5)
                if not teacher_id:
                    continue

                professor = Professor.objects.filter(teacher_identity_id=teacher_id).first()
                if not professor:
                    skipped_rows += 1
                    continue

                for index in range(1, 10):
                    subject_code_header = resolve_header_name(headers, [f'学科{index}代码', f'专业{index}代码'])
                    quota_type_header = resolve_header_name(headers, [f'专业类型{index}'])
                    quota_header = resolve_header_name(headers, [f'本次名额{index}', f'名额{index}'])
                    if subject_code_header is None or quota_header is None:
                        continue

                    subject_code = normalize_subject_code(row[header_map[subject_code_header]])
                    quota_type = str(row[header_map[quota_type_header]] or '').strip() if quota_type_header else ''
                    quota_value = to_int(row[header_map[quota_header]])
                    if not subject_code or quota_value <= 0:
                        continue
                    if quota_type and quota_type != '博士':
                        continue

                    subject = Subject.objects.filter(subject_code=subject_code, subject_type=2).first()
                    if not subject:
                        skipped_rows += 1
                        continue

                    quota, _ = ProfessorDoctorQuota.objects.get_or_create(professor=professor, subject=subject)
                    if conflict_action == 'replace':
                        if (quota.used_quota or 0) > quota_value:
                            skipped_rows += 1
                            continue
                        quota.total_quota = quota_value
                        quota.remaining_quota = quota_value - (quota.used_quota or 0)
                    else:
                        quota.total_quota = (quota.total_quota or 0) + quota_value
                        quota.remaining_quota = (quota.remaining_quota or 0) + quota_value

                    quota.save(update_fields=['total_quota', 'remaining_quota'])
                    touched_professors.add(professor.id)
                    updated_count += 1

            for professor in Professor.objects.filter(id__in=touched_professors):
                refresh_professor_summary_quotas(professor)

        return Response(
            {
                'detail': f'博士名额导入完成，更新 {updated_count} 条记录，跳过 {skipped_rows} 条。',
                'detail': f'Doctor quota import finished: updated {updated_count}, skipped {skipped_rows}.',
                'updated_count': updated_count,
                'skipped_rows': skipped_rows,
            },
            status=status.HTTP_200_OK,
        )


class DashboardStudentImportView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        upload = get_uploaded_file(request)
        if not upload:
            return Response({'detail': 'Please upload a student import file.'}, status=status.HTTP_400_BAD_REQUEST)

        import_type = str(request.data.get('import_type') or '').strip()
        update_quota = str(request.data.get('update_quota') or 'false').lower() in {'1', 'true', 'yes', 'on'}
        if import_type not in {'doctor', 'master_exam', 'master_recommend'}:
            return Response({'detail': 'Invalid import_type.'}, status=status.HTTP_400_BAD_REQUEST)

        repair_alternate_existing = str(request.data.get('repair_alternate_existing') or 'false').lower() in {'1', 'true', 'yes', 'on'}
        try:
            if import_type == 'doctor':
                result = self._import_doctor_students(upload, update_quota)
            elif import_type == 'master_exam':
                result = self._import_master_exam_students(upload, repair_alternate_existing=repair_alternate_existing)
            else:
                result = self._import_master_students(upload, import_type)
        except ValidationError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(result, status=status.HTTP_200_OK)

    def _import_doctor_students(self, upload, update_quota):
        workbook = load_workbook(upload, data_only=True)
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[1]]
        header_map = {header: index for index, header in enumerate(headers)}

        candidate_header = resolve_header_name(headers, ['考生编号'])
        name_header = resolve_header_name(headers, ['考生姓名', '姓名'])
        subject_header = resolve_header_name(headers, ['报考专业代码', '专业代码'])
        score_header = resolve_header_name(headers, ['复核成绩（满分100分）', '复试成绩', '复核成绩'])
        rank_header = resolve_header_name(headers, ['排名', '综合排名'])
        remark_header = resolve_header_name(headers, ['备注'])
        phone_header = resolve_header_name(headers, ['手机号', '手机号码'])
        if not all([candidate_header, name_header, subject_header, rank_header]):
            raise ValidationError('Doctor import template is missing required columns.')

        subject_quota_map = {}
        rows = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            candidate_number = str(row[header_map[candidate_header]] or '').strip()
            if not candidate_number:
                continue
            subject_code = normalize_subject_code(row[header_map[subject_header]])
            remark = str(row[header_map[remark_header]] or '').strip() if remark_header else ''
            rows.append(
                {
                    'candidate_number': candidate_number,
                    'name': str(row[header_map[name_header]] or '').strip(),
                    'subject_code': subject_code,
                    'score': row[header_map[score_header]] if score_header else None,
                    'rank': to_int(row[header_map[rank_header]], default=0),
                    'remark': remark,
                    'phone': str(row[header_map[phone_header]] or '').strip() if phone_header else '',
                }
            )
            if update_quota and remark == '选导师':
                subject_quota_map[subject_code] = subject_quota_map.get(subject_code, 0) + 1

        if update_quota:
            for subject_code, quota_value in subject_quota_map.items():
                subject = Subject.objects.filter(subject_code=subject_code, subject_type=2).first()
                if subject:
                    subject.total_admission_quota = quota_value
                    subject.save(update_fields=['total_admission_quota'])

        success_count = 0
        skipped_rows = 0
        with transaction.atomic():
            for row in rows:
                subject = Subject.objects.filter(subject_code=row['subject_code'], subject_type=2).first()
                if not subject:
                    skipped_rows += 1
                    continue
                if Student.objects.filter(candidate_number=row['candidate_number']).exists():
                    skipped_rows += 1
                    continue
                if User.objects.filter(username=row['candidate_number']).exists():
                    skipped_rows += 1
                    continue

                password = row['phone'] or row['candidate_number']
                user = User.objects.create_user(username=row['candidate_number'], password=password)
                total_quota = subject.total_admission_quota or 0
                rank = row['rank']
                is_alternate = bool(rank and total_quota and rank > total_quota)
                alternate_rank = rank - total_quota if is_alternate else None
                Student.objects.create(
                    user_name=user,
                    name=row['name'],
                    candidate_number=row['candidate_number'],
                    subject=subject,
                    student_type=3,
                    postgraduate_type=3,
                    study_mode=True,
                    phone_number=row['phone'],
                    secondary_exam_score=float(row['score']) if row['score'] not in (None, '') else None,
                    final_rank=rank or None,
                    is_alternate=is_alternate,
                    alternate_rank=alternate_rank,
                )
                success_count += 1

        return {'detail': f'Doctor student import finished: created {success_count}, skipped {skipped_rows}.'}

    def _import_master_exam_students(self, upload, repair_alternate_existing=False):
        workbook = load_workbook(upload, data_only=True)
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[1]]
        header_map = {header: index for index, header in enumerate(headers)}

        candidate_header = resolve_header_name(headers, ['考生编号'])
        name_header = resolve_header_name(headers, ['考生姓名', '姓名'])
        subject_code_header = resolve_header_name(headers, ['专业代码', '报考专业代码'])
        initial_score_header = resolve_header_name(headers, ['初试总成绩', '初试成绩'])
        secondary_score_header = resolve_header_name(headers, ['复试成绩'])
        final_score_header = resolve_header_name(headers, ['综合成绩'])
        final_rank_header = resolve_header_name(headers, ['排名', '综合排名'])
        campus_header = resolve_header_name(headers, ['校区'])
        choose_professor_header = resolve_header_name(headers, ['选导师'])
        alternate_header = resolve_header_name(headers, ['候补'])
        phone_header = resolve_header_name(headers, ['手机号', '手机号码'])

        required_headers = [
            candidate_header,
            name_header,
            subject_code_header,
            final_rank_header,
            campus_header,
            choose_professor_header,
            phone_header,
        ]
        if not all(required_headers):
            raise ValidationError('Master exam import template is missing required columns.')

        success_count = 0
        skipped_rows = 0

        with transaction.atomic():
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if not any(value not in (None, '') for value in row):
                    continue
                candidate_number = str(row[header_map[candidate_header]] or '').strip()
                name = str(row[header_map[name_header]] or '').strip()
                subject_code = normalize_subject_code(row[header_map[subject_code_header]])
                phone_number = str(row[header_map[phone_header]] or '').strip()
                campus = str(row[header_map[campus_header]] or '').strip()
                choose_professor = str(row[header_map[choose_professor_header]] or '').strip()
                alternate_rank_value = row[header_map[alternate_header]] if alternate_header else None
                alternate_display_value = str(alternate_rank_value or '').strip()
                final_rank = to_int(row[header_map[final_rank_header]], default=0)
                initial_score = row[header_map[initial_score_header]] if initial_score_header else None
                secondary_score = row[header_map[secondary_score_header]] if secondary_score_header else None
                final_score = row[header_map[final_score_header]] if final_score_header else None

                if not candidate_number or not name or not subject_code:
                    skipped_rows += 1
                    continue
                if not phone_number:
                    skipped_rows += 1
                    continue

                subject = Subject.objects.filter(subject_code=subject_code, subject_type__in=[0, 1]).first()
                if not subject:
                    skipped_rows += 1
                    continue

                if subject.subject_type == 1:
                    postgraduate_type = 2
                else:
                    if '烟台' in campus:
                        postgraduate_type = 4
                    elif '北京' in campus:
                        postgraduate_type = 1
                    else:
                        skipped_rows += 1
                        continue

                has_alternate_rank = to_int(alternate_rank_value, default=0) > 0
                is_alternate = choose_professor == '候补' or (choose_professor == '否' and has_alternate_rank)
                lock_selection = choose_professor == '否' and not has_alternate_rank
                alternate_rank = None
                if is_alternate:
                    alternate_rank = to_int(alternate_rank_value, default=0)
                    if alternate_rank <= 0:
                        skipped_rows += 1
                        continue

                if Student.objects.filter(candidate_number=candidate_number).exists():
                    skipped_rows += 1
                    continue
                if User.objects.filter(username=candidate_number).exists():
                    skipped_rows += 1
                    continue

                user = User.objects.create_user(username=candidate_number, password=f'xs{phone_number}')
                Student.objects.create(
                    user_name=user,
                    name=name,
                    candidate_number=candidate_number,
                    subject=subject,
                    student_type=2,
                    postgraduate_type=postgraduate_type,
                    phone_number=phone_number,
                    initial_exam_score=float(initial_score) if initial_score not in (None, '') else None,
                    secondary_exam_score=float(secondary_score) if secondary_score not in (None, '') else None,
                    final_rank=final_rank or None,
                    is_selected=lock_selection,
                    is_alternate=is_alternate,
                    alternate_rank=alternate_rank,
                )
                success_count += 1

        return {'detail': f'硕士统考生导入完成，创建 {success_count} 条，跳过 {skipped_rows} 条。'}

    def _import_master_students(self, upload, import_type):
        reader = csv.DictReader(TextIOWrapper(upload.file, encoding='utf-8-sig'))
        success_count = 0
        skipped_rows = 0

        with transaction.atomic():
            for row in reader:
                subject_code = normalize_subject_code(get_csv_value(row, ['专业代码', '报考专业代码']))
                candidate_number = str(get_csv_value(row, ['考生编号'])).strip()
                name = str(get_csv_value(row, ['姓名', '考生姓名'])).strip()
                phone_number = str(get_csv_value(row, ['手机号', '手机号码'])).strip()
                identify_number = str(get_csv_value(row, ['身份证号', '身份证号码'])).strip() or None
                final_rank = to_int(get_csv_value(row, ['综合排名', '排名']), default=0)
                score = get_csv_value(row, ['综合成绩', '复试成绩'])
                postgraduate_type = to_int(get_csv_value(row, ['研究生类型']), default=0)
                student_type = to_int(get_csv_value(row, ['学生类型']), default=0)

                subject = Subject.objects.filter(subject_code=subject_code).first()
                if not subject or not candidate_number:
                    skipped_rows += 1
                    continue
                if Student.objects.filter(candidate_number=candidate_number).exists():
                    skipped_rows += 1
                    continue
                if User.objects.filter(username=candidate_number).exists():
                    skipped_rows += 1
                    continue

                password = phone_number or candidate_number
                user = User.objects.create_user(username=candidate_number, password=password)
                total_quota = subject.total_admission_quota or 0
                is_alternate = bool(final_rank and total_quota and final_rank > total_quota)
                alternate_rank = final_rank - total_quota if is_alternate else None
                Student.objects.create(
                    user_name=user,
                    name=name,
                    candidate_number=candidate_number,
                    subject=subject,
                    identify_number=identify_number,
                    student_type=student_type or (1 if import_type == 'master_recommend' else 2),
                    postgraduate_type=postgraduate_type or 1,
                    phone_number=phone_number,
                    secondary_exam_score=float(score) if score not in (None, '') else None,
                    final_rank=final_rank or None,
                    is_alternate=is_alternate,
                    alternate_rank=alternate_rank,
                )
                success_count += 1

        label = '硕士推免生' if import_type == 'master_recommend' else '硕士统考生'
        return {'detail': f'{label} import finished: created {success_count}, skipped {skipped_rows}.'}


"""


class DashboardProfessorImportMasterQuotaView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        upload = get_uploaded_file(request)
        if not upload:
            return Response({'detail': 'Please upload a master quota XLSX file.'}, status=status.HTTP_400_BAD_REQUEST)

        sync_quotas = str(request.data.get('sync_quotas') or 'false').lower() in {'1', 'true', 'yes', 'on'}
        result = import_master_quota_workbook(upload, sync_quotas=sync_quotas)
        return Response(result, status=status.HTTP_200_OK)


class DashboardProfessorImportDoctorQuotaView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        upload = get_uploaded_file(request)
        if not upload:
            return Response({'detail': 'Please upload a doctor quota XLSX file.'}, status=status.HTTP_400_BAD_REQUEST)

        conflict_action = str(request.data.get('conflict_action') or 'replace').strip().lower()
        if conflict_action not in {'replace', 'add'}:
            return Response({'detail': 'conflict_action must be replace or add.'}, status=status.HTTP_400_BAD_REQUEST)

        workbook = load_workbook(upload, data_only=True)
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[1]]
        teacher_id_header = resolve_header_name(headers, ['\u5de5\u53f7'])
        if teacher_id_header is None:
            return Response({'detail': 'Teacher id column was not found.'}, status=status.HTTP_400_BAD_REQUEST)
        header_map = {header: index for index, header in enumerate(headers)}

        updated_count = 0
        skipped_rows = 0
        touched_professors = set()

        with transaction.atomic():
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                teacher_id = str(row[header_map[teacher_id_header]] or '').strip().zfill(5)
                if not teacher_id:
                    continue

                professor = Professor.objects.filter(teacher_identity_id=teacher_id).first()
                if not professor:
                    skipped_rows += 1
                    continue

                for index in range(1, 10):
                    subject_code_header = resolve_header_name(headers, [f'\u5b66\u79d1{index}\u4ee3\u7801', f'\u4e13\u4e1a{index}\u4ee3\u7801'])
                    quota_type_header = resolve_header_name(headers, [f'\u4e13\u4e1a\u7c7b\u578b{index}'])
                    quota_header = resolve_header_name(headers, [f'\u672c\u6b21\u540d\u989d{index}', f'\u540d\u989d{index}'])
                    if subject_code_header is None or quota_header is None:
                        continue

                    subject_code = normalize_subject_code(row[header_map[subject_code_header]])
                    quota_type = str(row[header_map[quota_type_header]] or '').strip() if quota_type_header else ''
                    quota_value = to_int(row[header_map[quota_header]])
                    if not subject_code or quota_value <= 0:
                        continue
                    if quota_type and quota_type != '\u535a\u58eb':
                        continue

                    subject = Subject.objects.filter(subject_code=subject_code, subject_type=2).first()
                    if not subject:
                        skipped_rows += 1
                        continue

                    quota, _ = ProfessorDoctorQuota.objects.get_or_create(professor=professor, subject=subject)
                    if conflict_action == 'replace':
                        if (quota.used_quota or 0) > quota_value:
                            skipped_rows += 1
                            continue
                        quota.total_quota = quota_value
                        quota.remaining_quota = quota_value - (quota.used_quota or 0)
                    else:
                        quota.total_quota = (quota.total_quota or 0) + quota_value
                        quota.remaining_quota = (quota.remaining_quota or 0) + quota_value

                    quota.save(update_fields=['total_quota', 'remaining_quota'])
                    touched_professors.add(professor.id)
                    updated_count += 1

            for professor in Professor.objects.filter(id__in=touched_professors):
                refresh_professor_summary_quotas(professor)

        return Response(
            {
                'detail': f'Doctor quota import finished: updated {updated_count}, skipped {skipped_rows}.',
                'updated_count': updated_count,
                'skipped_rows': skipped_rows,
            },
            status=status.HTTP_200_OK,
        )


class DashboardStudentImportView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        upload = get_uploaded_file(request)
        if not upload:
            return Response({'detail': 'Please upload a student import file.'}, status=status.HTTP_400_BAD_REQUEST)

        import_type = str(request.data.get('import_type') or '').strip()
        update_quota = str(request.data.get('update_quota') or 'false').lower() in {'1', 'true', 'yes', 'on'}
        repair_alternate_existing = str(request.data.get('repair_alternate_existing') or 'false').lower() in {'1', 'true', 'yes', 'on'}
        if import_type not in {'doctor', 'master_exam', 'master_recommend'}:
            return Response({'detail': 'Invalid import_type.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            if import_type == 'doctor':
                result = self._import_doctor_students(upload, update_quota)
            elif import_type == 'master_exam':
                result = self._import_master_exam_students(upload, repair_alternate_existing=repair_alternate_existing)
            else:
                result = self._import_master_students(upload, import_type)
        except ValidationError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(result, status=status.HTTP_200_OK)

    def _import_doctor_students(self, upload, update_quota):
        workbook = load_workbook(upload, data_only=True)
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[1]]
        header_map = {header: index for index, header in enumerate(headers)}

        candidate_header = resolve_header_name(headers, ['\u8003\u751f\u7f16\u53f7'])
        name_header = resolve_header_name(headers, ['\u8003\u751f\u59d3\u540d', '\u59d3\u540d'])
        subject_header = resolve_header_name(headers, ['\u62a5\u8003\u4e13\u4e1a\u4ee3\u7801', '\u4e13\u4e1a\u4ee3\u7801'])
        score_header = resolve_header_name(headers, ['\u590d\u6838\u6210\u7ee9\uff08\u6ee1\u5206100\u5206\uff09', '\u590d\u8bd5\u6210\u7ee9', '\u590d\u6838\u6210\u7ee9'])
        rank_header = resolve_header_name(headers, ['\u6392\u540d', '\u7efc\u5408\u6392\u540d'])
        remark_header = resolve_header_name(headers, ['\u5907\u6ce8'])
        phone_header = resolve_header_name(headers, ['\u624b\u673a\u53f7', '\u624b\u673a\u53f7\u7801'])
        if not all([candidate_header, name_header, subject_header, rank_header]):
            raise ValidationError('Doctor import template is missing required columns.')

        subject_quota_map = {}
        rows = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            candidate_number = str(row[header_map[candidate_header]] or '').strip()
            if not candidate_number:
                continue
            subject_code = normalize_subject_code(row[header_map[subject_header]])
            remark = str(row[header_map[remark_header]] or '').strip() if remark_header else ''
            rows.append(
                {
                    'candidate_number': candidate_number,
                    'name': str(row[header_map[name_header]] or '').strip(),
                    'subject_code': subject_code,
                    'score': row[header_map[score_header]] if score_header else None,
                    'rank': to_int(row[header_map[rank_header]], default=0),
                    'remark': remark,
                    'phone': str(row[header_map[phone_header]] or '').strip() if phone_header else '',
                }
            )
            if update_quota and remark == '\u9009\u5bfc\u5e08':
                subject_quota_map[subject_code] = subject_quota_map.get(subject_code, 0) + 1

        if update_quota:
            for subject_code, quota_value in subject_quota_map.items():
                subject = Subject.objects.filter(subject_code=subject_code, subject_type=2).first()
                if subject:
                    subject.total_admission_quota = quota_value
                    subject.save(update_fields=['total_admission_quota'])

        success_count = 0
        skipped_rows = 0
        with transaction.atomic():
            for row in rows:
                subject = Subject.objects.filter(subject_code=row['subject_code'], subject_type=2).first()
                if not subject:
                    skipped_rows += 1
                    continue
                if Student.objects.filter(candidate_number=row['candidate_number']).exists():
                    skipped_rows += 1
                    continue
                if User.objects.filter(username=row['candidate_number']).exists():
                    skipped_rows += 1
                    continue

                password = row['phone'] or row['candidate_number']
                user = User.objects.create_user(username=row['candidate_number'], password=password)
                total_quota = subject.total_admission_quota or 0
                rank = row['rank']
                is_alternate = bool(rank and total_quota and rank > total_quota)
                alternate_rank = rank - total_quota if is_alternate else None
                Student.objects.create(
                    user_name=user,
                    name=row['name'],
                    candidate_number=row['candidate_number'],
                    subject=subject,
                    student_type=3,
                    postgraduate_type=3,
                    study_mode=True,
                    phone_number=row['phone'],
                    secondary_exam_score=float(row['score']) if row['score'] not in (None, '') else None,
                    final_rank=rank or None,
                    is_alternate=is_alternate,
                    alternate_rank=alternate_rank,
                )
                success_count += 1

        return {'detail': f'Doctor student import finished: created {success_count}, skipped {skipped_rows}.'}

    def _import_master_exam_students(self, upload, repair_alternate_existing=False):
        workbook = load_workbook(upload, data_only=True)
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[1]]
        header_map = {header: index for index, header in enumerate(headers)}

        candidate_header = resolve_header_name(headers, ['考生编号'])
        name_header = resolve_header_name(headers, ['考生姓名', '姓名'])
        subject_code_header = resolve_header_name(headers, ['专业代码', '报考专业代码'])
        initial_score_header = resolve_header_name(headers, ['初试总成绩', '初试成绩'])
        secondary_score_header = resolve_header_name(headers, ['复试成绩'])
        final_score_header = resolve_header_name(headers, ['综合成绩'])
        final_rank_header = resolve_header_name(headers, ['排名', '综合排名'])
        campus_header = resolve_header_name(headers, ['校区'])
        choose_professor_header = resolve_header_name(headers, ['选导师'])
        alternate_header = resolve_header_name(headers, ['候补'])
        phone_header = resolve_header_name(headers, ['手机号', '手机号码'])

        required_headers = [
            candidate_header,
            name_header,
            subject_code_header,
            final_rank_header,
            campus_header,
            choose_professor_header,
            phone_header,
        ]
        if not all(required_headers):
            raise ValidationError('Master exam import template is missing required columns.')

        success_count = 0
        skipped_rows = 0
        repaired_count = 0
        subject_summary = {}
        skip_reason_summary = {}
        skipped_examples = []

        def add_skip_reason(reason, row_index, candidate_number='', name=''):
            nonlocal skipped_rows
            skipped_rows += 1
            skip_reason_summary[reason] = skip_reason_summary.get(reason, 0) + 1
            if len(skipped_examples) < 20:
                skipped_examples.append(
                    {
                        'row': row_index,
                        'candidate_number': candidate_number,
                        'name': name,
                        'reason': reason,
                    }
                )

        with transaction.atomic():
            for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(value not in (None, '') for value in row):
                    continue
                candidate_number = str(row[header_map[candidate_header]] or '').strip()
                name = str(row[header_map[name_header]] or '').strip()
                subject_code = normalize_subject_code(row[header_map[subject_code_header]])
                phone_number = str(row[header_map[phone_header]] or '').strip()
                campus = str(row[header_map[campus_header]] or '').strip()
                choose_professor = str(row[header_map[choose_professor_header]] or '').strip()
                alternate_rank_value = row[header_map[alternate_header]] if alternate_header else None
                alternate_display_value = str(alternate_rank_value or '').strip()
                final_rank = to_int(row[header_map[final_rank_header]], default=0)
                initial_score = row[header_map[initial_score_header]] if initial_score_header else None
                secondary_score = row[header_map[secondary_score_header]] if secondary_score_header else None
                final_score = row[header_map[final_score_header]] if final_score_header else None

                if not candidate_number or not name or not subject_code or not phone_number:
                    add_skip_reason('考生编号、姓名、专业代码或手机号为空', row_index, candidate_number, name)
                    continue

                subject = Subject.objects.filter(subject_code=subject_code, subject_type__in=[0, 1]).first()
                if not subject:
                    add_skip_reason(f'专业代码未匹配到硕士专业：{subject_code}', row_index, candidate_number, name)
                    continue

                if subject.subject_type == 1:
                    postgraduate_type = 2
                else:
                    if '烟台' in campus:
                        postgraduate_type = 4
                    elif '北京' in campus:
                        postgraduate_type = 1
                    else:
                        add_skip_reason(f'专硕校区无法识别：{campus or "空"}', row_index, candidate_number, name)
                        continue

                has_alternate_rank = to_int(alternate_rank_value, default=0) > 0
                is_alternate = choose_professor == '候补' or (choose_professor == '否' and has_alternate_rank)
                lock_selection = choose_professor == '否' and not has_alternate_rank
                if choose_professor not in {'是', '候补', '否'}:
                    add_skip_reason(f'选导师列值无法识别：{choose_professor or "空"}', row_index, candidate_number, name)
                    continue
                alternate_rank = None
                if is_alternate:
                    alternate_rank = to_int(alternate_rank_value, default=0)
                    if alternate_rank <= 0:
                        add_skip_reason('候补学生缺少有效候补排名', row_index, candidate_number, name)
                        continue

                existing_student = Student.objects.filter(candidate_number=candidate_number).first()
                if existing_student:
                    if repair_alternate_existing:
                        update_fields = []
                        if existing_student.is_alternate != is_alternate:
                            existing_student.is_alternate = is_alternate
                            update_fields.append('is_alternate')
                        if existing_student.alternate_rank != alternate_rank:
                            existing_student.alternate_rank = alternate_rank
                            update_fields.append('alternate_rank')
                        if existing_student.is_selected != lock_selection:
                            existing_student.is_selected = lock_selection
                            update_fields.append('is_selected')
                        if update_fields:
                            existing_student.save(update_fields=update_fields)
                        repaired_count += 1

                        summary_item = subject_summary.setdefault(
                            subject.id,
                            {
                                'subject_code': subject.subject_code,
                                'subject_name': subject.subject_name,
                                'created_count': 0,
                                'alternate_count': 0,
                                'normal_count': 0,
                                'locked_count': 0,
                                'repaired_count': 0,
                            },
                        )
                        summary_item['repaired_count'] += 1
                        if is_alternate:
                            summary_item['alternate_count'] += 1
                        elif lock_selection:
                            summary_item['locked_count'] += 1
                        else:
                            summary_item['normal_count'] += 1
                        continue

                    add_skip_reason(f'考生编号已存在：{candidate_number}', row_index, candidate_number, name)
                    continue
                if User.objects.filter(username=candidate_number).exists():
                    add_skip_reason(f'登录账号已存在：{candidate_number}', row_index, candidate_number, name)
                    continue

                user = User.objects.create_user(username=candidate_number, password=f'xs{phone_number}')
                Student.objects.create(
                    user_name=user,
                    name=name,
                    candidate_number=candidate_number,
                    subject=subject,
                    student_type=2,
                    postgraduate_type=postgraduate_type,
                    phone_number=phone_number,
                    initial_exam_score=float(initial_score) if initial_score not in (None, '') else None,
                    secondary_exam_score=float(secondary_score) if secondary_score not in (None, '') else None,
                    final_rank=final_rank or None,
                    is_selected=lock_selection,
                    is_alternate=is_alternate,
                    alternate_rank=alternate_rank,
                )
                success_count += 1

                summary_item = subject_summary.setdefault(
                    subject.id,
                    {
                        'subject_code': subject.subject_code,
                        'subject_name': subject.subject_name,
                        'created_count': 0,
                        'alternate_count': 0,
                        'normal_count': 0,
                        'locked_count': 0,
                        'repaired_count': 0,
                    },
                )
                summary_item['created_count'] += 1
                if is_alternate:
                    summary_item['alternate_count'] += 1
                elif lock_selection:
                    summary_item['locked_count'] += 1
                else:
                    summary_item['normal_count'] += 1

        summary = sorted(subject_summary.values(), key=lambda item: (item['subject_code'], item['subject_name']))
        return {
            'detail': f'硕士统考生导入完成，创建 {success_count} 条，修复 {repaired_count} 条，跳过 {skipped_rows} 条。',
            'created_count': success_count,
            'repaired_count': repaired_count,
            'skipped_rows': skipped_rows,
            'summary': summary,
            'skip_reason_summary': skip_reason_summary,
            'skipped_examples': skipped_examples,
        }

    def _import_master_students(self, upload, import_type):
        reader = csv.DictReader(TextIOWrapper(upload.file, encoding='utf-8-sig'))
        success_count = 0
        skipped_rows = 0
        subject_summary = {}

        with transaction.atomic():
            for row in reader:
                subject_code = normalize_subject_code(get_csv_value(row, ['\u4e13\u4e1a\u4ee3\u7801', '\u62a5\u8003\u4e13\u4e1a\u4ee3\u7801']))
                candidate_number = str(get_csv_value(row, ['\u8003\u751f\u7f16\u53f7'])).strip()
                name = str(get_csv_value(row, ['\u59d3\u540d', '\u8003\u751f\u59d3\u540d'])).strip()
                phone_number = str(get_csv_value(row, ['\u624b\u673a\u53f7', '\u624b\u673a\u53f7\u7801'])).strip()
                identify_number = str(get_csv_value(row, ['\u8eab\u4efd\u8bc1\u53f7', '\u8eab\u4efd\u8bc1\u53f7\u7801'])).strip() or None
                final_rank = to_int(get_csv_value(row, ['\u7efc\u5408\u6392\u540d', '\u6392\u540d']), default=0)
                score = get_csv_value(row, ['\u7efc\u5408\u6210\u7ee9', '\u590d\u8bd5\u6210\u7ee9'])
                postgraduate_type = to_int(get_csv_value(row, ['\u7814\u7a76\u751f\u7c7b\u578b']), default=0)
                student_type = to_int(get_csv_value(row, ['\u5b66\u751f\u7c7b\u578b']), default=0)

                subject = Subject.objects.filter(subject_code=subject_code).first()
                if not subject or not candidate_number:
                    skipped_rows += 1
                    continue
                if Student.objects.filter(candidate_number=candidate_number).exists():
                    skipped_rows += 1
                    continue
                if User.objects.filter(username=candidate_number).exists():
                    skipped_rows += 1
                    continue

                password = phone_number or candidate_number
                user = User.objects.create_user(username=candidate_number, password=password)
                total_quota = subject.total_admission_quota or 0
                is_alternate = bool(final_rank and total_quota and final_rank > total_quota)
                alternate_rank = final_rank - total_quota if is_alternate else None
                Student.objects.create(
                    user_name=user,
                    name=name,
                    candidate_number=candidate_number,
                    subject=subject,
                    identify_number=identify_number,
                    student_type=student_type or (1 if import_type == 'master_recommend' else 2),
                    postgraduate_type=postgraduate_type or 1,
                    phone_number=phone_number,
                    secondary_exam_score=float(score) if score not in (None, '') else None,
                    final_rank=final_rank or None,
                    is_alternate=is_alternate,
                    alternate_rank=alternate_rank,
                )
                success_count += 1
                summary_item = subject_summary.setdefault(
                    subject.id,
                    {
                        'subject_code': subject.subject_code,
                        'subject_name': subject.subject_name,
                        'created_count': 0,
                        'alternate_count': 0,
                        'normal_count': 0,
                    },
                )
                summary_item['created_count'] += 1
                if is_alternate:
                    summary_item['alternate_count'] += 1
                else:
                    summary_item['normal_count'] += 1

        label = '\u7855\u58eb\u63a8\u514d\u751f' if import_type == 'master_recommend' else '\u7855\u58eb\u7edf\u8003\u751f'
        summary = sorted(subject_summary.values(), key=lambda item: (item['subject_code'], item['subject_name']))
        return {
            'detail': f'{label} import finished: created {success_count}, skipped {skipped_rows}.',
            'created_count': success_count,
            'skipped_rows': skipped_rows,
            'summary': summary,
        }


class DashboardStudentListView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def get(self, request):
        page = max(int(request.query_params.get('page', 1)), 1)
        page_size = min(max(int(request.query_params.get('page_size', 10)), 1), 100)
        latest_choice_subquery = StudentProfessorChoice.objects.filter(student_id=OuterRef('pk')).order_by('-submit_date')

        queryset = build_student_queryset(request).select_related('subject').annotate(
            latest_choice_id=Subquery(latest_choice_subquery.values('id')[:1]),
        )
        order_by = request.query_params.get('order_by', 'final_rank')
        order_direction = request.query_params.get('order_direction', 'asc')

        order_map = {
            'name': 'name',
            'candidate_number': 'candidate_number',
            'subject_name': 'subject__subject_name',
            'admission_year': 'admission_year',
            'admission_batch': 'admission_batch__name',
            'can_login': 'can_login',
            'selection_display_enabled': 'selection_display_enabled',
            'student_type': 'student_type',
            'postgraduate_type': 'postgraduate_type',
            'final_rank': 'final_rank',
            'signature_table_review_status': 'signature_table_review_status',
            'is_selected': 'is_selected',
            'is_alternate': 'is_alternate',
            'is_giveup': 'is_giveup',
        }
        order_field = order_map.get(order_by, 'final_rank')
        if order_direction == 'desc':
            order_field = f'-{order_field}'
        queryset = queryset.order_by(order_field, 'id')

        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        students = list(queryset[start:end])

        choice_map = {
            choice.student_id: choice
            for choice in StudentProfessorChoice.objects.select_related('professor').filter(
                id__in=[student.latest_choice_id for student in students if student.latest_choice_id]
            )
        }
        for student in students:
            student.latest_choice = choice_map.get(student.latest_choice_id)

        serializer = StudentListSerializer(students, many=True)
        return Response(
            {'count': total, 'page': page, 'page_size': page_size, 'results': serializer.data},
            status=status.HTTP_200_OK,
        )

    def post(self, request):
        try:
            student = save_student_from_payload(serialize_student_form_data(request.data))
        except ValidationError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        create_audit_log(
            request,
            action='student.create',
            module='学生管理',
            target_type='student',
            target_id=student.pk,
            target_display=student.name,
            detail='创建学生。',
            after_data=snapshot_instance(student),
        )
        return Response(StudentDetailSerializer(student).data, status=status.HTTP_201_CREATED)


class DashboardAdmissionBatchListView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def get(self, request):
        queryset = build_admission_batch_queryset(request)
        order_by = request.query_params.get('order_by', 'admission_year')
        order_direction = request.query_params.get('order_direction', 'desc')
        order_map = {
            'name': 'name',
            'admission_year': 'admission_year',
            'batch_code': 'batch_code',
            'sort_order': 'sort_order',
            'is_active': 'is_active',
            'created_at': 'created_at',
        }
        order_field = order_map.get(order_by, 'admission_year')
        if order_direction == 'desc':
            order_field = f'-{order_field}'
        queryset = queryset.order_by(order_field, '-id')
        serializer = AdmissionBatchSerializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        try:
            admission_batch = save_admission_batch_from_payload(request.data)
        except ValidationError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        create_audit_log(
            request,
            action='admission_batch.create',
            module='招生批次',
            target_type='admission_batch',
            target_id=admission_batch.pk,
            target_display=admission_batch.name,
            detail='创建招生批次。',
            after_data=snapshot_instance(admission_batch),
        )
        return Response(AdmissionBatchSerializer(admission_batch).data, status=status.HTTP_201_CREATED)


class DashboardAdmissionBatchDetailView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def patch(self, request, pk):
        admission_batch = AdmissionBatch.objects.filter(pk=pk).first()
        if not admission_batch:
            return Response({'detail': '招生批次不存在。'}, status=status.HTTP_404_NOT_FOUND)
        before_data = snapshot_instance(admission_batch)
        try:
            admission_batch = save_admission_batch_from_payload(request.data, admission_batch=admission_batch)
        except ValidationError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        create_audit_log(
            request,
            action='admission_batch.update',
            module='招生批次',
            target_type='admission_batch',
            target_id=admission_batch.pk,
            target_display=admission_batch.name,
            detail='更新招生批次。',
            before_data=before_data,
            after_data=snapshot_instance(admission_batch),
        )
        return Response(AdmissionBatchSerializer(admission_batch).data, status=status.HTTP_200_OK)

    def delete(self, request, pk):
        admission_batch = AdmissionBatch.objects.filter(pk=pk).first()
        if not admission_batch:
            return Response({'detail': '招生批次不存在。'}, status=status.HTTP_404_NOT_FOUND)
        if admission_batch.students.exists():
            return Response({'detail': '该批次下仍有关联学生，暂不能删除。'}, status=status.HTTP_400_BAD_REQUEST)
        before_data = snapshot_instance(admission_batch)
        target_display = admission_batch.name
        admission_batch.delete()
        create_audit_log(
            request,
            action='admission_batch.delete',
            module='招生批次',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='admission_batch',
            target_id=pk,
            target_display=target_display,
            detail='删除招生批次。',
            before_data=before_data,
        )
        return Response({'detail': '招生批次已删除。'}, status=status.HTTP_200_OK)


class DashboardStudentDetailView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def get(self, request, pk):
        student = Student.objects.select_related('subject', 'admission_batch').filter(pk=pk).first()
        if not student:
            return Response({'detail': 'Student not found.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(StudentDetailSerializer(student).data, status=status.HTTP_200_OK)

    def patch(self, request, pk):
        student = Student.objects.select_related('user_name', 'subject', 'admission_batch').filter(pk=pk).first()
        if not student:
            return Response({'detail': 'Student not found.'}, status=status.HTTP_404_NOT_FOUND)
        before_data = snapshot_instance(student)
        try:
            student = save_student_from_payload(serialize_student_form_data(request.data), student=student)
        except ValidationError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        create_audit_log(
            request,
            action='student.update',
            module='学生管理',
            target_type='student',
            target_id=student.pk,
            target_display=student.name,
            detail='更新学生信息。',
            before_data=before_data,
            after_data=snapshot_instance(student),
        )
        return Response(StudentDetailSerializer(student).data, status=status.HTTP_200_OK)

    def delete(self, request, pk):
        student = Student.objects.select_related('user_name').filter(pk=pk).first()
        if not student:
            return Response({'detail': '学生不存在。'}, status=status.HTTP_404_NOT_FOUND)
        user = student.user_name
        before_data = snapshot_instance(student)
        target_display = student.name
        student.delete()
        if user:
            user.delete()
        create_audit_log(
            request,
            action='student.delete',
            module='学生管理',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='student',
            target_id=pk,
            target_display=target_display,
            detail='删除学生。',
            before_data=before_data,
        )
        return Response({'detail': '学生已删除。'}, status=status.HTTP_200_OK)


class DashboardAvailableStudentDisplaySettingView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def get(self, request):
        setting = get_available_student_display_setting()
        return Response(AvailableStudentDisplaySettingSerializer(setting).data, status=status.HTTP_200_OK)

    def patch(self, request):
        setting = get_available_student_display_setting()
        before_data = snapshot_instance(setting)
        setting = save_available_student_display_setting_from_payload(request.data)
        create_audit_log(
            request,
            action='student_display_setting.update',
            module='可选学生展示控制',
            target_type='available_student_display_setting',
            target_id=setting.pk,
            target_display='可选学生展示配置',
            detail='更新可选学生展示配置。',
            before_data=before_data,
            after_data=snapshot_instance(setting),
        )
        return Response(AvailableStudentDisplaySettingSerializer(setting).data, status=status.HTTP_200_OK)


class DashboardStudentBatchDeleteView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def post(self, request):
        ids, delete_all_filtered = parse_batch_delete_request(request)
        queryset = build_student_queryset(request).select_related('user_name')
        if not delete_all_filtered:
            queryset = queryset.filter(id__in=ids)
        if not queryset.exists():
            return Response({'detail': '没有可删除的学生记录。'}, status=status.HTTP_400_BAD_REQUEST)
        users = [student.user_name for student in queryset if student.user_name_id]
        target_names = list(queryset.values_list('name', flat=True))
        deleted_count = queryset.count()
        queryset.delete()
        for user in users:
            User.objects.filter(pk=user.pk).delete()
        create_audit_log(
            request,
            action='student.batch_delete',
            module='学生管理',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='student',
            target_display='批量删除学生',
            detail=f'批量删除 {deleted_count} 名学生。',
            before_data={'count': deleted_count, 'names': target_names, 'delete_all_filtered': delete_all_filtered},
        )
        return Response({'detail': f'已删除 {deleted_count} 名学生。'}, status=status.HTTP_200_OK)


class DashboardStudentPromoteAlternateView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def post(self, request, pk):
        student = Student.objects.select_related('subject').filter(pk=pk).first()
        if not student:
            return Response({'detail': 'Student not found.'}, status=status.HTTP_404_NOT_FOUND)
        if not student.is_alternate:
            return Response({'detail': 'Student is not currently marked as alternate.'}, status=status.HTTP_400_BAD_REQUEST)

        student.is_alternate = False
        student.alternate_rank = None
        student.save(update_fields=['is_alternate', 'alternate_rank'])
        normalize_alternate_ranks(student.subject, student.admission_year)
        create_audit_log(
            request,
            action='student.clear_alternate',
            module='候补管理',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='student',
            target_id=student.pk,
            target_display=student.name,
            detail='取消学生候补状态。',
            before_data={'is_alternate': True},
            after_data={'is_alternate': False, 'alternate_rank': None},
        )
        return Response({'detail': '已取消该学生的候补状态。'}, status=status.HTTP_200_OK)


class DashboardStudentRevokeGiveupView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def post(self, request, pk):
        student = Student.objects.select_related('subject').filter(pk=pk).first()
        if not student:
            return Response({'detail': 'Student not found.'}, status=status.HTTP_404_NOT_FOUND)
        if not student.is_giveup:
            return Response({'detail': 'Student has not given up admission.'}, status=status.HTTP_400_BAD_REQUEST)

        revoked_choice = (
            StudentProfessorChoice.objects.select_related(
                'student__subject',
                'professor__department',
                'shared_quota_pool',
            )
            .filter(student=student, status=5, chosen_by_professor=True)
            .order_by('-finish_time', '-submit_date', '-id')
            .first()
        )

        with transaction.atomic():
            if revoked_choice and not student.is_selected:
                try:
                    reinstate_revoked_choice(revoked_choice)
                except ValidationError as exc:
                    return Response({'detail': str(exc)}, status=status.HTTP_409_CONFLICT)
            else:
                student.is_giveup = False
                student.save(update_fields=['is_giveup'])
            normalize_alternate_ranks(student.subject, student.admission_year)
        create_audit_log(
            request,
            action='student.revoke_giveup',
            module='放弃录取',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='student',
            target_id=student.pk,
            target_display=student.name,
            detail='撤销学生放弃录取状态。',
            before_data={'is_giveup': True},
            after_data={'is_giveup': False, 'restored_choice_id': revoked_choice.id if revoked_choice else None},
        )
        if revoked_choice:
            return Response({'detail': '已撤销该学生的放弃录取状态，并恢复原录取名额。'}, status=status.HTTP_200_OK)
        return Response({'detail': '已撤销该学生的放弃录取状态。'}, status=status.HTTP_200_OK)


class DashboardStudentResetPasswordView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def post(self, request):
        ids = request.data.get('ids') or []
        queryset = Student.objects.select_related('user_name').filter(id__in=ids)
        updated_count = 0

        for student in queryset:
            if student.user_name:
                student.user_name.set_password(student.candidate_number)
                student.user_name.save(update_fields=['password'])
                updated_count += 1
        create_audit_log(
            request,
            action='student.reset_password',
            module='学生管理',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='student',
            target_display='批量重置学生密码',
            detail=f'重置 {updated_count} 位学生密码。',
            after_data={'ids': ids, 'updated_count': updated_count},
        )
        return Response({'detail': f'已重置 {updated_count} 位学生的密码。'}, status=status.HTTP_200_OK)


class DashboardStudentToggleLoginView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def post(self, request):
        ids = request.data.get('ids') or []
        can_login = parse_bool_param(request.data.get('can_login'))

        if can_login is None:
            return Response({'detail': '请指定目标登录状态。'}, status=status.HTTP_400_BAD_REQUEST)
        if not isinstance(ids, list) or not ids:
            return Response({'detail': '请先选择学生。'}, status=status.HTTP_400_BAD_REQUEST)

        queryset = Student.objects.select_related('user_name').filter(id__in=ids)
        if not queryset.exists():
            return Response({'detail': '未找到可操作的学生。'}, status=status.HTTP_400_BAD_REQUEST)

        changed_students = []
        unchanged_count = 0

        for student in queryset:
            if student.can_login == can_login:
                unchanged_count += 1
                continue
            before_can_login = student.can_login
            student.can_login = can_login
            student.save(update_fields=['can_login'])
            changed_students.append(
                {
                    'id': student.id,
                    'name': student.name,
                    'candidate_number': student.candidate_number,
                    'before_can_login': before_can_login,
                    'after_can_login': can_login,
                }
            )

        if not changed_students:
            return Response(
                {'detail': '所选学生的登录状态已经是目标状态。', 'updated_count': 0, 'unchanged_count': unchanged_count},
                status=status.HTTP_200_OK,
            )

        action_text = '开启' if can_login else '关闭'
        create_audit_log(
            request,
            action='student.toggle_login',
            module='学生管理',
            level=DashboardAuditLog.LEVEL_WARNING if not can_login else DashboardAuditLog.LEVEL_INFO,
            target_type='student',
            target_display=f'批量{action_text}学生登录',
            detail=f'批量{action_text} {len(changed_students)} 名学生的小程序登录权限。',
            before_data={'ids': ids},
            after_data={
                'can_login': can_login,
                'updated_count': len(changed_students),
                'unchanged_count': unchanged_count,
                'students': changed_students,
            },
        )
        return Response(
            {
                'detail': f'已{action_text} {len(changed_students)} 名学生的小程序登录权限。',
                'updated_count': len(changed_students),
                'unchanged_count': unchanged_count,
            },
            status=status.HTTP_200_OK,
        )


class DashboardStudentToggleSelectionDisplayView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def post(self, request):
        ids = request.data.get('ids') or []
        selection_display_enabled = parse_bool_param(request.data.get('selection_display_enabled'))

        if selection_display_enabled is None:
            return Response({'detail': '请指定目标展示状态。'}, status=status.HTTP_400_BAD_REQUEST)
        if not isinstance(ids, list) or not ids:
            return Response({'detail': '请先选择学生。'}, status=status.HTTP_400_BAD_REQUEST)

        queryset = Student.objects.filter(id__in=ids)
        if not queryset.exists():
            return Response({'detail': '未找到可操作的学生。'}, status=status.HTTP_400_BAD_REQUEST)

        changed_students = []
        unchanged_count = 0

        for student in queryset:
            if student.selection_display_enabled == selection_display_enabled:
                unchanged_count += 1
                continue
            before_value = student.selection_display_enabled
            student.selection_display_enabled = selection_display_enabled
            student.save(update_fields=['selection_display_enabled'])
            changed_students.append(
                {
                    'id': student.id,
                    'name': student.name,
                    'candidate_number': student.candidate_number,
                    'before_selection_display_enabled': before_value,
                    'after_selection_display_enabled': selection_display_enabled,
                }
            )

        if not changed_students:
            return Response(
                {'detail': '所选学生的展示状态已经是目标状态。', 'updated_count': 0, 'unchanged_count': unchanged_count},
                status=status.HTTP_200_OK,
            )

        action_text = '开启' if selection_display_enabled else '关闭'
        create_audit_log(
            request,
            action='student.toggle_selection_display',
            module='可选学生展示控制',
            level=DashboardAuditLog.LEVEL_WARNING if not selection_display_enabled else DashboardAuditLog.LEVEL_INFO,
            target_type='student',
            target_display=f'批量{action_text}学生展示',
            detail=f'批量{action_text} {len(changed_students)} 名学生的可选学生池展示状态。',
            before_data={'ids': ids},
            after_data={
                'selection_display_enabled': selection_display_enabled,
                'updated_count': len(changed_students),
                'unchanged_count': unchanged_count,
                'students': changed_students,
            },
        )
        return Response(
            {
                'detail': f'已{action_text} {len(changed_students)} 名学生的可选学生池展示状态。',
                'updated_count': len(changed_students),
                'unchanged_count': unchanged_count,
            },
            status=status.HTTP_200_OK,
        )


class DashboardStudentBatchDownloadView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def post(self, request):
        ids = request.data.get('ids') or []
        file_type = request.data.get('file_type')
        if file_type not in {'signature', 'giveup'}:
            return Response({'detail': 'file_type 参数不合法。'}, status=status.HTTP_400_BAD_REQUEST)

        queryset = Student.objects.filter(id__in=ids).order_by('id')
        buffer = io.BytesIO()
        downloaded_count = 0

        with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for student in queryset:
                file_id = student.signature_table if file_type == 'signature' else student.giveup_signature_table
                if not file_id:
                    continue

                try:
                    download_url = get_file_download_url(file_id) if not str(file_id).startswith(('http://', 'https://')) else file_id
                    file_response = requests.get(download_url, timeout=20)
                    file_response.raise_for_status()
                    suffix = 'pdf'
                    zip_file.writestr(f'{student.candidate_number}_{student.name}.{suffix}', file_response.content)
                    downloaded_count += 1
                except Exception:
                    continue

        if downloaded_count == 0:
            return Response({'detail': '没有可下载的文件。'}, status=status.HTTP_400_BAD_REQUEST)

        response = HttpResponse(buffer.getvalue(), content_type='application/zip')
        filename = '学生互选表.zip' if file_type == 'signature' else '学生放弃说明表.zip'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        create_audit_log(
            request,
            action='student.batch_download',
            module='材料访问',
            target_type='student',
            target_display=filename,
            detail=f'批量下载 {downloaded_count} 份学生材料。',
            after_data={'ids': ids, 'file_type': file_type, 'downloaded_count': downloaded_count},
        )
        return response


class DashboardChoiceListView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def get(self, request):
        page = max(int(request.query_params.get('page', 1)), 1)
        page_size = min(max(int(request.query_params.get('page_size', 10)), 1), 100)
        queryset = build_choice_queryset(request)
        order_by = request.query_params.get('order_by', 'submit_date')
        order_direction = request.query_params.get('order_direction', 'desc')

        order_map = {
            'student_name': 'student__name',
            'candidate_number': 'student__candidate_number',
            'professor_name': 'professor__name',
            'professor_teacher_identity_id': 'professor__teacher_identity_id',
            'department_name': 'professor__department__department_name',
            'subject_name': 'student__subject__subject_name',
            'status': 'status',
            'submit_date': 'submit_date',
            'finish_time': 'finish_time',
        }
        order_field = order_map.get(order_by, 'submit_date')
        if order_direction == 'desc':
            order_field = f'-{order_field}'
        queryset = queryset.order_by(order_field, '-id')

        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        serializer = ChoiceListSerializer(queryset[start:end], many=True)
        return Response(
            {'count': total, 'page': page, 'page_size': page_size, 'results': serializer.data},
            status=status.HTTP_200_OK,
        )


class DashboardChoiceCancelApprovedView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def post(self, request):
        ids = request.data.get('ids') or []
        queryset = StudentProfessorChoice.objects.select_related(
            'student__subject',
            'professor__department',
        ).filter(id__in=ids, status=1)
        if not queryset.exists():
            return Response({'detail': 'No approved choices were selected.'}, status=status.HTTP_400_BAD_REQUEST)

        cancelled_count = 0
        with transaction.atomic():
            for choice in queryset:
                cancel_approved_choice(choice)
                cancelled_count += 1
        create_audit_log(
            request,
            action='choice.cancel_approved',
            module='双选记录',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='choice',
            target_display='撤销已同意双选记录',
            detail=f'撤销 {cancelled_count} 条已同意双选记录。',
            after_data={'ids': ids, 'cancelled_count': cancelled_count},
        )
        return Response({'detail': f'已撤销 {cancelled_count} 条已同意的双选记录。'}, status=status.HTTP_200_OK)


class DashboardChoiceBatchDeleteView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def post(self, request):
        ids, delete_all_filtered = parse_batch_delete_request(request)
        queryset = build_choice_queryset(request)
        if not delete_all_filtered:
            queryset = queryset.filter(id__in=ids)
        if not queryset.exists():
            return Response({'detail': '没有可删除的双选记录。'}, status=status.HTTP_400_BAD_REQUEST)
        deleted_count = queryset.count()
        deleted_ids = list(queryset.values_list('id', flat=True))
        queryset.delete()
        create_audit_log(
            request,
            action='choice.batch_delete',
            module='双选记录',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='choice',
            target_display='批量删除双选记录',
            detail=f'删除 {deleted_count} 条双选记录。',
            before_data={'ids': deleted_ids, 'delete_all_filtered': delete_all_filtered},
        )
        return Response({'detail': f'已删除 {deleted_count} 条双选记录。'}, status=status.HTTP_200_OK)


class DashboardChoiceRejectWaitingNoQuotaView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def post(self, request):
        rejected_count = reject_waiting_choices_without_quota()
        create_audit_log(
            request,
            action='choice.reject_waiting_no_quota',
            module='双选记录',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='choice',
            target_display='无名额待处理申请',
            detail=f'自动拒绝 {rejected_count} 条无名额待处理申请。',
            after_data={'rejected_count': rejected_count},
        )
        return Response({'detail': f'已自动拒绝 {rejected_count} 条无名额的待处理申请。'}, status=status.HTTP_200_OK)


class DashboardChoiceCancelWaitingGiveupView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def post(self, request):
        cancelled_count = cancel_waiting_choices_for_giveup_students()
        create_audit_log(
            request,
            action='choice.cancel_waiting_giveup',
            module='双选记录',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='choice',
            target_display='放弃录取导致失效的待处理申请',
            detail=f'取消 {cancelled_count} 条待处理申请。',
            after_data={'cancelled_count': cancelled_count},
        )
        return Response({'detail': f'已取消 {cancelled_count} 条因学生放弃而失效的待处理申请。'}, status=status.HTTP_200_OK)


class DashboardChoiceExportSelectedView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def get(self, request):
        queryset = StudentProfessorChoice.objects.select_related('student__subject', 'professor').filter(
            status=1,
            chosen_by_professor=True,
        )

        student_type_map = {
            1: '硕士推免生',
            2: '硕士统考生',
            3: '博士统考生',
        }

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = '互选结果'
        worksheet.append([
            '专业代码',
            '专业名称',
            '考生编号',
            '姓名',
            '初试成绩',
            '复试成绩',
            '综合排名',
            '培养层次',
            '培养类型',
            '手机号',
            '导师',
            '招生批次',
        ])

        for choice in queryset:
            student = choice.student
            professor = choice.professor
            level = '博士' if student.postgraduate_type == 3 else '硕士'
            degree_type = '学术学位' if student.postgraduate_type in [2, 3] else '专业学位'
            worksheet.append([
                student.subject.subject_code if student.subject else '',
                student.subject.subject_name if student.subject else '',
                student.candidate_number,
                student.name,
                student.initial_exam_score,
                student.secondary_exam_score,
                student.final_rank,
                level,
                degree_type,
                student.phone_number,
                professor.name if professor else '',
                student_type_map.get(student.student_type, str(student.student_type)),
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="selected-choices.xlsx"'
        workbook.save(response)
        create_audit_log(
            request,
            action='choice.export_selected',
            module='双选记录',
            target_type='choice',
            target_display='已录取双选导出',
            detail='导出已同意且已选中的双选记录。',
        )
        return response


class DashboardReviewRecordListView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def get(self, request):
        page = max(int(request.query_params.get('page', 1)), 1)
        page_size = min(max(int(request.query_params.get('page_size', 10)), 1), 100)
        queryset = build_review_record_queryset(request)
        order_by = request.query_params.get('order_by', 'submit_time')
        order_direction = request.query_params.get('order_direction', 'desc')

        order_map = {
            'student_name': 'student__name',
            'candidate_number': 'student__candidate_number',
            'professor_name': 'professor__name',
            'reviewer_name': 'reviewer__name',
            'subject_name': 'student__subject__subject_name',
            'status': 'status',
            'submit_time': 'submit_time',
            'review_time': 'review_time',
        }
        order_field = order_map.get(order_by, 'submit_time')
        if order_direction == 'desc':
            order_field = f'-{order_field}'
        queryset = queryset.order_by(order_field, '-id')

        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        serializer = ReviewRecordListSerializer(queryset[start:end], many=True)
        return Response(
            {'count': total, 'page': page, 'page_size': page_size, 'results': serializer.data},
            status=status.HTTP_200_OK,
        )


class DashboardReviewRecordDetailView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def get(self, request, pk):
        review_record = ReviewRecord.objects.select_related(
            'student__subject',
            'professor__department',
            'reviewer',
        ).prefetch_related(
            'professor__master_quotas__subject',
            'professor__doctor_quotas__subject',
        ).filter(pk=pk).first()
        if not review_record:
            return Response({'detail': 'Review record not found.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(ReviewRecordDetailSerializer(review_record).data, status=status.HTTP_200_OK)

    def delete(self, request, pk):
        review_record = ReviewRecord.objects.filter(pk=pk).first()
        if not review_record:
            return Response({'detail': '审核记录不存在。'}, status=status.HTTP_404_NOT_FOUND)
        before_data = snapshot_instance(review_record)
        review_record.delete()
        create_audit_log(
            request,
            action='review.delete',
            module='审核记录',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='review_record',
            target_id=pk,
            target_display=f'审核记录 #{pk}',
            detail='删除审核记录。',
            before_data=before_data,
        )
        return Response({'detail': '审核记录已删除。'}, status=status.HTTP_200_OK)


class DashboardReviewRecordBatchDeleteView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def post(self, request):
        ids, delete_all_filtered = parse_batch_delete_request(request)
        queryset = build_review_record_queryset(request)
        if not delete_all_filtered:
            queryset = queryset.filter(id__in=ids)
        if not queryset.exists():
            return Response({'detail': '没有可删除的审核记录。'}, status=status.HTTP_400_BAD_REQUEST)
        deleted_count = queryset.count()
        deleted_ids = list(queryset.values_list('id', flat=True))
        queryset.delete()
        create_audit_log(
            request,
            action='review.batch_delete',
            module='审核记录',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='review_record',
            target_display='批量删除审核记录',
            detail=f'删除 {deleted_count} 条审核记录。',
            before_data={'ids': deleted_ids, 'delete_all_filtered': delete_all_filtered},
        )
        return Response({'detail': f'已删除 {deleted_count} 条审核记录。'}, status=status.HTTP_200_OK)


class DashboardReviewRecordBatchApproveView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def post(self, request):
        ids = request.data.get('ids') or []
        queryset = ReviewRecord.objects.filter(status=3)
        if ids:
            queryset = queryset.filter(id__in=ids)
        student_ids = list(queryset.values_list('student_id', flat=True))
        updated_count = queryset.update(status=1, review_status=False, review_time=timezone.now())
        if student_ids:
            Student.objects.filter(id__in=student_ids).update(signature_table_review_status=1)
        create_audit_log(
            request,
            action='review.batch_approve',
            module='审核记录',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='review_record',
            target_display='批量通过审核记录',
            detail=f'批量通过 {updated_count} 条审核记录。',
            after_data={'ids': ids, 'updated_count': updated_count},
        )
        return Response({'detail': f'已批量通过 {updated_count} 条审核记录。'}, status=status.HTTP_200_OK)


class DashboardDepartmentListView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def get(self, request):
        queryset = build_department_queryset(request)
        order_by = request.query_params.get('order_by', 'id')
        order_direction = request.query_params.get('order_direction', 'asc')
        order_map = {
            'department_name': 'department_name',
            'subject_count': 'subject_count',
            'total_academic_quota': 'total_academic_quota',
            'total_professional_quota': 'total_professional_quota',
            'total_professional_yt_quota': 'total_professional_yt_quota',
            'total_doctor_quota': 'total_doctor_quota',
        }
        order_field = order_map.get(order_by, 'id')
        if order_direction == 'desc':
            order_field = f'-{order_field}'
        queryset = queryset.order_by(order_field, 'id')
        serializer = DepartmentListSerializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        try:
            department = save_department_from_payload(request.data)
        except ValidationError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        create_audit_log(
            request,
            action='department.create',
            module='专业方向',
            target_type='department',
            target_id=department.pk,
            target_display=department.department_name,
            detail='创建方向。',
            after_data=snapshot_instance(department),
        )
        return Response(DepartmentListSerializer(department).data, status=status.HTTP_201_CREATED)


class DashboardDepartmentBatchDeleteView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def post(self, request):
        ids, delete_all_filtered = parse_batch_delete_request(request)
        queryset = build_department_queryset(request)
        if not delete_all_filtered:
            queryset = queryset.filter(id__in=ids)
        if not queryset.exists():
            return Response({'detail': '没有可删除的方向记录。'}, status=status.HTTP_400_BAD_REQUEST)
        user_ids = list(
            User.objects.filter(professor__department__in=queryset).values_list('id', flat=True).distinct()
        )
        names = list(queryset.values_list('department_name', flat=True))
        deleted_count = queryset.count()
        queryset.delete()
        if user_ids:
            User.objects.filter(id__in=user_ids).delete()
        create_audit_log(
            request,
            action='department.batch_delete',
            module='专业方向',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='department',
            target_display='批量删除方向',
            detail=f'删除 {deleted_count} 个方向。',
            before_data={'names': names, 'delete_all_filtered': delete_all_filtered},
        )
        return Response({'detail': f'已删除 {deleted_count} 个方向。'}, status=status.HTTP_200_OK)


class DashboardDepartmentQuotaUpdateView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def patch(self, request, pk):
        department = Department.objects.filter(pk=pk).first()
        if not department:
            return Response({'detail': 'Department not found.'}, status=status.HTTP_404_NOT_FOUND)

        before_data = snapshot_instance(department)
        try:
            save_department_from_payload(request.data, department=department)
        except ValidationError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        department.refresh_from_db()
        create_audit_log(
            request,
            action='department.update',
            module='专业方向',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='department',
            target_id=department.pk,
            target_display=department.department_name,
            detail='更新方向信息和名额。',
            before_data=before_data,
            after_data=snapshot_instance(department),
        )
        return Response({'detail': '方向名额更新成功。'}, status=status.HTTP_200_OK)


class DashboardSubjectListView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def get(self, request):
        page = max(int(request.query_params.get('page', 1)), 1)
        page_size = min(max(int(request.query_params.get('page_size', 20)), 1), 100)
        queryset = build_subject_queryset(request)
        order_by = request.query_params.get('order_by', 'subject_code')
        order_direction = request.query_params.get('order_direction', 'asc')

        order_map = {
            'subject_code': 'subject_code',
            'subject_name': 'subject_name',
            'subject_type': 'subject_type',
            'total_admission_quota': 'total_admission_quota',
            'student_count': 'student_count',
            'selected_student_count': 'selected_student_count',
            'alternate_student_count': 'alternate_student_count',
            'giveup_student_count': 'giveup_student_count',
        }
        order_field = order_map.get(order_by, 'subject_code')
        if order_direction == 'desc':
            order_field = f'-{order_field}'
        queryset = queryset.order_by(order_field, 'id')

        subjects = list(queryset)
        subject_ids = [subject.id for subject in subjects]

        master_quota_map = {
            item['subject']: item['total']
            for item in ProfessorMasterQuota.objects.filter(subject_id__in=subject_ids).values('subject').annotate(
                total=Coalesce(Sum('total_quota'), Value(0))
            )
        }
        doctor_quota_map = {
            item['subject']: item['total']
            for item in ProfessorDoctorQuota.objects.filter(subject_id__in=subject_ids).values('subject').annotate(
                total=Coalesce(Sum('total_quota'), Value(0))
            )
        }
        shared_quota_map = {
            item['subjects']: item['total']
            for item in ProfessorSharedQuotaPool.objects.filter(subjects__id__in=subject_ids, is_active=True).values('subjects').annotate(
                total=Coalesce(Sum('total_quota'), Value(0))
            )
        }
        shared_quota_labels_map = {}
        for pool in ProfessorSharedQuotaPool.objects.select_related('professor').prefetch_related('subjects').filter(subjects__id__in=subject_ids, is_active=True):
            subject_names = [subject.subject_name for subject in pool.subjects.all()]
            label = f"{pool.pool_name} / {pool.professor.name} / 剩余 {pool.remaining_quota}/{pool.total_quota}"
            for subject in pool.subjects.all():
                shared_quota_labels_map.setdefault(subject.id, []).append(label)

        for subject in subjects:
            fixed_total = doctor_quota_map.get(subject.id, 0) if subject.subject_type == 2 else master_quota_map.get(subject.id, 0)
            subject.assigned_quota_total = fixed_total + shared_quota_map.get(subject.id, 0)
            labels = shared_quota_labels_map.get(subject.id, [])
            subject.shared_quota_pool_count = len(labels)
            subject.shared_quota_pool_labels = labels

        total = len(subjects)
        start = (page - 1) * page_size
        end = start + page_size
        serializer = SubjectListSerializer(subjects[start:end], many=True)
        return Response(
            {'count': total, 'page': page, 'page_size': page_size, 'results': serializer.data},
            status=status.HTTP_200_OK,
        )

    def post(self, request):
        try:
            subject = save_subject_from_payload(request.data)
        except ValidationError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        create_audit_log(
            request,
            action='subject.create',
            module='专业方向',
            target_type='subject',
            target_id=subject.pk,
            target_display=subject.subject_name,
            detail='创建专业。',
            after_data=snapshot_instance(subject),
        )
        return Response(SubjectListSerializer(subject).data, status=status.HTTP_201_CREATED)


class DashboardSubjectBatchDeleteView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def post(self, request):
        ids, delete_all_filtered = parse_batch_delete_request(request)
        queryset = build_subject_queryset(request)
        if not delete_all_filtered:
            queryset = queryset.filter(id__in=ids)
        if not queryset.exists():
            return Response({'detail': '没有可删除的专业记录。'}, status=status.HTTP_400_BAD_REQUEST)
        deleted_count = queryset.count()
        names = list(queryset.values_list('subject_name', flat=True))
        queryset.delete()
        create_audit_log(
            request,
            action='subject.batch_delete',
            module='专业方向',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='subject',
            target_display='批量删除专业',
            detail=f'删除 {deleted_count} 个专业。',
            before_data={'names': names, 'delete_all_filtered': delete_all_filtered},
        )
        return Response({'detail': f'已删除 {deleted_count} 个专业。'}, status=status.HTTP_200_OK)


class DashboardSubjectQuotaUpdateView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def patch(self, request, pk):
        subject = Subject.objects.filter(pk=pk).first()
        if not subject:
            return Response({'detail': 'Subject not found.'}, status=status.HTTP_404_NOT_FOUND)

        old_quota = subject.total_admission_quota or 0
        before_data = snapshot_instance(subject)
        try:
            subject = save_subject_from_payload(request.data, subject=subject)
        except ValidationError as exc:
            message = exc.messages[0] if hasattr(exc, 'messages') and exc.messages else str(exc)
            return Response({'detail': message}, status=status.HTTP_400_BAD_REQUEST)

        updated_count = sync_student_alternate_status(subject)
        normalize_alternate_ranks(subject)
        create_audit_log(
            request,
            action='subject.update',
            module='专业方向',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='subject',
            target_id=subject.pk,
            target_display=subject.subject_name,
            detail='更新专业信息和总名额。',
            before_data=before_data,
            after_data=snapshot_instance(subject, extra={'alternate_updates': updated_count}),
        )
        return Response(
            {
                'detail': 'Subject updated successfully.',
                'old_quota': old_quota,
                'new_quota': subject.total_admission_quota,
                'alternate_updates': updated_count,
            },
            status=status.HTTP_200_OK,
        )
        """
        return Response(
            {
                'detail': '专业总名额更新成功。',
                'old_quota': old_quota,
                'new_quota': new_quota,
                'alternate_updates': updated_count,
            },
            status=status.HTTP_200_OK,
        )
        """


class DashboardDoctorQuotaListView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def get(self, request):
        page = max(int(request.query_params.get('page', 1)), 1)
        page_size = min(max(int(request.query_params.get('page_size', 20)), 1), 200)
        queryset = ProfessorDoctorQuota.objects.select_related('professor__department', 'subject')

        search = request.query_params.get('search', '').strip()
        professor_id = request.query_params.get('professor_id')
        subject_id = request.query_params.get('subject_id')
        department_id = request.query_params.get('department_id')
        order_by = request.query_params.get('order_by', 'teacher_identity_id')
        order_direction = request.query_params.get('order_direction', 'asc')

        if search:
            queryset = queryset.filter(
                Q(professor__name__icontains=search) |
                Q(professor__teacher_identity_id__icontains=search) |
                Q(subject__subject_name__icontains=search) |
                Q(subject__subject_code__icontains=search)
            )
        if professor_id:
            queryset = queryset.filter(professor_id=professor_id)
        if subject_id:
            queryset = queryset.filter(subject_id=subject_id)
        if department_id:
            queryset = queryset.filter(professor__department_id=department_id)

        order_map = {
            'professor_name': 'professor__name',
            'teacher_identity_id': 'professor__teacher_identity_id',
            'subject_name': 'subject__subject_name',
            'subject_code': 'subject__subject_code',
            'total_quota': 'total_quota',
            'used_quota': 'used_quota',
            'remaining_quota': 'remaining_quota',
        }
        order_field = order_map.get(order_by, 'professor__teacher_identity_id')
        if order_direction == 'desc':
            order_field = f'-{order_field}'
        queryset = queryset.order_by(order_field, 'id')

        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        serializer = ProfessorDoctorQuotaSerializer(queryset[start:end], many=True)
        return Response({'count': total, 'page': page, 'page_size': page_size, 'results': serializer.data}, status=status.HTTP_200_OK)

    def post(self, request):
        try:
            quota = save_doctor_quota_from_payload(request.data)
        except ValidationError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        create_audit_log(
            request,
            action='doctor_quota.create',
            module='导师博士专业名额',
            target_type='doctor_quota',
            target_id=quota.pk,
            target_display=f'{quota.professor.name} - {quota.subject.subject_name}',
            detail='创建导师博士专业名额。',
            after_data=snapshot_instance(quota),
        )
        return Response(ProfessorDoctorQuotaSerializer(quota).data, status=status.HTTP_201_CREATED)


class DashboardDoctorQuotaDetailView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def patch(self, request, pk):
        quota = ProfessorDoctorQuota.objects.filter(pk=pk).first()
        if not quota:
            return Response({'detail': '博士专业名额不存在。'}, status=status.HTTP_404_NOT_FOUND)
        before_data = snapshot_instance(quota)
        try:
            quota = save_doctor_quota_from_payload(request.data, quota=quota)
        except ValidationError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        create_audit_log(
            request,
            action='doctor_quota.update',
            module='导师博士专业名额',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='doctor_quota',
            target_id=quota.pk,
            target_display=f'{quota.professor.name} - {quota.subject.subject_name}',
            detail='更新导师博士专业名额。',
            before_data=before_data,
            after_data=snapshot_instance(quota),
        )
        return Response(ProfessorDoctorQuotaSerializer(quota).data, status=status.HTTP_200_OK)

    def delete(self, request, pk):
        quota = ProfessorDoctorQuota.objects.select_related('professor').filter(pk=pk).first()
        if not quota:
            return Response({'detail': '博士专业名额不存在。'}, status=status.HTTP_404_NOT_FOUND)
        professor = quota.professor
        before_data = snapshot_instance(quota)
        target_display = f'{quota.professor.name} - {quota.subject.subject_name}'
        quota.delete()
        refresh_professor_summary_quotas(professor)
        create_audit_log(
            request,
            action='doctor_quota.delete',
            module='导师博士专业名额',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='doctor_quota',
            target_id=pk,
            target_display=target_display,
            detail='删除导师博士专业名额。',
            before_data=before_data,
        )
        return Response({'detail': '博士专业名额已删除。'}, status=status.HTTP_200_OK)


class DashboardDoctorQuotaClearAllView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def post(self, request):
        quotas = list(ProfessorDoctorQuota.objects.select_related('professor', 'subject'))
        if not quotas:
            return Response({'detail': '当前没有可删除的博士专业名额记录。'}, status=status.HTTP_200_OK)

        touched_professors = {quota.professor for quota in quotas if quota.professor_id}
        deleted_count = len(quotas)
        with transaction.atomic():
            before_data = [
                snapshot_instance(quota, extra={'professor_name': quota.professor.name, 'subject_name': quota.subject.subject_name})
                for quota in quotas
            ]
            ProfessorDoctorQuota.objects.filter(id__in=[quota.id for quota in quotas]).delete()
            for professor in touched_professors:
                refresh_professor_summary_quotas(professor)
            create_audit_log(
                request,
                action='doctor_quota.clear_all',
                module='导师博士专业名额',
                level=DashboardAuditLog.LEVEL_WARNING,
                target_type='doctor_quota',
                target_display='全部博士专业名额',
                detail=f'一键删除全部导师博士专业名额，共 {deleted_count} 条。',
                before_data={'records': before_data[:100], 'count': deleted_count},
            )
        return Response({'detail': f'已删除全部博士专业名额记录，共 {deleted_count} 条。'}, status=status.HTTP_200_OK)


class DashboardMasterQuotaListView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def get(self, request):
        page = max(int(request.query_params.get('page', 1)), 1)
        page_size = min(max(int(request.query_params.get('page_size', 20)), 1), 200)
        queryset = ProfessorMasterQuota.objects.select_related('professor__department', 'subject')

        search = request.query_params.get('search', '').strip()
        professor_id = request.query_params.get('professor_id')
        subject_id = request.query_params.get('subject_id')
        department_id = request.query_params.get('department_id')
        subject_type = request.query_params.get('subject_type')
        order_by = request.query_params.get('order_by', 'teacher_identity_id')
        order_direction = request.query_params.get('order_direction', 'asc')

        if search:
            queryset = queryset.filter(
                Q(professor__name__icontains=search) |
                Q(professor__teacher_identity_id__icontains=search) |
                Q(subject__subject_name__icontains=search) |
                Q(subject__subject_code__icontains=search)
            )
        if professor_id:
            queryset = queryset.filter(professor_id=professor_id)
        if subject_id:
            queryset = queryset.filter(subject_id=subject_id)
        if department_id:
            queryset = queryset.filter(professor__department_id=department_id)
        if subject_type not in (None, ''):
            queryset = queryset.filter(subject__subject_type=subject_type)

        order_map = {
            'professor_name': 'professor__name',
            'teacher_identity_id': 'professor__teacher_identity_id',
            'subject_name': 'subject__subject_name',
            'subject_code': 'subject__subject_code',
            'beijing_quota': 'beijing_quota',
            'beijing_remaining_quota': 'beijing_remaining_quota',
            'yantai_quota': 'yantai_quota',
            'yantai_remaining_quota': 'yantai_remaining_quota',
            'total_quota': 'total_quota',
        }
        order_field = order_map.get(order_by, 'professor__teacher_identity_id')
        if order_direction == 'desc':
            order_field = f'-{order_field}'
        queryset = queryset.order_by(order_field, 'id')

        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        serializer = ProfessorMasterQuotaSerializer(queryset[start:end], many=True)
        return Response({'count': total, 'page': page, 'page_size': page_size, 'results': serializer.data}, status=status.HTTP_200_OK)

    def post(self, request):
        try:
            quota = save_master_quota_from_payload(request.data)
        except ValidationError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        create_audit_log(
            request,
            action='master_quota.create',
            module='导师硕士专业名额',
            target_type='master_quota',
            target_id=quota.pk,
            target_display=f'{quota.professor.name} - {quota.subject.subject_name}',
            detail='创建导师硕士专业名额。',
            after_data=snapshot_instance(quota),
        )
        return Response(ProfessorMasterQuotaSerializer(quota).data, status=status.HTTP_201_CREATED)


class DashboardMasterQuotaDetailView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def patch(self, request, pk):
        quota = ProfessorMasterQuota.objects.filter(pk=pk).first()
        if not quota:
            return Response({'detail': '硕士专业名额不存在。'}, status=status.HTTP_404_NOT_FOUND)
        before_data = snapshot_instance(quota)
        try:
            quota = save_master_quota_from_payload(request.data, quota=quota)
        except ValidationError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        create_audit_log(
            request,
            action='master_quota.update',
            module='导师硕士专业名额',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='master_quota',
            target_id=quota.pk,
            target_display=f'{quota.professor.name} - {quota.subject.subject_name}',
            detail='更新导师硕士专业名额。',
            before_data=before_data,
            after_data=snapshot_instance(quota),
        )
        return Response(ProfessorMasterQuotaSerializer(quota).data, status=status.HTTP_200_OK)

    def delete(self, request, pk):
        quota = ProfessorMasterQuota.objects.select_related('professor').filter(pk=pk).first()
        if not quota:
            return Response({'detail': '硕士专业名额不存在。'}, status=status.HTTP_404_NOT_FOUND)
        professor = quota.professor
        before_data = snapshot_instance(quota)
        target_display = f'{quota.professor.name} - {quota.subject.subject_name}'
        quota.delete()
        refresh_professor_summary_quotas(professor)
        create_audit_log(
            request,
            action='master_quota.delete',
            module='导师硕士专业名额',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='master_quota',
            target_id=pk,
            target_display=target_display,
            detail='删除导师硕士专业名额。',
            before_data=before_data,
        )
        return Response({'detail': '硕士专业名额已删除。'}, status=status.HTTP_200_OK)


class DashboardMasterQuotaClearAllView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def post(self, request):
        quotas = list(ProfessorMasterQuota.objects.select_related('professor', 'subject'))
        if not quotas:
            return Response({'detail': '当前没有可删除的硕士专业名额记录。'}, status=status.HTTP_200_OK)

        touched_professors = {quota.professor for quota in quotas if quota.professor_id}
        deleted_count = len(quotas)
        with transaction.atomic():
            before_data = [
                snapshot_instance(quota, extra={'professor_name': quota.professor.name, 'subject_name': quota.subject.subject_name})
                for quota in quotas
            ]
            ProfessorMasterQuota.objects.filter(id__in=[quota.id for quota in quotas]).delete()
            for professor in touched_professors:
                refresh_professor_summary_quotas(professor)
            create_audit_log(
                request,
                action='master_quota.clear_all',
                module='导师硕士专业名额',
                level=DashboardAuditLog.LEVEL_WARNING,
                target_type='master_quota',
                target_display='全部硕士专业名额',
                detail=f'一键删除全部导师硕士专业名额，共 {deleted_count} 条。',
                before_data={'records': before_data[:100], 'count': deleted_count},
            )
        return Response({'detail': f'已删除全部硕士专业名额记录，共 {deleted_count} 条。'}, status=status.HTTP_200_OK)


class DashboardSharedQuotaPoolListView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def get(self, request):
        page = max(int(request.query_params.get('page', 1)), 1)
        page_size = min(max(int(request.query_params.get('page_size', 20)), 1), 200)
        queryset = ProfessorSharedQuotaPool.objects.select_related('professor__department').prefetch_related('subjects')

        search = request.query_params.get('search', '').strip()
        professor_id = request.query_params.get('professor_id')
        department_id = request.query_params.get('department_id')
        quota_scope = request.query_params.get('quota_scope')
        campus = request.query_params.get('campus')
        order_by = request.query_params.get('order_by', 'teacher_identity_id')
        order_direction = request.query_params.get('order_direction', 'asc')

        if search:
            queryset = queryset.filter(
                Q(pool_name__icontains=search)
                | Q(professor__name__icontains=search)
                | Q(professor__teacher_identity_id__icontains=search)
                | Q(subjects__subject_name__icontains=search)
                | Q(subjects__subject_code__icontains=search)
            ).distinct()
        if professor_id:
            queryset = queryset.filter(professor_id=professor_id)
        if department_id:
            queryset = queryset.filter(professor__department_id=department_id)
        if quota_scope:
            queryset = queryset.filter(quota_scope=quota_scope)
        if campus:
            queryset = queryset.filter(campus=campus)

        order_map = {
            'pool_name': 'pool_name',
            'professor_name': 'professor__name',
            'teacher_identity_id': 'professor__teacher_identity_id',
            'quota_scope': 'quota_scope',
            'campus': 'campus',
            'total_quota': 'total_quota',
            'used_quota': 'used_quota',
            'remaining_quota': 'remaining_quota',
        }
        order_field = order_map.get(order_by, 'professor__teacher_identity_id')
        if order_direction == 'desc':
            order_field = f'-{order_field}'
        queryset = queryset.order_by(order_field, 'id')

        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        serializer = SharedQuotaPoolSerializer(queryset[start:end], many=True)
        return Response({'count': total, 'page': page, 'page_size': page_size, 'results': serializer.data}, status=status.HTTP_200_OK)

    def post(self, request):
        try:
            pool = save_shared_quota_pool_from_payload(request.data)
        except ValidationError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        create_audit_log(
            request,
            action='shared_quota_pool.create',
            module='共享名额池',
            target_type='shared_quota_pool',
            target_id=pool.pk,
            target_display=pool.pool_name,
            detail='创建共享名额池。',
            after_data=snapshot_instance(pool, extra={'subject_ids': list(pool.subjects.values_list('id', flat=True))}),
        )
        return Response(SharedQuotaPoolSerializer(pool).data, status=status.HTTP_201_CREATED)


class DashboardSharedQuotaPoolDetailView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def patch(self, request, pk):
        pool = ProfessorSharedQuotaPool.objects.prefetch_related('subjects').filter(pk=pk).first()
        if not pool:
            return Response({'detail': '共享名额池不存在。'}, status=status.HTTP_404_NOT_FOUND)
        before_data = snapshot_instance(pool, extra={'subject_ids': list(pool.subjects.values_list('id', flat=True))})
        try:
            pool = save_shared_quota_pool_from_payload(request.data, pool=pool)
        except ValidationError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        create_audit_log(
            request,
            action='shared_quota_pool.update',
            module='共享名额池',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='shared_quota_pool',
            target_id=pool.pk,
            target_display=pool.pool_name,
            detail='更新共享名额池。',
            before_data=before_data,
            after_data=snapshot_instance(pool, extra={'subject_ids': list(pool.subjects.values_list('id', flat=True))}),
        )
        return Response(SharedQuotaPoolSerializer(pool).data, status=status.HTTP_200_OK)

    def delete(self, request, pk):
        pool = ProfessorSharedQuotaPool.objects.prefetch_related('subjects').select_related('professor').filter(pk=pk).first()
        if not pool:
            return Response({'detail': '共享名额池不存在。'}, status=status.HTTP_404_NOT_FOUND)
        if pool.used_quota > 0:
            return Response({'detail': '该共享名额池已有已用名额，不能直接删除。'}, status=status.HTTP_400_BAD_REQUEST)
        professor = pool.professor
        before_data = snapshot_instance(pool, extra={'subject_ids': list(pool.subjects.values_list('id', flat=True))})
        target_display = pool.pool_name
        pool.delete()
        recalculate_professor_quota_summary(professor)
        create_audit_log(
            request,
            action='shared_quota_pool.delete',
            module='共享名额池',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='shared_quota_pool',
            target_id=pk,
            target_display=target_display,
            detail='删除共享名额池。',
            before_data=before_data,
        )
        return Response({'detail': '共享名额池已删除。'}, status=status.HTTP_200_OK)


class DashboardSharedQuotaPoolClearAllView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def post(self, request):
        blocking_count = ProfessorSharedQuotaPool.objects.filter(used_quota__gt=0).count()
        if blocking_count:
            return Response(
                {'detail': f'当前有 {blocking_count} 条共享名额池已有已用名额，不能一键删除，请先处理已用记录。'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        pools = list(ProfessorSharedQuotaPool.objects.prefetch_related('subjects').select_related('professor'))
        if not pools:
            return Response({'detail': '当前没有可删除的共享名额池记录。'}, status=status.HTTP_200_OK)

        touched_professors = {pool.professor for pool in pools if pool.professor_id}
        deleted_count = len(pools)
        with transaction.atomic():
            before_data = [
                snapshot_instance(pool, extra={'subject_ids': list(pool.subjects.values_list('id', flat=True))})
                for pool in pools
            ]
            ProfessorSharedQuotaPool.objects.filter(id__in=[pool.id for pool in pools]).delete()
            for professor in touched_professors:
                recalculate_professor_quota_summary(professor)
            create_audit_log(
                request,
                action='shared_quota_pool.clear_all',
                module='共享名额池',
                level=DashboardAuditLog.LEVEL_WARNING,
                target_type='shared_quota_pool',
                target_display='全部共享名额池',
                detail=f'一键删除全部共享名额池，共 {deleted_count} 条。',
                before_data={'records': before_data[:100], 'count': deleted_count},
            )
        return Response({'detail': f'已删除全部共享名额池记录，共 {deleted_count} 条。'}, status=status.HTTP_200_OK)


class DashboardWeChatAccountListView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def get(self, request):
        page = max(int(request.query_params.get('page', 1)), 1)
        page_size = min(max(int(request.query_params.get('page_size', 20)), 1), 200)
        queryset = WeChatAccount.objects.select_related('user')

        search = request.query_params.get('search', '').strip()
        order_by = request.query_params.get('order_by', 'username')
        order_direction = request.query_params.get('order_direction', 'asc')
        if search:
            queryset = queryset.filter(
                Q(user__username__icontains=search) |
                Q(openid__icontains=search) |
                Q(user__professor__name__icontains=search) |
                Q(user__student__name__icontains=search)
            ).distinct()

        order_map = {
            'username': 'user__username',
            'openid': 'openid',
        }
        order_field = order_map.get(order_by, 'user__username')
        if order_direction == 'desc':
            order_field = f'-{order_field}'
        queryset = queryset.order_by(order_field, 'id')

        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        serializer = WeChatAccountSerializer(queryset[start:end], many=True)
        return Response({'count': total, 'page': page, 'page_size': page_size, 'results': serializer.data}, status=status.HTTP_200_OK)

    def post(self, request):
        try:
            account = save_wechat_account_from_payload(request.data)
        except ValidationError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        create_audit_log(
            request,
            action='wechat_account.create',
            module='微信账号绑定',
            target_type='wechat_account',
            target_id=account.pk,
            target_display=account.openid,
            detail='创建微信账号绑定。',
            after_data=snapshot_instance(account),
        )
        return Response(WeChatAccountSerializer(account).data, status=status.HTTP_201_CREATED)


class DashboardWeChatAccountDetailView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def patch(self, request, pk):
        account = WeChatAccount.objects.filter(pk=pk).first()
        if not account:
            return Response({'detail': '微信绑定记录不存在。'}, status=status.HTTP_404_NOT_FOUND)
        before_data = snapshot_instance(account)
        try:
            account = save_wechat_account_from_payload(request.data, account=account)
        except ValidationError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        create_audit_log(
            request,
            action='wechat_account.update',
            module='微信账号绑定',
            target_type='wechat_account',
            target_id=account.pk,
            target_display=account.openid,
            detail='更新微信账号绑定。',
            before_data=before_data,
            after_data=snapshot_instance(account),
        )
        return Response(WeChatAccountSerializer(account).data, status=status.HTTP_200_OK)

    def delete(self, request, pk):
        account = WeChatAccount.objects.filter(pk=pk).first()
        if not account:
            return Response({'detail': '微信绑定记录不存在。'}, status=status.HTTP_404_NOT_FOUND)
        before_data = snapshot_instance(account)
        target_display = account.openid
        account.delete()
        create_audit_log(
            request,
            action='wechat_account.delete',
            module='微信账号绑定',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='wechat_account',
            target_id=pk,
            target_display=target_display,
            detail='删除微信账号绑定。',
            before_data=before_data,
        )
        return Response({'detail': '微信绑定记录已删除。'}, status=status.HTTP_200_OK)


class DashboardTokenListView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def get(self, request):
        page = max(int(request.query_params.get('page', 1)), 1)
        page_size = min(max(int(request.query_params.get('page_size', 20)), 1), 200)
        queryset = Token.objects.select_related('user')

        search = request.query_params.get('search', '').strip()
        user_type = request.query_params.get('user_type')
        order_by = request.query_params.get('order_by', 'created')
        order_direction = request.query_params.get('order_direction', 'desc')
        if search:
            queryset = queryset.filter(
                Q(user__username__icontains=search) |
                Q(user__professor__name__icontains=search) |
                Q(user__student__name__icontains=search) |
                Q(key__icontains=search)
            ).distinct()
        if user_type == 'professor':
            queryset = queryset.filter(user__professor__isnull=False)
        elif user_type == 'student':
            queryset = queryset.filter(user__student__isnull=False)
        elif user_type == 'staff':
            queryset = queryset.filter(user__is_staff=True, user__professor__isnull=True, user__student__isnull=True)

        order_map = {
            'username': 'user__username',
            'created': 'created',
            'key': 'key',
        }
        order_field = order_map.get(order_by, 'created')
        if order_direction == 'desc':
            order_field = f'-{order_field}'
        queryset = queryset.order_by(order_field, 'key')

        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        serializer = DashboardTokenSerializer(queryset[start:end], many=True)
        return Response({'count': total, 'page': page, 'page_size': page_size, 'results': serializer.data}, status=status.HTTP_200_OK)

    def post(self, request):
        user_id = request.data.get('user_id')
        username = str(request.data.get('username') or '').strip()
        if not user_id and not username:
            return Response({'detail': '用户不能为空。'}, status=status.HTTP_400_BAD_REQUEST)
        user = User.objects.filter(pk=user_id).first() if user_id else User.objects.filter(username=username).first()
        if not user:
            return Response({'detail': '用户不存在。'}, status=status.HTTP_404_NOT_FOUND)
        Token.objects.filter(user=user).delete()
        token = Token.objects.create(user=user)
        create_audit_log(
            request,
            action='token.create',
            module='认证令牌',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='token',
            target_id=token.key,
            target_display=get_display_name_for_user(user),
            detail='创建认证令牌。',
            after_data={'user_id': user.pk, 'token': token.key},
        )
        return Response(DashboardTokenSerializer(token).data, status=status.HTTP_201_CREATED)


class DashboardTokenDetailView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def delete(self, request, key):
        token = Token.objects.filter(key=key).first()
        if not token:
            return Response({'detail': '认证令牌不存在。'}, status=status.HTTP_404_NOT_FOUND)
        before_data = {'key': token.key, 'user_id': token.user_id}
        target_display = token.user.username if token.user_id else key
        token.delete()
        create_audit_log(
            request,
            action='token.delete',
            module='认证令牌',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='token',
            target_id=key,
            target_display=target_display,
            detail='删除认证令牌。',
            before_data=before_data,
        )
        return Response({'detail': '认证令牌已删除。'}, status=status.HTTP_200_OK)


class DashboardTokenRegenerateView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def post(self, request, key):
        token = Token.objects.select_related('user').filter(key=key).first()
        if not token:
            return Response({'detail': '认证令牌不存在。'}, status=status.HTTP_404_NOT_FOUND)
        user = token.user
        token.delete()
        new_token = Token.objects.create(user=user)
        create_audit_log(
            request,
            action='token.regenerate',
            module='认证令牌',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='token',
            target_id=new_token.key,
            target_display=get_display_name_for_user(user),
            detail='重新生成认证令牌。',
            before_data={'old_key': key},
            after_data={'new_key': new_token.key, 'user_id': user.pk},
        )
        return Response(DashboardTokenSerializer(new_token).data, status=status.HTTP_200_OK)


class DashboardAlternateListView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def get(self, request):
        page = max(int(request.query_params.get('page', 1)), 1)
        page_size = min(max(int(request.query_params.get('page_size', 20)), 1), 100)
        available_admission_years = list(
            Student.objects.filter(is_alternate=True)
            .exclude(admission_year__isnull=True)
            .order_by('-admission_year')
            .values_list('admission_year', flat=True)
            .distinct()
        )
        queryset = build_alternate_queryset(request).select_related('subject')
        order_by = request.query_params.get('order_by', 'alternate_rank')
        order_direction = request.query_params.get('order_direction', 'asc')

        order_map = {
            'name': 'name',
            'candidate_number': 'candidate_number',
            'subject_name': 'subject__subject_name',
            'admission_year': 'admission_year',
            'final_rank': 'final_rank',
            'alternate_rank': 'alternate_rank',
            'is_giveup': 'is_giveup',
        }
        order_field = order_map.get(order_by, 'alternate_rank')
        if order_direction == 'desc':
            order_field = f'-{order_field}'
        queryset = queryset.order_by(order_field, 'id')

        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        serializer = AlternateStudentSerializer(queryset[start:end], many=True)
        return Response(
            {
                'count': total,
                'page': page,
                'page_size': page_size,
                'results': serializer.data,
                'available_admission_years': available_admission_years,
            },
            status=status.HTTP_200_OK,
        )


class DashboardAlternatePromoteNextView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def post(self, request):
        subject_id = request.data.get('subject_id')
        if not subject_id:
            return Response({'detail': 'subject_id is required.'}, status=status.HTTP_400_BAD_REQUEST)

        subject = Subject.objects.filter(pk=subject_id).first()
        if not subject:
            return Response({'detail': 'Subject not found.'}, status=status.HTTP_404_NOT_FOUND)

        admission_year = request.data.get('admission_year')
        promoted_student, reason = promote_next_alternate(
            subject,
            require_available_quota=True,
            admission_year=to_int(admission_year, default=None) if admission_year not in (None, '') else None,
        )
        if not promoted_student:
            if reason == 'no_quota':
                return Response({'detail': '该专业当前没有可回补的可用名额，无法递补候补学生。'}, status=status.HTTP_409_CONFLICT)
            return Response({'detail': '该专业没有可递补的候补学生。'}, status=status.HTTP_400_BAD_REQUEST)

        create_audit_log(
            request,
            action='alternate.promote_next',
            module='候补管理',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='subject',
            target_id=subject.pk,
            target_display=subject.subject_name,
            detail=f'递补候补学生 {promoted_student.name}。',
            after_data={'student_id': promoted_student.id, 'student_name': promoted_student.name},
        )
        return Response(
            {'detail': f'已递补候补学生 {promoted_student.name}。', 'student_id': promoted_student.id, 'student_name': promoted_student.name},
            status=status.HTTP_200_OK,
        )


class DashboardAlternateBatchDeleteView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def post(self, request):
        ids, delete_all_filtered = parse_batch_delete_request(request)
        queryset = build_alternate_queryset(request).select_related('user_name')
        if not delete_all_filtered:
            queryset = queryset.filter(id__in=ids)
        if not queryset.exists():
            return Response({'detail': '没有可删除的候补记录。'}, status=status.HTTP_400_BAD_REQUEST)
        users = [student.user_name for student in queryset if student.user_name_id]
        names = list(queryset.values_list('name', flat=True))
        deleted_count = queryset.count()
        queryset.delete()
        for user in users:
            User.objects.filter(pk=user.pk).delete()
        create_audit_log(
            request,
            action='alternate.batch_delete',
            module='候补管理',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='student',
            target_display='批量删除候补记录',
            detail=f'删除 {deleted_count} 条候补记录。',
            before_data={'names': names, 'delete_all_filtered': delete_all_filtered},
        )
        return Response({'detail': f'已删除 {deleted_count} 条候补记录。'}, status=status.HTTP_200_OK)


class DashboardGiveupListView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def get(self, request):
        page = max(int(request.query_params.get('page', 1)), 1)
        page_size = min(max(int(request.query_params.get('page_size', 20)), 1), 100)
        queryset = build_giveup_queryset(request).select_related('subject')
        order_by = request.query_params.get('order_by', 'id')
        order_direction = request.query_params.get('order_direction', 'desc')

        order_map = {
            'name': 'name',
            'candidate_number': 'candidate_number',
            'subject_name': 'subject__subject_name',
            'final_rank': 'final_rank',
            'is_selected': 'is_selected',
        }
        order_field = order_map.get(order_by, 'id')
        if order_direction == 'desc':
            order_field = f'-{order_field}'
        queryset = queryset.order_by(order_field, '-id')

        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        serializer = GiveupStudentSerializer(queryset[start:end], many=True)
        return Response(
            {'count': total, 'page': page, 'page_size': page_size, 'results': serializer.data},
            status=status.HTTP_200_OK,
        )


class DashboardGiveupBatchDeleteView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsDashboardAdmin]

    def post(self, request):
        ids, delete_all_filtered = parse_batch_delete_request(request)
        queryset = build_giveup_queryset(request).select_related('user_name')
        if not delete_all_filtered:
            queryset = queryset.filter(id__in=ids)
        if not queryset.exists():
            return Response({'detail': '没有可删除的放弃录取记录。'}, status=status.HTTP_400_BAD_REQUEST)
        users = [student.user_name for student in queryset if student.user_name_id]
        names = list(queryset.values_list('name', flat=True))
        deleted_count = queryset.count()
        queryset.delete()
        for user in users:
            User.objects.filter(pk=user.pk).delete()
        create_audit_log(
            request,
            action='giveup.batch_delete',
            module='放弃录取',
            level=DashboardAuditLog.LEVEL_WARNING,
            target_type='student',
            target_display='批量删除放弃录取记录',
            detail=f'删除 {deleted_count} 条放弃录取记录。',
            before_data={'names': names, 'delete_all_filtered': delete_all_filtered},
        )
        return Response({'detail': f'已删除 {deleted_count} 条放弃录取记录。'}, status=status.HTTP_200_OK)

